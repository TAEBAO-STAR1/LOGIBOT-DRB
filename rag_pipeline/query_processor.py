import os
import time
import json
import logging
import hashlib
import re
import smtplib
import openai
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
from pydantic import PrivateAttr
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.language_models.llms import LLM
from langchain_qdrant import Qdrant
from langchain_ollama import OllamaEmbeddings 
from qdrant_client import QdrantClient, models
from qdrant_client.models import Distance, VectorParams, PointStruct, Filter, FieldCondition, MatchValue, MatchAny
from ddgs import DDGS
from functools import lru_cache
import threading
import pandas as pd
import fitz

load_dotenv()
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger("httpx").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# 환경 변수
QDRANT_COLLECTION = os.getenv("QDRANT_COLLECTION", "logistics_data")
QDRANT_HOST = os.getenv("QDRANT_HOST", "localhost")
QDRANT_PORT = os.getenv("QDRANT_PORT", "6333")
QDRANT_API_KEY = os.getenv("QDRANT_API_KEY", None)
LEARNING_COLLECTION = os.getenv("LEARNING_COLLECTION", "learning_history")
BAD_FEEDBACK_COLLECTION = os.getenv("BAD_FEEDBACK_COLLECTION", "bad_feedback_history")

# 이메일 설정
EMAIL_ENABLED = os.getenv("EMAIL_ENABLED", "false").lower() == "true"
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
EMAIL_FROM = os.getenv("EMAIL_FROM", SMTP_USERNAME)
EMAIL_TO = os.getenv("EMAIL_TO", "").split(",")

# 로깅 전용 컬렉션
QUERY_LOG_COLLECTION = "query_logs"
ANSWER_LOG_COLLECTION = "answer_logs"

ONPREMISE_API_URL = os.getenv("ONPREMISE_API_URL", "http://192.168.1.128:9800")
ONPREMISE_MODEL   = os.getenv("ONPREMISE_MODEL",   "gpt-oss-120b")
ONPREMISE_API_KEY = os.getenv("ONPREMISE_API_KEY",  "")
ONPREMISE_TIMEOUT = int(os.getenv("ONPREMISE_TIMEOUT", "120"))
OLLAMA_EMBEDDING_MODEL = os.getenv("OLLAMA_EMBEDDING_MODEL", "granite-embedding:278m")
OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://localhost:11434")

# 캐시 설정
CACHE_TTL = 600  # 10분 — 같은 질문 반복 시 일관된 답변 보장
response_cache = {}
cache_lock = threading.Lock()

# ══════════════════════════════════════════════════════════════
# 시뮬레이터 유도 안내
# ══════════════════════════════════════════════════════════════
SIMULATOR_GUIDE = {
    "컨베어벨트배차": {
        "keywords": [
            # 직경 계산
            "직경 계산", "롤 직경", "직경을 계산", "직경이 얼마",
            "직경 얼마", "직경 구해", "직경 알려",
            "롤 계산", "1롤일때", "롤일때",
            "컨베어 계산", "컨베이어 계산",
            "m 1롤", "m일때 직경", "길이 직경",
            # 컨베어 + 차량/배차 조합
            "컨베어벨트 배차", "컨베이어 배차", "컨베어 배차",
            "컨베어벨트배차", "컨베이어배차",
            "컨베어 차량", "컨베이어 차량", "컨베어벨트 차량",
            "컨베어 몇 대", "컨베이어 몇 대", "컨베어 몇대",
            "컨베어 차 몇", "컨베어 트럭", "컨베어 운송",
            # ROLL 수 + 배차/차량
            "ROLL 배차", "roll 배차", "롤 배차",
            "ROLL 차량", "roll 차량", "롤 차량",
            "ROLL 몇 대", "roll 몇 대", "롤 몇 대",
            "롤 몇대", "ROLL 몇대",
            # 컨베어 + 톤/중량 조합
            "컨베어 톤", "컨베이어 톤", "컨베어벨트 톤",
            "컨베어 중량", "컨베이어 중량",
            # 컨베어 맥락에서 배차/차량 수 질문 패턴
            "배차해야", "배차 해야", "배차할까", "배차 할까",
            "몇 대를", "몇대를", "몇 대야", "몇대야",
            "톤 배차", "ton 배차", "ton 차량",
            "컨베어 8톤", "컨베어 9톤", "컨베어 11톤", "컨베어 18톤", "컨베어 25톤",
            "컨베이어 8톤", "컨베이어 9톤", "컨베이어 11톤",
            "톤 차량 몇", "톤 차량 몇 대", "t 차량 몇",
            "컨베어 차량 몇", "컨베이어 차량 몇",
            "롤인데 차량", "롤이고 차량", "롤 차량 배차",
            "배차해줘", "배차해 줘", "배차 부탁",
        ],
        "msg": (
            "💡 **해당 내용은 시뮬레이터를 이용해주세요!**\n\n좌측 메뉴의 **[시뮬레이터]** 를 클릭하시면 정확한 계산 결과를 확인하실 수 있습니다."
        ),
    },
    "국내운임비교": {
        "keywords": [
            # 운임 직접 언급
            "운임 비교", "운임비교", "운송비 얼마", "택배비 얼마",
            "화물 요금", "배송비 얼마", "톤수별 운임", "운임 계산",
            # 지역 + 운송/배차 조합
            "부산에서", "에서 광주", "에서 서울", "에서 대구", "에서 인천",
            "에서 수원", "에서 대전", "에서 울산", "에서 창원",
            # 중량 + 운송/배차/방법 조합
            "kg 운송", "ton 운송", "톤 운송",
            "kg 배차", "ton 배차",
            "최적의 배차", "최적 배차", "배차를 알려", "배차 알려줘",
            "운송 방법", "운반 방법", "화물 어떻게 보내", "택배 어떻게 보내",
            "직송이야", "화물이야", "택배야", "화물로 보내",

        ],
        "msg": (
            "💡 **해당 내용은 시뮬레이터를 이용해주세요!**\n\n좌측 메뉴의 **[시뮬레이터]** 를 클릭하시면 정확한 계산 결과를 확인하실 수 있습니다."
        ),
    },
    "수출포장량": {
        "keywords": [
            "박스 몇 개", "몇 박스", "수출 포장 몇",
            "650 몇개", "1090 몇개", "마대 몇개", "박스 수량 계산",
            "몇 박스 나와", "몇개 나와",
            # 수출 포장량 계산 요청 패턴
            "수출포장량", "수출 포장량", "포장량 산출", "포장량 계산",
            "자재그룹별", "박스로 포장", "kg 포장", "포장할때",
            "포장 산출", "박스 산출", "마대 산출",
            "수출 박스", "650박스", "1090박스",
        ],
        "msg": (
            "💡 **해당 내용은 시뮬레이터를 이용해주세요!**\n\n좌측 메뉴의 **[시뮬레이터]** 를 클릭하시면 정확한 계산 결과를 확인하실 수 있습니다."
        ),
    },
    "국내최적배차": {
        "keywords": [
            "크롤러 배차", "러버트랙 배차", "RT 배차",
            "크롤러 몇 대", "크롤러 차량 몇", "배차 계산",
            "차량 몇 대 필요", "크롤러 몇 톤", "러버트랙 몇 대",
            # 자재코드 + 수량 + 배차 조합 질문
            "중량과 포장단위", "포장단위", "출고를 위한 배차", "출고 배차",
            "개의 중량", "개 배차", "개 차량", "수량 배차",
            "수량과 배차", "수량 차량", "몇개 배차", "몇 개 배차",
            "포장단위 배차", "중량 배차 차량", "출고 차량",
        ],
        "msg": (
            "💡 **해당 내용은 시뮬레이터를 이용해주세요!**\n\n좌측 메뉴의 **[시뮬레이터]** 를 클릭하시면 정확한 계산 결과를 확인하실 수 있습니다."
        ),
    },
}

def check_simulator_intent(query: str) -> str | None:
    """시뮬레이터 처리 가능 질문이면 유도 메시지 반환, 아니면 None"""
    import re as _re
    q = query.strip()

    # ── 1) 자재코드 + 배차/차량 맥락 → 시뮬레이터 자동 유도 (최우선) ──
    # 키워드보다 먼저 체크: 자재코드가 있으면 제품 유형으로 정확하게 유도
    # 자재코드 패턴: 7자리 숫자
    _code_pat = _re.search(r"(?<!\d)(\d{7})(?!\d)", q)
    if _code_pat:
        _code = int(_code_pat.group(1))
        # 배차/차량 맥락 키워드
        _dispatch_kw = [
            "배차", "차량", "몇 톤", "몇톤", "출고", "운송",
            "적재", "싣", "몇 대", "배달", "보내",
        ]
        _is_dispatch = any(k in q for k in _dispatch_kw)

        if _is_dispatch:
            # 자재코드 범위로 시뮬레이터 결정
            # 컨베어벨트 코드: B0x로 시작하는 7자리 (예: 1093605, 9000xxx)
            # 크롤러 코드: 6xxxxxxx (예: 6004216, 6005489)
            # 수출 코드: B01/B02/N18/N19 시작
            if 6000000 <= _code <= 6999999:
                return (
                    "💡 **해당 내용은 시뮬레이터를 이용해주세요!**\n\n좌측 메뉴의 **[시뮬레이터]** 를 클릭하시면 정확한 계산 결과를 확인하실 수 있습니다."
                )
            elif 1000000 <= _code <= 1999999 or 9000000 <= _code <= 9099999:
                return (
                    "💡 **해당 내용은 시뮬레이터를 이용해주세요!**\n\n좌측 메뉴의 **[시뮬레이터]** 를 클릭하시면 정확한 계산 결과를 확인하실 수 있습니다."
                )
            else:
                # 코드 범위 불명확 → 일반 배차 시뮬레이터 유도
                return (
                    "💡 **해당 내용은 시뮬레이터를 이용해주세요!**\n\n좌측 메뉴의 **[시뮬레이터]** 를 클릭하시면 정확한 계산 결과를 확인하실 수 있습니다."
                )

    # ── 2) 키워드 기반 시뮬레이터 유도 (자재코드 없을 때) ────────────
    for sim_name, info in SIMULATOR_GUIDE.items():
        if any(kw in q for kw in info["keywords"]):
            return info["msg"]

    return None

# ══════════════════════════════════════════════════════════════
# 격주 기사 노선 기준
# ══════════════════════════════════════════════════════════════
# 5/28(목)이 속한 주 월요일 = 5/26 → B주(이용구 노선 운행주)
BIWEEKLY_ANCHOR_DATE = date(2026, 5, 25)  # 5/28 적용 시작주의 월요일 (B주=이용구 기준)

def get_week_group(target: date = None) -> str:
    """오늘(또는 target)이 A주인지 B주인지 반환. B주=이용구, A주=심효섭"""
    if target is None:
        target = date.today()
    monday = target - timedelta(days=target.weekday())
    weeks_elapsed = (monday - BIWEEKLY_ANCHOR_DATE).days // 7
    return "B" if weeks_elapsed % 2 == 0 else "A"

# ══════════════════════════════════════════════════════════════
# 프롬프트 시스템 — 도메인별 분리 구조
# ══════════════════════════════════════════════════════════════

# ── 통합 프롬프트 (V4 단일 프롬프트 방식 복원 + V5 도메인 분기 유지) ──────
# V4와 동일하게 모든 규칙을 하나로 통합 → 도메인 분류 오류 시에도 규칙 누락 없음
_PROMPT_SYSTEM = """당신은 DRB 동일고무벨트 물류팀 전문 AI 어시스턴트입니다.
아래 [참고 데이터]만을 근거로 답변하세요.
참고 데이터에 없는 내용은 "해당 정보를 찾을 수 없습니다"라고 말하세요.

[답변 작성 규칙]
1. 출처 표기 금지: "문서 1", "[문서 2]" 같은 내부 참조 번호를 절대 포함하지 마세요.
2. 데이터 외 정보 생성 금지: [참고 데이터]에 없는 단어·수치·분류를 임의로 만들거나 조합하지 마세요.
3. LaTeX 수식 금지: 계산식은 반드시 자연어로만 작성하세요.
   올바른 예시: 포의 총두께 = (6.4 + 2.4) + (3 × 0.95 - 0.2) = 11.65 mm
4. 핵심 먼저: 불필요한 서론·인사 없이 핵심 답변부터 바로 시작하세요.
5. 간결하게: 질문에 직접 관련 없는 부가 설명·추측은 포함하지 마세요.
6. 정보 통합: 여러 데이터에 걸친 정보는 중복 없이 하나로 합쳐 서술하세요.
7. 비교·목록은 표로: 여러 항목을 나열하거나 비교할 때는 Markdown 표를 사용하세요.
8. 없는 정보: 참고 데이터에 없는 내용은 "해당 정보를 찾을 수 없습니다"라고 말하세요.
9. 배차 계산: '몇 톤', '배차', '차량' 포함 질문 →
   ① 자재코드로 1파렛트당 최대 적재수량 확인
   ② 총 수량 ÷ 1파렛트당 수량 = 필요 파렛트 수
   ③ 파렛트 사이즈와 차량 적재함 폭·길이 비교
   ④ 조건을 충족하는 가장 작은 차량 추천
10. 컨베어벨트 직경 계산: '직경', '롤 직경' 포함 질문 →
    ① 자재코드로 상고무두께·하고무두께·PLY·코팅후포두께 확인
    ② 포의 총두께 = (상고무두께 + 하고무두께) + (PLY × 코팅후포두께 - 0.2)
    ③ 롤 직경(m) = √(포의 총두께 ÷ 1000 × 4 × 길이(M) ÷ 3.142 + 0.09)
    ④ 길이 미입력 시 ①②만 계산 후 "컨베어벨트 길이(M)를 알려주세요" 요청
    ⑤ 최종 롤 직경(m) + mm 단위 환산값 함께 제시
11. 담당자/인원 조회:
    [클레임·문의처 질문 — '클레임','불량','문제','누구한테','연락해야','어디에' 포함 시]
    ① 해당 공정 주임 1명만 선택
    ② 해당 공정 담당 팀원 1명만 선택
    ③ 그 외 인원 절대 포함 금지
    ④ 답변 형식: "○○ 관련 문의는 아래 담당자에게 연락해 주세요.\n- 홍길동 주임 (내선: 1234)"
    [일반 인원 조회 — '몇 명','전체','현황','팀원' 포함 시]
    표 형식: | 성명 | 직책 | 담당공정 | 내선 |
    전화번호 0이면 "직통번호 없음 (내선 문의)"
    [업무 영역 매핑]
    국내/내수 → 담당공정에 "내수" | 수출/해외 → "수출" | 컨베어/크롤러 → "컨베어" 또는 "크롤러"
    중부 → "중부" | 베트남 → "베트남" | 지입기사 → 부산공장/중부물류센터 기사
12. 지입기사 납품 동선: '지입기사','납품 동선','노선','동선' 포함 질문 →
    월~금 행이 모두 있는 기사 = 주5일 매일 운행 (특정 요일만 운행이라고 절대 말하지 마세요)
    기사별로 요일별 납품 동선 표로 정리: | 요일 | 납품 동선 |
13. 자재코드(7자리 숫자) 단독 질문 →
    ① 자재 종류 자동 식별 ② 자재내역·자재그룹·중량 등 기본정보 제시
    ③ 추가 키워드('중량','배차','직경') 있으면 해당 계산도 수행
14. 전동수출 파렛트 CBM: 파렛트 1개 = 1.1m × 1.1m × 2.2m = 2.662 CBM, N파렛트 = 2.662 × N
    내수용 PLT = 전동내수 공정 파렛트, 수출용 PLT = 전동수출 공정 파렛트
    PLT/박스 리스트 질문: 해당 공정의 모든 규격을 빠짐없이 표로 제시
    "내수용"="전동내수", "수출용"="전동수출"로 매핑하여 조회
15. 국내 출고 운송방식: 부산시내 150kg/경남권 300kg/장거리 800kg 기준으로 화물·택배 vs 직송 판단
16. 박스 적재: 600박스=1PLT당 8박스, 650박스=20박스, 1090박스=4박스
17. 수출 컨테이너: 혼합 조합(40ft+20ft)도 제시, 잔여 공간 최소 조합 추천"""

# ── 도메인별 추가 규칙 (V4 단일 프롬프트에 추가로 domain별 세부 지침만 남김) ──
_DOMAIN_RULES = {

    "conveyor": """
[컨베어벨트 전용 규칙]
1. 직경 계산 순서:
   ① 자재코드로 상고무두께·하고무두께·PLY·코팅후포두께 확인
   ② 포의 총두께 = (상고무두께 + 하고무두께) + (PLY × 코팅후포두께 - 0.2)
   ③ 롤 직경(m) = √(포의 총두께 ÷ 1000 × 4 × 길이(M) ÷ 3.142 + 0.09)
   ④ 길이 미입력 시 ①②만 계산 후 "컨베어벨트 길이(M)를 알려주세요" 요청
   ⑤ 최종 답변: 롤 직경(m) + mm 단위 환산값 함께 제시
2. 공백 셀 처리: 같은 포규격 패턴을 참고해 계산
3. 자재내역 해석: ME NN-200 800X3X6.0X2.0 → 폭800mm PLY3 상고무6.0mm 하고무2.0mm
4. 표 활용: 계산 과정은 단계별로, 결과는 표로 정리
5. 컨베어벨트 배차 질문 (ROLL수 + 중량 제시 시):
   ① 총중량(ton) 기준으로 [차량 데이터]의 최대중량(ton) 범위와 비교
   ② 총중량이 차량 1대 최대중량 이하 → 해당 차량 1대 추천
   ③ 초과 시 분할 배차 계산: 몇 대 필요한지 올림 계산
   ④ 결과 표: | 총중량 | 추천차량 | 필요대수 | 비고 |
   ⑤ 로브이(Low-bed)는 높이 2.6m 이상 제품에만 적용""",

    "crawler": """
[크롤러 러버트랙 전용 규칙]
1. 배차 계산 순서:
   ① 자재코드로 1PC당중량·1파렛트최대적재수 확인
   ② 필요파렛트수 = 총수량 ÷ 1파렛트최대적재수 (소수점 올림)
   ③ 파렛트사이즈(L×W)와 [차량 데이터] 적재함 폭·길이 비교
   ④ 조건 충족하는 가장 작은 차량 추천
2. 결과 표: | 자재코드 | 수량 | 1PC중량 | 총중량 | 필요PLT | 추천차량 |
3. 자재코드 해석: RT550X60LX90.0P_A5RD → RT(러버트랙) 550(폭) 60(링크수) 90(피치mm)""",

    "sidewall": """
[주름혔벨트(사이드월) 전용 규칙]
1. 우든박스 사이즈는 [주름혹벨트 우든박스 사이즈 데이터] 시트 참조
2. 자재코드로 자재내역·수량(M)·우든박스 규격(W×L×H) 확인 후 제시
3. '수량(M)' = 주문 길이(M) 기준값. 동일 자재코드에 규격이 여러 개면 수량(M)별로 모두 제시
4. 총중량 계산: (수량_M × 순중량_KG) + 우든박스_중량_KG
5. 결과 표: | 수량(M) | 우든박스 규격(W×L×H) | 우든박스 중량 추정 | 총중량 예상 |""",

    "export": """
[수출 포장량 전용 규칙]
1. 박스 종류별 1파렛트당 적재수량:
   600박스=8개/PLT, 650박스=20개/PLT, 1090박스=4개/PLT
2. 파렛트 규격:
   600박스: 1200×800×730mm (패키징: 1200×800×1460mm)
   650박스: 1100×1100×2200mm
   1090박스: 1100×1100×1110mm (패키징: 1100×1100×2220mm)
3. CBM 계산: 파렛트 1개 = 2.662 CBM
4. 컨테이너 선택: 20ft=최대10PLT, 40ft=최대20PLT, 혼합 조합도 제시""",

    "domestic": """
[국내 운송방식 전용 규칙]
1. 자재코드로 1PC당 중량 확인 → 총 중량 계산
2. 도착지 구간별 기준:
   부산시내: 150kg 이하 화물/택배, 초과 직송
   경남권(녹산·대저·명지): 300kg 이하 화물/택배, 초과 직송
   장거리(서울·광주·대구 등): 800kg 이하 화물/택배, 초과 직송
3. 웹 검색 절대 금지""",

    "driver_route": """
[지입기사 납품 동선 전용 규칙]
1. 월~금 행이 모두 있는 기사 = 주5일 매일 운행 ("특정 요일만"이라고 절대 말하지 마세요)
2. 요일별 도착지가 다른 것 = 그날의 납품 코스(동선)가 다름
3. 모든 기사(부산공장 + 중부물류센터) 빠짐없이 포함
4. 기사별 요일별 납품 동선 표: | 요일 | 납품 동선 |
5. 이번 주 운행 기사 판별: 격주 앵커(2026-05-25 = B주 = 이용구) 기준
6. 거래처 관련 질문 ("○○ 거래처 담당자", "○○에 납품 담당", "거래처 문의처" 등):
   → 납품 동선 데이터의 거래처명은 기사 노선 확인용이며,
     거래처 담당자·클레임·문의는 물류팀 담당자인 **김동우 팀원 (내선: 9133)**에게 문의 안내""",

    "personnel": """
[담당자 조회 전용 규칙]
1. 인원 수 질문("몇 명", "총 몇 명", "인원은"):
   → 데이터에서 직접 세어 "물류팀 총 ○명입니다" 형식으로 답변
   → 리스트를 그대로 나열하지 말고 숫자로 직접 답변
   → 세부 구성도 요청 시: "사무직 ○명, 현장직 ○명, 지입기사 ○명"
2. 전체 현황 요청("현황", "명단", "알려줘"): 표 형식 | 성명 | 직책 | 담당공정 | 내선 |
   전화번호 0이면 "직통번호 없음 (내선 문의)"
3. 직책자 질문: 팀장·주임·기정만 출력 (사원·팀원·지입기사 제외)
   ※ 단, 지게차/외주 기사 현황 질문은 예외 → 아래 규칙6 적용
4. 클레임/문의처: 해당 공정 주임 1명 + 팀원 1명만, 그 외 절대 포함 금지
   형식: "○○ 관련 문의는 아래 담당자에게 연락해 주세요.\n- 홍길동 주임 (내선: 1234)"
5. 지입기사 질문: 이름·연락처·차량 종류 제시
6. 지게차·외주 기사 현황 질문("지게차 기사", "지게차 기사님", "외주 기사"):
   → [참고 데이터]의 운영규칙에서 지게차 관련 Q&A를 찾아 그대로 답변
   → "지게차 기사님 총 3명 (4.5톤 3대, 10톤 1대)" 형식으로 제시
   → 데이터에 이름/연락처가 없으면 "담당자(신태환 팀원, 내선 9067)에게 문의" 안내""",

    "vehicle": """
[차량 제원 전용 규칙]
1. 차량 데이터 표 형식: | 차량톤수 | 최대중량(ton) | 적재함길이(m) | 적재함폭(m) |
2. 특이사항: 높이 2.6m 이상 → 로브이(Low-bed) 선택, 윙바디는 2.4m 미만
3. 최대중량 초과 적재 절대 금지""",

    "operation_rule": """
[운영 규칙 전용]
1. Q&A 데이터에서 질문과 가장 유사한 답변을 찾아 제시
   - 질문과 완전히 동일하지 않아도 됨. 의미상 유사한 Q&A를 적극 활용
   - 여러 Q&A에 걸친 내용이면 통합해서 하나의 답변으로 제시
2. 시간·마감 관련: 구체적인 시각까지 정확히 전달
3. 추가 안내 금지: 데이터에 없는 절차·주의사항을 임의로 덧붙이지 마세요
4. 데이터에 완전히 없는 내용만 "물류팀 담당자에게 문의해 주세요" 안내""",

    "general": """
[일반 규칙]
1. [참고 데이터]에 질문과 관련된 Q&A가 있으면 그 답변을 최우선으로 활용
2. 여러 데이터에 걸친 정보는 하나로 통합해서 자연스럽게 설명
3. 비교/목록 데이터는 Markdown 표로 정리
4. 수량 질문: 데이터를 빠짐없이 세어 정확한 합계 제시
5. 자재코드(7자리 숫자) 단독 입력: 자재 종류 자동 식별 후 기본정보 제시
6. [참고 데이터]에 관련 내용이 없으면 "해당 정보를 내부 데이터에서 찾을 수 없습니다" 안내""",
}

_PROMPT_BASE = """{system_role}

{domain_rules}

[이전 대화]
{conversation_context}

[과거 유사 답변]
{history_context}

[참고 데이터]
{context}

[질문]
{input}

답변:"""

# ── 도메인별 프롬프트 빌더 (캐시 없음 — 프롬프트 변경 즉시 반영) ──
def get_domain_prompt(domain: str) -> ChatPromptTemplate:
    """도메인별 ChatPromptTemplate 반환"""
    rule = _DOMAIN_RULES.get(domain, _DOMAIN_RULES["general"])
    template = (
        _PROMPT_BASE
        .replace("{system_role}", _PROMPT_SYSTEM)
        .replace("{domain_rules}", rule)
    )
    return ChatPromptTemplate.from_template(template)

# 기존 코드 호환용 (RAGChainWrapper 초기화 시 사용)
PROMPT_TEMPLATE = (
    _PROMPT_BASE
    .replace("{system_role}", _PROMPT_SYSTEM)
    .replace("{domain_rules}", _DOMAIN_RULES["general"])
)

class EmailNotifier:
    """부정 피드백 이메일 알림 시스템"""
    
    def __init__(self):
        self.enabled = EMAIL_ENABLED
        self.smtp_server = SMTP_SERVER
        self.smtp_port = SMTP_PORT
        self.username = SMTP_USERNAME
        self.password = SMTP_PASSWORD
        self.email_from = EMAIL_FROM
        self.email_to = [email.strip() for email in EMAIL_TO if email.strip()]
        
        if self.enabled:
            if not self.username or not self.password:
                logger.warning("⚠️ 이메일 알림이 활성화되었지만 SMTP 계정 정보가 없습니다.")
                self.enabled = False
            elif not self.email_to:
                logger.warning("⚠️ 이메일 알림이 활성화되었지만 수신자가 지정되지 않았습니다.")
                self.enabled = False
            else:
                logger.info(f"✅ 이메일 알림 활성화: {', '.join(self.email_to)}")
    
    def send_bad_feedback_notification(self, feedback_data: Dict):
        """부정 피드백 발생 시 이메일 전송"""
        if not self.enabled:
            return
        
        try:
            # 이메일 본문 구성
            html_content = self._create_html_email(feedback_data)
            
            # 이메일 메시지 생성
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f"[물류 AI 챗봇] 부정 피드백 알림 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            # 발신자 표시명만 보이고 실제 계정 주소는 숨김
            from email.utils import formataddr
            msg['From'] = formataddr(("DRB LOGIBOT-AI", self.email_from))
            msg['To'] = ', '.join(self.email_to)
            
            # HTML 파트 추가
            html_part = MIMEText(html_content, 'html', 'utf-8')
            msg.attach(html_part)
            
            # SMTP 서버 연결 및 전송
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.username, self.password)
                server.send_message(msg)
            
            logger.info(f"✅ 부정 피드백 이메일 전송 완료: {feedback_data['query'][:30]}...")
            
        except Exception as e:
            logger.error(f"❌ 이메일 전송 실패: {e}")

    def send_improvement_request(self, content: str, team: str):
        """개선 요청 이메일 전송"""
        if not self.enabled:
            return False
        try:
            from email.utils import formataddr
            ts = datetime.now().strftime("%Y년 %m월 %d일 %H:%M")
            html_content = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<style>
  body {{ font-family: 'Malgun Gothic', sans-serif; background:#f4f6f9; margin:0; padding:20px; }}
  .wrap {{ max-width:640px; margin:auto; background:#fff; border-radius:12px;
           box-shadow:0 4px 16px rgba(0,0,0,.1); overflow:hidden; }}
  .header {{ background:linear-gradient(135deg,#667eea,#764ba2); padding:28px 32px; color:#fff; }}
  .header h1 {{ margin:0; font-size:20px; }}
  .header p  {{ margin:6px 0 0; opacity:.85; font-size:13px; }}
  .body {{ padding:28px 32px; }}
  .meta {{ background:#f8f9fc; border-radius:8px; padding:14px 18px;
           font-size:13px; color:#555; margin-bottom:20px; }}
  .meta span {{ font-weight:700; color:#333; }}
  .content-box {{ background:#fff8e1; border-left:4px solid #f59e0b;
                  border-radius:6px; padding:16px 20px; font-size:14px;
                  line-height:1.8; color:#333; white-space:pre-wrap; }}
  .footer {{ background:#f4f6f9; padding:16px 32px; font-size:12px;
             color:#888; text-align:center; }}
</style></head><body>
<div class="wrap">
  <div class="header">
    <h1>💡 개선 요청이 접수되었습니다</h1>
    <p>DRB LOGIBOT-AI · 개선 요청 알림</p>
  </div>
  <div class="body">
    <div class="meta">
      <p>📅 접수 일시 : <span>{ts}</span></p>
      <p>👥 요청 모드 : <span>{team}</span></p>
    </div>
    <p style="font-size:14px;font-weight:700;color:#444;margin-bottom:10px;">📝 요청 내용</p>
    <div class="content-box">{content}</div>
  </div>
  <div class="footer">본 메일은 DRB LOGIBOT-AI 개선 요청 시스템에서 자동 발송되었습니다.</div>
</div>
</body></html>"""
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f"[LOGIBOT] 개선 요청 접수 - {datetime.now().strftime('%Y-%m-%d %H:%M')} ({team})"
            msg['From']    = formataddr(("DRB LOGIBOT-AI", self.email_from))
            msg['To']      = ', '.join(self.email_to)
            msg.attach(MIMEText(html_content, 'html', 'utf-8'))
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.username, self.password)
                server.send_message(msg)
            logger.info(f"✅ 개선 요청 이메일 전송 완료 ({team})")
            return True
        except Exception as e:
            logger.error(f"❌ 개선 요청 이메일 전송 실패: {e}")
            return False

    def _create_html_email(self, feedback_data: Dict) -> str:
        """HTML 형식의 이메일 본문 생성 (구조화 버전)"""
        query     = feedback_data.get('query', 'N/A')
        answer    = feedback_data.get('answer', 'N/A')
        timestamp = feedback_data.get('timestamp', 'N/A')
        reason    = feedback_data.get('reason', '사유 미입력')

        # 타임스탬프 포맷팅
        try:
            from datetime import datetime as _dt
            ts_display = _dt.fromisoformat(timestamp).strftime("%Y년 %m월 %d일 %H:%M")
        except Exception:
            ts_display = timestamp

        # 답변 마크다운 → HTML 변환 (굵게, 리스트, 표, 줄바꿈)
        def md_to_html(text: str) -> str:
            import re as _re
            # 표(table) 변환
            lines = text.split('\n')
            result, in_table, table_buf = [], False, []
            for line in lines:
                if '|' in line and line.strip().startswith('|'):
                    in_table = True
                    table_buf.append(line)
                else:
                    if in_table:
                        result.append(_render_md_table(table_buf))
                        table_buf = []; in_table = False
                    result.append(line)
            if in_table:
                result.append(_render_md_table(table_buf))
            text = '\n'.join(result)
            # **볼드**
            text = _re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', text)
            # ## 제목
            text = _re.sub(r'^#{1,3}\s+(.+)$', r'<h4 style="margin:10px 0 4px;color:#1e40af;">\1</h4>', text, flags=_re.MULTILINE)
            # - 리스트
            text = _re.sub(r'^[\-\*]\s+(.+)$', r'<li>\1</li>', text, flags=_re.MULTILINE)
            text = _re.sub(r'(<li>.*?</li>(\s*<li>.*?</li>)*)', r'<ul style="margin:6px 0;padding-left:20px;">\1</ul>', text, flags=_re.DOTALL)
            # 줄바꿈
            text = text.replace('\n', '<br>')
            return text

        def _render_md_table(lines):
            import re as _re
            rows = [l for l in lines if not _re.match(r'^\s*\|[\s\-:]+\|\s*$', l)]
            html = '<table style="border-collapse:collapse;width:100%;margin:10px 0;font-size:13px;">'
            for i, row in enumerate(rows):
                cells = [c.strip() for c in row.strip().strip('|').split('|')]
                tag = 'th' if i == 0 else 'td'
                style = ('background:#1e40af;color:white;padding:8px 10px;text-align:left;'
                         if i == 0 else
                         f'padding:7px 10px;border-bottom:1px solid #e2e8f0;background:{"#f8fafc" if i%2==0 else "white"};')
                html += '<tr>' + ''.join(f'<{tag} style="{style}">{c}</{tag}>' for c in cells) + '</tr>'
            html += '</table>'
            return html

        answer_html = md_to_html(answer[:3000] + ("..." if len(answer) > 3000 else ""))
        reason_display = reason if reason.strip() else "사유 미입력"

        html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  body{{font-family:'Malgun Gothic','Apple SD Gothic Neo',sans-serif;background:#f1f5f9;margin:0;padding:20px;}}
  .wrap{{max-width:680px;margin:0 auto;}}
  .header{{background:linear-gradient(135deg,#dc2626,#9f1239);color:white;padding:28px 32px;border-radius:12px 12px 0 0;}}
  .header h1{{margin:0 0 6px;font-size:22px;}}
  .header p{{margin:0;font-size:13px;opacity:.85;}}
  .body{{background:white;padding:28px 32px;border-radius:0 0 12px 12px;border:1px solid #e2e8f0;border-top:none;}}
  .section{{margin-bottom:22px;}}
  .section-title{{font-size:13px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px;}}
  .card{{background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:14px 16px;font-size:14px;color:#334155;line-height:1.7;}}
  .card.red{{border-left:4px solid #dc2626;}}
  .card.blue{{border-left:4px solid #3b82f6;}}
  .card.amber{{border-left:4px solid #f59e0b;}}
  .reason-tag{{display:inline-block;background:#fee2e2;color:#b91c1c;padding:3px 10px;border-radius:20px;font-size:12px;font-weight:600;margin:2px;}}
  .loop-section{{background:linear-gradient(135deg,#eff6ff,#f0fdf4);border:1px solid #bfdbfe;border-radius:12px;padding:20px 24px;margin-bottom:22px;text-align:center;}}
  .loop-section p{{margin:0 0 14px;font-size:13px;color:#475569;line-height:1.7;}}
  .loop-btn{{display:inline-block;background:linear-gradient(135deg,#0ea5e9,#6366f1);color:#ffffff !important;text-decoration:none !important;padding:12px 28px;border-radius:8px;font-size:14px;font-weight:700;letter-spacing:.3px;box-shadow:0 4px 12px rgba(99,102,241,.35);}}
  .loop-btn:hover{{opacity:.9;}}
  .loop-hint{{margin:10px 0 0;font-size:11px;color:#94a3b8;}}
  .meta{{font-size:12px;color:#94a3b8;margin-top:20px;text-align:right;border-top:1px solid #f1f5f9;padding-top:12px;}}
  .footer{{text-align:center;margin-top:18px;font-size:11px;color:#94a3b8;}}
</style>
</head>
<body>
<div class="wrap">
  <div class="header">
    <h1>👎 부정 피드백 알림</h1>
    <p>DRB 물류 AI 챗봇 · 답변 품질 개선이 필요합니다</p>
  </div>
  <div class="body">

    <div class="section">
      <div class="section-title">📝 사용자 질문</div>
      <div class="card red">{query}</div>
    </div>

    <div class="section">
      <div class="section-title">🤖 AI 답변</div>
      <div class="card blue">{answer_html}</div>
    </div>

    <div class="section">
      <div class="section-title">⚠️ 부정 피드백 사유</div>
      <div class="card amber">
        {''.join(f'<span class="reason-tag">{r.strip()}</span>' for r in reason_display.replace('| 추가의견:', '|').split('/'))}
        {'<br><span style="font-size:13px;color:#78350f;margin-top:8px;display:block;">💬 추가 의견: ' + reason_display.split('추가의견:')[1].strip() + '</span>' if '추가의견:' in reason_display else ''}
      </div>
    </div>

    <div class="loop-section">
      <p>
        📋 <strong>팀원 여러분의 소중한 의견이 AI 학습 데이터 개선에 직접 반영됩니다.</strong><br>
        아래 버튼을 클릭하여 학습 데이터 문서에 개선 내용을 직접 기록해 주세요.<br>
        작성해 주신 피드백은 더 정확한 AI 답변을 만드는 데 큰 힘이 됩니다. 🙏
      </p>
      <a href="https://drbworld-my.sharepoint.com/:x:/p/shin_tae_hwan/IQAg1gyD-21uSKlKo0KBbSq7ATansZ4KE7m3nctSSvyd7I4?e=ooGJuI"
         class="loop-btn" target="_blank">
        📝 &nbsp;학습 데이터 개선 문서 열기
      </a>
      <div class="loop-hint">※ 사내 Microsoft 365 계정으로 로그인이 필요할 수 있습니다</div>
    </div>

    <div class="meta">⏰ 피드백 시각: {ts_display}</div>
  </div>
  <div class="footer">
    이 메일은 부정 피드백 발생 시 자동으로 발송됩니다 · DRB LOGIBOT-AI
  </div>
</div>
</body>
</html>"""
        return html

class OnPremiseGemmaLLM(LLM):
    """온프레미스 LLM (openai 호환 API)"""
    api_url: str = ONPREMISE_API_URL
    model:   str = ONPREMISE_MODEL
    api_key: str = ONPREMISE_API_KEY
    timeout: int = ONPREMISE_TIMEOUT
    max_retries: int = 3
    temperature: float = 0.2
    _last_call_ts:    float = PrivateAttr(default=0.0)
    _rate_limit_delay: float = PrivateAttr(default=1.0)

    @property
    def _identifying_params(self) -> Dict[str, Any]:
        return {"api_url": self.api_url, "model": self.model}

    @property
    def _llm_type(self) -> str:
        return "onpremise_llm"

    def _enforce_rate_limit(self):
        elapsed = time.time() - self._last_call_ts
        if elapsed < self._rate_limit_delay:
            time.sleep(self._rate_limit_delay - elapsed)
        self._last_call_ts = time.time()

    def _get_client(self) -> openai.OpenAI:
        return openai.OpenAI(
            api_key=self.api_key or "dummy",  # 키 없을 때 dummy 허용
            base_url=self.api_url,
        )

    def _call(self, prompt: str, stop: Optional[List[str]] = None) -> str:
        return self._call_with_max_tokens(prompt, max_tokens=4096)

    def _call_with_max_tokens(self, prompt: str, max_tokens: int = 2048) -> str:
        self._enforce_rate_limit()
        client = self._get_client()

        for attempt in range(1, self.max_retries + 1):
            try:
                response = client.chat.completions.create(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                    temperature=self.temperature,
                    max_tokens=max_tokens,
                    top_p=1,
                    timeout=self.timeout,
                )
                content = response.choices[0].message.content
                return content or ""

            except openai.RateLimitError:
                logger.warning(f"Rate limit (시도 {attempt})")
                if attempt < self.max_retries:
                    time.sleep(5 * attempt)
                else:
                    raise RuntimeError("Rate limit exceeded")

            except openai.APITimeoutError:
                logger.warning(f"API timeout (시도 {attempt})")
                if attempt < self.max_retries:
                    time.sleep(2 * attempt)
                else:
                    raise RuntimeError(f"API timeout after {self.max_retries} retries")

            except Exception as e:
                logger.error(f"API 호출 오류 (시도 {attempt}): {e}")
                if attempt < self.max_retries:
                    time.sleep(2 * attempt)
                else:
                    raise

        logger.error("_call_with_max_tokens: 모든 재시도 실패")
        return ""


class RAGChainWrapper:
    """RAG 체인 래퍼 (검색 최적화)"""

    # 시트명 → 도메인 매핑
    # source값(시트명) → domain 매핑 (data_loader V5 기준)
    SHEET_TO_DOMAIN = {
        # source 값 (시트명) 기준
        "컨베어벨트 규격 데이터"          : "conveyor",
        "주름혹벨트 우든박스 사이즈 데이터" : "sidewall",
        "크롤러 러버트랙 규격 데이터"       : "crawler",
        # domain 값 직접 매핑 (V5에서 source 대신 domain이 저장된 경우)
        "conveyor"   : "conveyor",
        "sidewall"   : "sidewall",
        "crawler"    : "crawler",
    }

    def _sheet_to_domain(self, sheet_name: str) -> Optional[str]:
        """시트명(source) 또는 domain값 → 도메인 반환"""
        # 정확 일치 먼저
        if sheet_name in self.SHEET_TO_DOMAIN:
            return self.SHEET_TO_DOMAIN[sheet_name]
        # 부분 포함
        for key, domain in self.SHEET_TO_DOMAIN.items():
            if key in sheet_name or sheet_name in key:
                return domain
        return None

    def _domain_from_code(self, code: str) -> Optional[str]:
        """코드만으로 도메인 판별 (매핑 캐시 사용)"""
        val = self._code_sheet_map.get(str(code))
        if not val:
            return None
        if val in ("conveyor", "sidewall", "crawler"):
            return val
        return self._sheet_to_domain(val)

    def __init__(self, vectorstore, llm, embeddings, qdrant_client):
        self.vectorstore = vectorstore
        self.llm = llm
        self.embeddings = embeddings
        self.qdrant_client = qdrant_client
        self.prompt_template = ChatPromptTemplate.from_template(PROMPT_TEMPLATE)
        self._code_sheet_map: dict = {}
        self._build_code_sheet_map()

    def _build_code_sheet_map(self):
        """
        Qdrant logistics_data에서 material_code 페이로드를 읽어
        {코드: domain} 매핑 딕셔너리를 빌드.
        data_loader V5: material_code/domain/source 최상위 키로 저장
        """
        try:
            offset = None
            batch_size = 500
            while True:
                result = self.qdrant_client.scroll(
                    collection_name=QDRANT_COLLECTION,
                    limit=batch_size,
                    offset=offset,
                    with_payload=True,
                    with_vectors=False
                )
                points, next_offset = result
                for point in points:
                    p = point.payload
                    code = p.get("material_code")
                    if not code:
                        code = p.get("metadata", {}).get("material_code")
                    if not code:
                        continue
                    sheet = (
                        p.get("source") or
                        p.get("sheet_name") or
                        p.get("metadata", {}).get("sheet_name") or
                        p.get("domain", "")
                    )
                    if code and sheet:
                        self._code_sheet_map[str(code)] = sheet

                if next_offset is None or len(points) < batch_size:
                    break
                offset = next_offset

            logger.info(f"✅ 코드-시트 매핑 빌드 완료: {len(self._code_sheet_map)}개 코드")
        except Exception as e:
            logger.warning(f"코드-시트 매핑 빌드 실패 (fallback 사용): {e}")
            self._code_sheet_map = {}

    @lru_cache(maxsize=100)
    def extract_material_code(self, query: str) -> Optional[str]:
        """자재코드(7자리) 추출 - 한글 앞뒤 경계도 처리"""
        patterns = [
            r'자재코드[:\s]*(\d{7})',
            r'코드[:\s]*(\d{7})',
            r'품번[:\s]*(\d{7})',
            # 숫자 앞뒤로 숫자가 없는 경우 (한글/공백/문장부호 포함)
            r'(?<![0-9])(\d{7})(?![0-9])',
        ]
        for pattern in patterns:
            match = re.search(pattern, query)
            if match:
                return match.group(1)
        return None
    
    def keyword_search_optimized(self, material_code: str, limit: int = 5):
        """
        자재코드 필터 검색.
        1차: 최상위 material_code 필터 (data_loader 최신 버전)
        2차: metadata.material_code 필터 (구버전 호환)
        """
        points_found = []
        try:
            # 1차: 최상위 키
            result = self.qdrant_client.scroll(
                collection_name=QDRANT_COLLECTION,
                scroll_filter=Filter(
                    must=[FieldCondition(key="material_code",
                                        match=MatchValue(value=material_code))]
                ),
                limit=limit, with_payload=True, with_vectors=False
            )
            points_found = result[0]
        except Exception as e:
            logger.warning(f"1차 코드 검색 실패: {e}")

        # 2차: metadata 중첩 키 (구버전)
        if not points_found:
            try:
                result = self.qdrant_client.scroll(
                    collection_name=QDRANT_COLLECTION,
                    scroll_filter=Filter(
                        must=[FieldCondition(key="metadata.material_code",
                                            match=MatchValue(value=material_code))]
                    ),
                    limit=limit, with_payload=True, with_vectors=False
                )
                points_found = result[0]
            except Exception as e:
                logger.warning(f"2차 코드 검색 실패: {e}")

        matched_docs = []
        for point in points_found:
            payload = point.payload
            # text 키 우선 (data_loader V5), 없으면 page_content (구버전 호환)
            content = payload.get('text') or payload.get('page_content', '')
            matched_docs.append({
                'id'      : point.id,
                'content' : content,
                'metadata': {k: v for k, v in payload.items() if k != 'text'},
                'score'   : 1.0
            })

        logger.info(f"키워드 검색 [{material_code}]: {len(matched_docs)}개")
        return matched_docs
        
    def _detect_domain(self, query: str, keyword_doc_content: str = "") -> str:
        """
        도메인 판별 — 3단계 전략:
          1단계) 확정적 코드 판별 (자재코드 매핑 / 제품명 / 기사 이름)
          2단계) 구조적 컨텍스트 판별 (업무 맥락이 명확한 키워드만 최소한으로 유지)
          3단계) LLM 경량 분류 (1~2단계에서 판별 못한 애매한 케이스)
        """
        combined = query + " " + keyword_doc_content

        # ══════════════════════════════════════════
        # 1단계: 확정적 코드 판별 (LLM 불필요, 100% 정확)
        # ══════════════════════════════════════════

        # 1-1. 기사 이름 직접 매칭 → driver_route (이름은 오해 여지 없음)
        DRIVER_NAMES = ["김병일", "김영철", "이용구", "심효섭"]
        if any(n in query for n in DRIVER_NAMES):
            return "driver_route"

        # 1-1-b. 노선 거래처명 감지 → driver_route
        # 거래처명은 사내 고유 명칭이라 LLM이 모를 수 있음
        _ROUTE_CUSTOMERS = [
            "GM대우 KD", "대신화물", "동양 ENG", "동양알앤비", "동일벨트",
            "동일벨트산업", "동일부품", "동일삼광산업", "동일상공사", "동일상사",
            "동일알앤씨", "동일종합물산", "동일종합산업", "동일종합상사", "동일팬벨트",
            "동일팬유통", "명진", "모던테크", "반도상사", "삼흥정밀",
            "서울팬벨트", "성보산업사", "성우상사", "승리상사",
            "신우기업", "신흥알앤테크", "신흥폴리테크", "유니온벨티노", "유일상사",
            "인왕산업", "제일산업사", "조선통상", "조일상공", "태영산업사",
            "한국벨트", "한길산업", "현대내자", "현대테크젠", "흥국상사", "흥진사",
        ]
        # 공백 제거 후 비교 (예: "현대 테크젠" → "현대테크젠")
        _q_no_space = query.replace(" ", "")
        if any(k.replace(" ", "") in _q_no_space for k in _ROUTE_CUSTOMERS):
            logger.info(f"노선 거래처명 감지 → 도메인: driver_route")
            return "driver_route"

        # 1-2. 제품 고유명 → 기술 도메인 (오분류 시 계산 오답 위험)
        if any(k in combined for k in ["컨베어벨트", "컨베이어벨트", "컨베이어 벨트"]):
            # 담당자 질문 제외
            if not any(k in combined for k in ["담당", "누구", "연락", "직책"]):
                return "conveyor"
        if any(k in combined for k in ["러버트랙", "Rubber Track", "rubber track"]):
            if not any(k in combined for k in ["담당", "누구", "연락", "직책"]):
                return "crawler"
        if any(k in combined for k in ["주름혹벨트", "주름혹 벨트"]):
            return "sidewall"

        # 1-3. 자재코드 매핑 (가장 정확한 판별)
        code = self.extract_material_code(query)
        if code:
            domain_from_code = self._domain_from_code(code)
            if domain_from_code:
                logger.info(f"코드 {code} → 도메인: {domain_from_code}")
                return domain_from_code
            # 코드 있는데 캐시 미스 → 키워드 보조
            if any(k in query for k in ["포장", "총 중량", "우든", "박스 무게"]):
                return "sidewall"

        # 1-4. "컨베어" 단독 + 배차/차량 → 시뮬레이터 유도 대상 (conveyor)
        if any(k in combined for k in ["컨베어", "컨베이어"]):
            if not any(k in combined for k in ["담당", "누구", "연락"]):
                return "conveyor"

        # 1-5. "크롤러" 단독
        if "크롤러" in combined and not any(k in combined for k in ["담당", "누구", "연락"]):
            return "crawler"

        # ══════════════════════════════════════════
        # 2단계: 구조적 컨텍스트 판별 (명확한 맥락 키워드만 유지)
        # ══════════════════════════════════════════

        # 2-1. 기사 노선/동선 — 이름 없이도 맥락이 명확한 것만
        if any(k in combined for k in ["납품 동선", "납품동선", "기사 노선", "납품 노선",
                                         "지입기사", "지입 기사", "납품경로"]):
            return "driver_route"

        # 2-2. 수출 포장 — B01/B02 같은 사내 전용 코드는 LLM이 모름
        if any(k in combined for k in ["B01", "B02", "N18", "N19", "CBM", "cbm",
                                         "수출 파렛트", "컨테이너", "수출 포장량"]):
            return "export"

        # 2-3. 파렛트/박스 규격 조회 (배차/운송 맥락 없을 때만)
        _is_dispatch = any(k in query for k in ["배차", "차량", "운송", "몇 톤", "몇톤"])
        if any(k in combined for k in ["PLT", "plt", "파렛트", "박스 규격", "PE포", "받침목",
                                         "포장자재", "포장재", "포장 종류", "포장 자재"]):
            if not _is_dispatch:
                return "pallet_box"

        # 2-4. 차량 제원 명시 질문
        if any(k in combined for k in ["적재함 길이", "적재함 폭", "차량 제원", "차량제원",
                                         "톤수별 차량", "차량 규격"]):
            return "vehicle"

        # ══════════════════════════════════════════
        # 3단계: LLM 경량 분류 (1~2단계 미판별 → 의도 파악 필요)
        # ══════════════════════════════════════════
        # 대상: personnel / operation_rule / domestic / general
        # 이 4개는 표현이 너무 다양해서 키워드로 커버 불가
        try:
            classify_prompt = f"""당신은 물류팀 사내 챗봇의 질문 분류기입니다. 반드시 아래 4가지 중 하나만 영어로 답하세요. 다른 말은 절대 금지.

personnel      : 팀원/담당자/기사 정보 (연락처, 인원현황, 누가 담당인지, 지게차 기사 현황)
operation_rule : 물류 업무 절차·규칙·방법 질문 전반. 아래를 모두 포함:
                 마감시간(주문/출고/배차/직배차/용차/직송), 배차 신청·절차,
                 지게차 요청, 출고 변경·취소, 포장 방법·분단, 운임 기준,
                 샘플 발송·수령, 창고·보관, 짐 이동, 업무 문의 등
                 ※ "직배차량","직배차","고정 용차","직배"처럼 표현이 달라도
                   배차·마감·출고 맥락이면 operation_rule
domestic       : 국내 운송수단 선택·비교 (화물vs택배vs직송, 운임비교, 어떻게 보낼지)
general        : 위 3가지 외

예시) "이정희 주임 연락처" → personnel
예시) "지게차 어떻게 신청해" → operation_rule
예시) "3PLT 배차 요청 몇 톤 차량" → operation_rule
예시) "인천까지 배차 언제 신청" → operation_rule
예시) "택배vs직송 어떤게 나을까" → domestic
예시) "샘플을 보내려고 하는데 어떻게 해야해" → operation_rule
예시) "벨트 샘플 수령 방법" → operation_rule
예시) "샘플 발송 프로세스" → operation_rule
예시) "직배차량 마감시간" → operation_rule
예시) "직배차 마감 언제야" → operation_rule
예시) "고정 용차 마감시간" → operation_rule
예시) "직송으로 보낼때 추가 운임 항목" → operation_rule
예시) "직송 추가운임 발생하는 경우" → operation_rule
예시) "직송으로 보내면 얼마야" → domestic
예시) "3PLT 운송 몇톤 차량 배차요청" → operation_rule
예시) "파렛트 운송 차량 톤수 알려줘" → operation_rule
예시) "제주도에 배송하려고 하는데 어떻게 해야해" → operation_rule
예시) "제주도 배송 되나요" → operation_rule
예시) "동양알앤비 납품 담당자 누구야" → driver_route
예시) "유일상사에 납품하려는데 문의처" → driver_route

질문: {query}
답:"""

            # 도메인 분류는 결정적 응답 필요 → temperature=0 강제
            _orig_temp = self.llm.temperature
            self.llm.temperature = 0.0
            try:
                raw_intent = self.llm._call(classify_prompt).strip().lower()
            finally:
                self.llm.temperature = _orig_temp

            # 응답 파싱 (LLM이 여분 텍스트를 붙일 수 있으므로 포함 검사)
            if "personnel" in raw_intent:
                domain = "personnel"
            elif "operation" in raw_intent:
                domain = "operation_rule"
            elif "domestic" in raw_intent:
                domain = "domestic"
            else:
                domain = "general"

            logger.info(f"LLM 도메인 분류: '{raw_intent}' → {domain}")
            return domain

        except Exception as e:
            logger.warning(f"LLM 분류 실패({e}) → general 폴백")
            return "general"

    def fetch_whole_docs(self, sheet_names: list, limit: int = 5):
        """
        data_loader V5 기준: source 키로 저장된 문서를 페이로드 필터로 가져옴.
        저장 구조: payload = {**meta, "text": ...}  (text/domain/source 최상위)
        """
        from langchain_core.documents import Document
        docs = []
        try:
            for sheet in sheet_names:
                points_found = []

                # 1차: source 키 필터 (data_loader V5 기준)
                try:
                    result = self.qdrant_client.scroll(
                        collection_name=QDRANT_COLLECTION,
                        scroll_filter=Filter(
                            must=[FieldCondition(
                                key="source",
                                match=MatchValue(value=sheet)
                            )]
                        ),
                        limit=limit,
                        with_payload=True,
                        with_vectors=False
                    )
                    points_found = result[0]
                except Exception:
                    pass

                # 2차: domain 키 필터 (source 매칭 실패 시)
                if not points_found:
                    # source 시트명 → domain 매핑
                    _SHEET_TO_DOMAIN = {
                        "컨베어벨트 직경 산출 수식":          "conveyor_formula",
                        "주름혹벨트 우든박스 사이즈 데이터":   "sidewall",
                        "차량 데이터":                        "vehicle",
                        "물류팀 운영 규칙":                   "operation_rule",
                        "용차 차량 노선 데이터":              "route",
                        "지입 차량(기사) 노선 데이터":        "driver_route",
                        "포장량 산출 데이터":                 "packaging",
                        "수출 포장량 산출 수식":              "export_rule",
                        "물류팀 현황 데이터":                 "personnel",
                        "크롤러 러버트랙 규격 데이터":        "crawler",
                        "파렛트, 박스 데이터":               "pallet_box",
                    }
                    domain_val = _SHEET_TO_DOMAIN.get(sheet)
                    if domain_val:
                        try:
                            result = self.qdrant_client.scroll(
                                collection_name=QDRANT_COLLECTION,
                                scroll_filter=Filter(
                                    must=[FieldCondition(
                                        key="domain",
                                        match=MatchValue(value=domain_val)
                                    )]
                                ),
                                limit=limit,
                                with_payload=True,
                                with_vectors=False
                            )
                            points_found = result[0]
                        except Exception:
                            pass

                # 3차: 구버전 sheet_name 필터 (이전 data_loader 호환)
                if not points_found:
                    try:
                        result = self.qdrant_client.scroll(
                            collection_name=QDRANT_COLLECTION,
                            scroll_filter=Filter(
                                must=[FieldCondition(
                                    key="sheet_name",
                                    match=MatchValue(value=sheet)
                                )]
                            ),
                            limit=limit,
                            with_payload=True,
                            with_vectors=False
                        )
                        points_found = result[0]
                    except Exception:
                        pass

                for point in points_found:
                    payload = point.payload
                    # text 키 우선, 없으면 page_content (구버전 호환)
                    text = payload.get('text') or payload.get('page_content', '')
                    meta = {k: v for k, v in payload.items() if k != 'text'}
                    doc = Document(page_content=text, metadata=meta)
                    if doc.page_content:
                        docs.append((doc, 0.9))

                logger.info(f"fetch_whole_docs [{sheet}]: {len(points_found)}개 조회")

        except Exception as e:
            logger.warning(f"WHOLE 문서 보완 실패: {e}")
        return docs

    def fetch_driver_docs_by_group(self, target_date=None) -> list:
        """
        driver_route 전체 문서를 가져옴 (공통 + A + B 모두).
        격주 필터링은 _format_driver_route_answer 내부에서 수행.
        → 이번 주 미운행 기사를 명시적으로 질문해도 '다음 주 노선 미리보기'로 답변 가능.
        """
        from langchain_core.documents import Document
        wg = get_week_group(target_date)
        docs = []
        try:
            # 공통 + A + B 전체 조회
            result = self.qdrant_client.scroll(
                collection_name=QDRANT_COLLECTION,
                scroll_filter=Filter(
                    must=[FieldCondition(
                        key="domain",
                        match=MatchValue(value="driver_route")
                    )]
                ),
                limit=100,
                with_payload=True,
                with_vectors=False,
            )
            for point in result[0]:
                payload = point.payload
                text = payload.get("text", payload.get("page_content", ""))
                if text:
                    doc = Document(
                        page_content=text,
                        metadata={k: v for k, v in payload.items() if k != "text"},
                    )
                    docs.append((doc, 0.9))
            logger.info(f"driver_route 전체 조회(이번주={wg}): {len(docs)}건")
        except Exception as e:
            logger.warning(f"driver_route domain 필터 실패 → source fallback: {e}")
            docs = self.fetch_whole_docs(["지입 차량(기사) 노선 데이터"], limit=100)

        if not docs:
            logger.warning("driver_route 조회 0건 → 재적재 필요")

        return docs

    # 도메인별 보완 시트 매핑
    DOMAIN_SUPPLEMENT_SHEETS = {
        # Fix: conveyor에 차량 데이터 추가 → 컨베어 배차 질문 답변 가능
        "conveyor"      : ["컨베어벨트 직경 산출 수식", "차량 데이터", "물류팀 운영 규칙"],
        "sidewall"      : ["주름혹벨트 우든박스 사이즈 데이터"],
        "crawler"       : ["차량 데이터"],
        "domestic"      : ["물류팀 운영 규칙", "용차 차량 노선 데이터", "차량 데이터"],
        "export"        : ["수출 포장량 산출 수식", "포장량 산출 데이터", "물류팀 운영 규칙"],
        "driver_route"  : ["지입 차량(기사) 노선 데이터"],
        "personnel"     : ["물류팀 현황 데이터"],
        "operation_rule": ["물류팀 운영 규칙", "물류팀 현황 데이터"],
        # Fix: vehicle 도메인 신규 추가 → 차량 제원 질문 전용
        "vehicle"       : ["차량 데이터"],
        "pallet_box"    : ["파렛트, 박스 데이터"],  # Fix: PLT/박스 도메인 추가
        # general: 운영규칙 먼저 → 차량 데이터 순
        "general"       : ["물류팀 운영 규칙", "차량 데이터", "물류팀 현황 데이터"],
    }

    def hybrid_search(self, query: str, k: int = 50):
        material_code = self.extract_material_code(query)
        results = []

        if material_code:
            keyword_results = self.keyword_search_optimized(material_code, limit=5)

            if keyword_results:
                from langchain_core.documents import Document
                for result in keyword_results:
                    doc = Document(
                        page_content=result['content'],
                        metadata=result['metadata']
                    )
                    results.append((doc, result['score']))

                # 코드-시트 매핑 우선, 없으면 문서 내용으로 도메인 판별
                first_doc_content = keyword_results[0].get('content', '')
                domain = self._detect_domain(query, first_doc_content)
                supplement_sheets = list(self.DOMAIN_SUPPLEMENT_SHEETS.get(domain, []))

                # ★ domestic 도메인: 자재코드 기반 도메인 시트도 추가 보완
                #   (운송방식 판단에는 물류팀 운영 규칙 + 자재 중량 데이터 모두 필요)
                if domain == "domestic":
                    code_domain = self._domain_from_code(material_code)
                    if code_domain:
                        extra_sheets = self.DOMAIN_SUPPLEMENT_SHEETS.get(code_domain, [])
                        for s in extra_sheets:
                            if s not in supplement_sheets:
                                supplement_sheets = [s] + supplement_sheets
                    logger.info(f"domestic 보완 시트 (자재+운송): {supplement_sheets}")

                if supplement_sheets:
                    whole_docs = self.fetch_whole_docs(supplement_sheets, limit=2)
                    results.extend(whole_docs)
                    logger.info(f"코드 검색: {len(keyword_results)}개 | 도메인: {domain} | 보완: {supplement_sheets} ({len(whole_docs)}개)")
                else:
                    logger.info(f"코드 검색: {len(keyword_results)}개 | 도메인: {domain} | 보완 없음")
                return results

            else:
                # Qdrant 필터 검색 실패 → 코드-시트 캐시로 도메인 판별 후 벡터 검색 보완
                domain_from_code = self._domain_from_code(material_code)
                if domain_from_code:
                    logger.warning(f"코드 {material_code} Qdrant 조회 실패 → 도메인 {domain_from_code}로 벡터 검색 보완")
                    supplement_sheets = list(self.DOMAIN_SUPPLEMENT_SHEETS.get(domain_from_code, []))
                    # ★ domestic 도메인이면 물류팀 운영 규칙도 추가
                    detected = self._detect_domain(query)
                    if detected == "domestic":
                        for s in self.DOMAIN_SUPPLEMENT_SHEETS.get("domestic", []):
                            if s not in supplement_sheets:
                                supplement_sheets.append(s)
                    if supplement_sheets:
                        whole_docs = self.fetch_whole_docs(supplement_sheets, limit=3)
                        results.extend(whole_docs)
                # 결과가 없으면 아래 벡터 검색으로 fallback

        try:
            vector_results = self.vectorstore.similarity_search_with_score(query, k=50)  # Fix: V4 동일 50개

            # ── 도메인별 유사도 임계값 차등 적용 ─────────────────────────────
            _domain_now = self._detect_domain(query)
            SCORE_THRESHOLD = {
                "conveyor"       : 0.15,
                "crawler"        : 0.15,
                "sidewall"       : 0.15,
                "domestic"       : 0.15,
                "export"         : 0.15,
                "personnel"      : 0.12,
                "operation_rule" : 0.12,
                "vehicle"        : 0.12,
                "pallet_box"     : 0.12,
                "driver_route"   : 0.12,
                "general"        : 0.12,
            }
            _threshold = SCORE_THRESHOLD.get(_domain_now, 0.35)

            if len(vector_results) > 3:
                filtered_results = [
                    (doc, score) for doc, score in vector_results if score >= _threshold
                ]
                if len(filtered_results) < 2:
                    filtered_results = [
                        (doc, score) for doc, score in vector_results if score >= 0.10
                    ]
                    logger.info(f"임계값 완화: {_threshold} → 0.10 (결과 부족)")
                logger.info(f"도메인={_domain_now} 임계값={_threshold} 결과={len(filtered_results)}개")
            else:
                filtered_results = vector_results

            # 이미 results에 코드 기반 문서가 있으면 합산
            if results:
                existing = {doc.page_content[:50] for doc, _ in filtered_results}
                for doc, score in results:
                    if doc.page_content[:50] not in existing:
                        filtered_results.insert(0, (doc, score))
                        existing.add(doc.page_content[:50])

            domain = self._detect_domain(query)

            # 자재코드 질문인데 벡터 결과만 있는 경우 → 해당 도메인 시트 문서 강제 보완
            if material_code and domain not in ("driver_route", "general"):
                supplement_sheets = self.DOMAIN_SUPPLEMENT_SHEETS.get(domain, [])
                if supplement_sheets:
                    existing = {doc.page_content[:50] for doc, _ in filtered_results}
                    extra = self.fetch_whole_docs(supplement_sheets, limit=2)
                    for doc, score in extra:
                        if doc.page_content[:50] not in existing:
                            filtered_results.append((doc, score))
                    logger.info(f"코드 벡터 fallback 보완: 도메인={domain}, +{len(extra)}개")

            # general 도메인 중 배차/차량 질문 → 차량 데이터 외 타도메인 문서 제거
            _vehicle_query_kw = ["배차 할", "배차할", "배차 가능", "차량 목록", "차량 종류",
                                  "차량 리스트", "어떤 차량", "차량에는 어떤", "차량 있어"]
            if domain == "general" and any(k in query for k in _vehicle_query_kw):
                veh_docs = self.fetch_whole_docs(["차량 데이터"], limit=20)
                filtered_results = [
                    (doc, score) for doc, score in filtered_results
                    if doc.metadata.get("domain") == "vehicle"
                    or doc.metadata.get("source") == "차량 데이터"
                ]
                existing = {doc.page_content[:50] for doc, _ in filtered_results}
                for doc, score in veh_docs:
                    if doc.page_content[:50] not in existing:
                        filtered_results.append((doc, score))
                logger.info(f"general 차량목록 질문 → 차량 데이터만 {len(filtered_results)}개")

            # general 도메인 → 운영 규칙 강제 보완 (샘플/주문/배송 등 일반 업무 질문)
            if domain in ("general", "operation_rule"):
                op_docs = self.fetch_whole_docs(["물류팀 운영 규칙"], limit=100)  # summary 1개 + Q&A 전체 (51개 이상 대비)
                if op_docs:
                    # summary 청크를 맨 앞으로 정렬 → context 앞부분에 전체 Q&A 요약 배치
                    # 개별 Q&A를 앞에, summary를 뒤에 배치 → 유사 Q&A가 context 앞부분에 오도록
                    qa_first = sorted(op_docs, key=lambda x: 1 if x[0].metadata.get("type") == "summary" else 0)
                    existing = {doc.page_content[:50] for doc, _ in filtered_results}
                    added = 0
                    for doc, score in qa_first:
                        if doc.page_content[:50] not in existing:
                            filtered_results.append((doc, score))  # 뒤에 추가 (Q&A 먼저, summary 나중)
                            existing.add(doc.page_content[:50])
                            added += 1
                    logger.info(f"general/operation_rule 운영규칙 보완: +{added}개")

                    # 샘플 관련 질문이면 샘플 문서를 컨텍스트 앞으로 강제 이동
                    _sample_kw = ["샘플", "sample", "Sample"]
                    if any(k in query for k in _sample_kw):
                        _sample_docs = [
                            (d, s) for d, s in filtered_results
                            if "샘플" in d.page_content
                        ]
                        _other_docs = [
                            (d, s) for d, s in filtered_results
                            if "샘플" not in d.page_content
                        ]
                        filtered_results = _sample_docs + _other_docs
                        if _sample_docs:
                            logger.info(f"샘플 문서 {len(_sample_docs)}개 컨텍스트 앞에 배치")

            # pallet_box 도메인 → 파렛트/박스 데이터 전체 강제 보완
            if domain == "pallet_box":
                plt_docs = self.fetch_whole_docs(["파렛트, 박스 데이터"], limit=20)
                if plt_docs:
                    filtered_results = [
                        (doc, score) for doc, score in filtered_results
                        if doc.metadata.get("domain") == "pallet_box"
                        or doc.metadata.get("source") == "파렛트, 박스 데이터"
                    ]
                    existing = {doc.page_content[:50] for doc, _ in filtered_results}
                    for doc, score in plt_docs:
                        if doc.page_content[:50] not in existing:
                            filtered_results.append((doc, score))
                    logger.info(f"pallet_box 강제 보완: {len(filtered_results)}개")

            # vehicle 도메인 → 차량 데이터만 강제 보완, 타도메인 혼입 제거
            if domain == "vehicle":
                veh_docs = self.fetch_whole_docs(["차량 데이터"], limit=20)
                if veh_docs:
                    # 벡터 결과에서 vehicle 이외 도메인 문서 제거 → 조합 오답 방지
                    filtered_results = [
                        (doc, score) for doc, score in filtered_results
                        if doc.metadata.get("domain") == "vehicle"
                        or doc.metadata.get("source") == "차량 데이터"
                    ]
                    existing = {doc.page_content[:50] for doc, _ in filtered_results}
                    for doc, score in veh_docs:
                        if doc.page_content[:50] not in existing:
                            filtered_results.append((doc, score))
                    logger.info(f"vehicle 강제 보완: 타도메인 제거 후 {len(filtered_results)}개")

            # personnel 도메인 → 물류팀 현황 데이터 전체 강제 보완
            # Fix: personnel은 23개 문서 → limit=30으로 전체 보장
            if domain == "personnel":
                # 운영규칙도 함께 보완 (지게차 요청 등이 personnel로 분류될 때 대비)
                per_docs = self.fetch_whole_docs(["물류팀 현황 데이터", "물류팀 운영 규칙"], limit=30)
                if per_docs:
                    existing = {doc.page_content[:50] for doc, _ in filtered_results}
                    added = 0
                    for doc, score in per_docs:
                        if doc.page_content[:50] not in existing:
                            filtered_results.append((doc, score))
                            existing.add(doc.page_content[:50])
                            added += 1
                    # 벡터 결과에서 personnel 이외 도메인 문서 제거 → 데이터 혼입 방지
                    # ★ 운영규칙도 허용: 지게차 기사 현황 등이 운영규칙 시트에 존재
                    filtered_results = [
                        (doc, score) for doc, score in filtered_results
                        if doc.metadata.get("domain") == "personnel"
                        or doc.metadata.get("source") == "물류팀 현황 데이터"
                        or doc.metadata.get("source") == "물류팀 운영 규칙"
                        or doc.metadata.get("domain") == "operation_rule"
                    ]
                    logger.info(f"personnel 강제 보완: +{added}개, 타도메인 문서 제거 후 {len(filtered_results)}개")

            # 지입기사 납품 동선: 전체 20개 문서가 필요 (A+B+공통)
            if domain == "driver_route":
                driver_docs = self.fetch_driver_docs_by_group()
                # 20개 미만이면 일부 route_group 누락 → 엑셀 직접 로드
                if len(driver_docs) < 18:
                    logger.warning(f"driver_route {len(driver_docs)}건 → 엑셀 직접 로드로 보완")
                    excel_ctx = build_driver_context_from_excel()
                    if excel_ctx:
                        from langchain_core.documents import Document as _Doc
                        excel_doc = _Doc(page_content=excel_ctx, metadata={"domain":"driver_route","source":"excel_direct"})
                        driver_docs = [(excel_doc, 1.0)]
                if driver_docs:
                    filtered_results = driver_docs
                    logger.info(f"지입기사 노선 전용 컨텍스트: {len(driver_docs)}개")
                else:
                    driver_kw = ["지입", "기사", "노선", "납품", "동선"]
                    driver_filtered = [
                        (doc, score) for doc, score in filtered_results
                        if any(kw in doc.page_content for kw in driver_kw)
                        or "지입 차량" in doc.metadata.get("sheet_name", "")
                    ]
                    if driver_filtered:
                        filtered_results = driver_filtered
                        logger.info(f"지입기사 벡터 결과 필터링: {len(driver_filtered)}개")
                    else:
                        logger.warning("지입기사 노선 문서 조회 실패 — 벡터 검색 전체 결과 사용")

            logger.info(f"벡터 검색 결과 확보: {len(filtered_results)}개")
            return filtered_results
        except Exception as e:
            logger.error(f"벡터 검색 실패: {e}")
            return results if results else []
          
class LoggingSystem:
    """질문/답변 로깅 시스템"""
    def __init__(self, qdrant_client: QdrantClient, embeddings: OllamaEmbeddings):
        self.client = qdrant_client
        self.embeddings = embeddings
        self._ensure_log_collections()
    
    def _ensure_log_collections(self):
        """로깅 전용 컬렉션 생성"""
        try:
            collections = self.client.get_collections().collections
            existing_names = [c.name for c in collections]
            
            sample_vector = self.embeddings.embed_query("test")
            vector_size = len(sample_vector)
            
            if QUERY_LOG_COLLECTION not in existing_names:
                self.client.create_collection(
                    collection_name=QUERY_LOG_COLLECTION,
                    vectors_config=VectorParams(size=vector_size, distance=Distance.COSINE)
                )
                logger.info(f"✅ {QUERY_LOG_COLLECTION} 생성 (로깅 전용)")
            
            if ANSWER_LOG_COLLECTION not in existing_names:
                self.client.create_collection(
                    collection_name=ANSWER_LOG_COLLECTION,
                    vectors_config=VectorParams(size=vector_size, distance=Distance.COSINE)
                )
                logger.info(f"✅ {ANSWER_LOG_COLLECTION} 생성 (로깅 전용)")
                
        except Exception as e:
            logger.error(f"로그 컬렉션 오류: {e}")
    
    def log_query(self, query: str, metadata: Dict = None, team: str = "",
                  cache_hit: bool = False, hour: int = None):
        """질문 로깅"""
        try:
            query_vector = self.embeddings.embed_query(query)
            query_id = hashlib.md5(
                f"{query}_{datetime.now().isoformat()}".encode()
            ).hexdigest()

            now = datetime.now()
            payload = {
                "query": query,
                "timestamp": now.isoformat(),
                "hour": hour if hour is not None else now.hour,          # 시간대별 분포
                "cache_hit": cache_hit,                                   # 캐시 히트 여부
                "team": team,
                "metadata": json.dumps(metadata or {}, ensure_ascii=False)
            }
            
            self.client.upsert(
                collection_name=QUERY_LOG_COLLECTION,
                points=[PointStruct(id=query_id, vector=query_vector, payload=payload)]
            )
            
            return query_id
            
        except Exception as e:
            logger.error(f"질문 로깅 실패: {e}")
            return None
    
    def log_answer(self, query: str, answer: str, sources: List[Dict],
                   metadata: Dict = None, team: str = "",
                   response_ms: int = 0, domain: str = "",
                   session_turn: int = 0, hour: int = None):
        """답변 로깅"""
        try:
            answer_vector = self.embeddings.embed_query(answer)
            answer_id = hashlib.md5(
                f"{answer}_{datetime.now().isoformat()}".encode()
            ).hexdigest()

            # 미해결 질문 감지
            UNANSWERED_PATTERNS = [
                "찾을 수 없습니다", "정보를 찾지 못했습니다", "담당자에게 직접 문의",
                "데이터가 없습니다", "일시적인 오류", "답변 생성에 실패",
                "관련 정보를 찾을 수 없습니다"
            ]
            is_unanswered = any(p in answer for p in UNANSWERED_PATTERNS)

            now = datetime.now()
            payload = {
                "query": query,
                "answer": answer,
                # page_content 제외 → 출처 메타정보만 저장 (용량 최적화)
                "sources": json.dumps([
                    {k: v for k, v in s.items() if k != "page_content"}
                    for s in (sources or [])
                ], ensure_ascii=False),
                "timestamp": now.isoformat(),
                "hour": hour if hour is not None else now.hour,          # 시간대별 분포
                "answer_length": len(answer),
                "source_count": len(sources),
                "team": team,
                "response_ms": response_ms,                              # 응답시간(ms)
                "domain": domain,                                        # 도메인 판별 결과
                "session_turn": session_turn,                            # 대화 턴 수
                "is_unanswered": is_unanswered,                         # 미해결 질문 여부
                "doc_click_count": 0,                                    # 📎 클릭 카운트 (초기값, 이후 갱신)
                "metadata": json.dumps(metadata or {}, ensure_ascii=False)
            }
            
            self.client.upsert(
                collection_name=ANSWER_LOG_COLLECTION,
                points=[PointStruct(id=answer_id, vector=answer_vector, payload=payload)]
            )

            return answer_id
            
        except Exception as e:
            logger.error(f"답변 로깅 실패: {e}")
            return None


class LearningSystem:
    """학습 시스템 + 이메일 알림"""
    def __init__(self, qdrant_client: QdrantClient, embeddings: OllamaEmbeddings, 
                 email_notifier: Optional[EmailNotifier] = None):
        self.client = qdrant_client
        self.embeddings = embeddings
        self.good_collection = LEARNING_COLLECTION
        self.bad_collection = BAD_FEEDBACK_COLLECTION
        self.email_notifier = email_notifier
        self._ensure_collections()
    
    def _ensure_collections(self):
        """학습용 컬렉션 생성"""
        try:
            collections = self.client.get_collections().collections
            existing_names = [c.name for c in collections]
            
            sample_vector = self.embeddings.embed_query("test")
            vector_size = len(sample_vector)
            
            if self.good_collection not in existing_names:
                self.client.create_collection(
                    collection_name=self.good_collection,
                    vectors_config=VectorParams(size=vector_size, distance=Distance.COSINE)
                )
                logger.info(f"✅ {self.good_collection} 생성 (학습용)")
            
            if self.bad_collection not in existing_names:
                self.client.create_collection(
                    collection_name=self.bad_collection,
                    vectors_config=VectorParams(size=vector_size, distance=Distance.COSINE)
                )
                logger.info(f"✅ {self.bad_collection} 생성 (학습용)")
                
        except Exception as e:
            logger.error(f"컬렉션 오류: {e}")
    
    def save_interaction(self, query: str, answer: str, sources: List[Dict], 
                        feedback_score: Optional[float] = None, team: str = ""):
        """👍 긍정 피드백 상호작용 저장 (따봉 클릭 시에만 호출)"""
        try:
            query_vector = self.embeddings.embed_query(query)
            
            interaction_id = hashlib.md5(
                f"{query}_{datetime.now().isoformat()}".encode()
            ).hexdigest()
            
            metadata = {
                "query": query,
                "answer": answer,
                # page_content 제외 → 출처 메타정보만 저장 (용량 최적화)
                "sources": json.dumps([
                    {k: v for k, v in s.items() if k != "page_content"}
                    for s in (sources or [])
                ], ensure_ascii=False),
                "timestamp": datetime.now().isoformat(),
                "feedback_score": feedback_score or 1.0,
                "usage_count": 1,
                "avg_quality": sum(s.get('score', 0) for s in sources) / len(sources) if sources else 0.0,
                "team": team,  # 부서 모드 (국내영업팀/해외영업팀/트랙영업팀)
            }
            
            self.client.upsert(
                collection_name=self.good_collection,
                points=[PointStruct(id=interaction_id, vector=query_vector, payload=metadata)]
            )
            logger.info(f"💾 긍정 피드백 저장: [{team}] {query[:30]}...")
            
        except Exception as e:
            logger.error(f"상호작용 저장 실패: {e}")
    
    def save_bad_feedback(self, query: str, answer: str, sources: List[Dict],
                          reason: str = "", team: str = ""):
        """
        부정 피드백 저장 + 이메일 알림 발송 (사유 포함)
        """
        try:
            query_vector = self.embeddings.embed_query(query)
            feedback_id = hashlib.md5(
                f"{query}_{datetime.now().isoformat()}_bad".encode()
            ).hexdigest()
            timestamp = datetime.now().isoformat()

            metadata = {
                "query": query,
                "answer": answer,
                # page_content 제외 → 출처 메타정보만 저장 (용량 최적화)
                "sources": json.dumps([
                    {k: v for k, v in s.items() if k != "page_content"}
                    for s in (sources or [])
                ], ensure_ascii=False),
                "timestamp": timestamp,
                "feedback_type": "bad",
                "reason": reason,
                "team": team,  # 부서 모드
            }

            self.client.upsert(
                collection_name=self.bad_collection,
                points=[PointStruct(id=feedback_id, vector=query_vector, payload=metadata)]
            )
            logger.info(f"💾 부정 피드백 저장: {query[:30]}... | 사유: {reason[:30]}")

            if self.email_notifier and self.email_notifier.enabled:
                feedback_data = {
                    "query": query,
                    "answer": answer,
                    "sources": sources,
                    "timestamp": timestamp,
                    "reason": reason,
                }
                email_thread = threading.Thread(
                    target=self.email_notifier.send_bad_feedback_notification,
                    args=(feedback_data,),
                    daemon=True
                )
                email_thread.start()

        except Exception as e:
            logger.error(f"Bad 피드백 실패: {e}")
    
    def search_similar_interactions(self, query: str, limit: int = 3, 
                                   min_score: float = 0.7) -> List[Dict]:
        """과거 유사 상호작용 검색"""
        try:
            query_vector = self.embeddings.embed_query(query)
            
            results = self.client.search(
                collection_name=self.good_collection,
                query_vector=query_vector,
                limit=limit,
                score_threshold=min_score
            )
            
            interactions = []
            for result in results:
                interactions.append({
                    'query': result.payload.get('query'),
                    'answer': result.payload.get('answer'),
                    'score': result.score,
                    'usage_count': result.payload.get('usage_count', 1)
                })
                
                # 사용 횟수 업데이트
                self.client.set_payload(
                    collection_name=self.good_collection,
                    payload={'usage_count': result.payload.get('usage_count', 1) + 1},
                    points=[result.id]
                )
            
            return interactions
            
        except Exception as e:
            logger.error(f"과거 데이터 검색 실패: {e}")
            return []
    
    def update_feedback(self, query: str, feedback_score: float, answer: str = "",
                        sources: List[Dict] = [], reason: str = "", team: str = ""):
        """피드백 업데이트 — 👍 긍정이면 learning_history 저장, 👎 부정이면 bad_feedback 저장"""
        try:
            if feedback_score >= 0.5:
                # 👍 따봉: learning_history에 새 포인트로 저장
                self.save_interaction(query, answer, sources,
                                      feedback_score=feedback_score, team=team)
            else:
                # 👎 부정 피드백 시 사유 포함 저장 + 이메일
                self.save_bad_feedback(query, answer, sources, reason=reason, team=team)
        except Exception as e:
            logger.error(f"피드백 업데이트 실패: {e}")



# ──────────────────────────────────────────────────────────────────────────────
# 시뮬레이터 조회 횟수 카운터 (Qdrant search_count 컬렉션)
# ──────────────────────────────────────────────────────────────────────────────
SEARCH_COUNT_COLLECTION = os.getenv("SEARCH_COUNT_COLLECTION", "search_count")

class SearchCountSystem:
    """
    시뮬레이터별 일자별 조회 횟수를 Qdrant search_count 컬렉션에 기록.
    포인트 ID = {simulator}_{YYYY-MM-DD} 의 MD5 해시 (정수)
    payload: simulator / date / count / last_updated / team
    벡터: 더미 1차원 (검색 불필요, 집계 전용)
    """
    SIMULATORS = [
        "국내운임비교",
        "국내최적배차",
        "컨베어벨트배차",
        "수출포장량",
        "크롤러배차",
    ]

    def __init__(self, qdrant_client: QdrantClient):
        self.client = qdrant_client
        self._ensure_collection()

    def _ensure_collection(self):
        try:
            existing = [c.name for c in self.client.get_collections().collections]
            if SEARCH_COUNT_COLLECTION not in existing:
                # 벡터 검색 불필요 → 1차원 더미 벡터
                self.client.create_collection(
                    collection_name=SEARCH_COUNT_COLLECTION,
                    vectors_config=VectorParams(size=1, distance=Distance.COSINE)
                )
                logger.info(f"✅ {SEARCH_COUNT_COLLECTION} 컬렉션 생성 (시뮬레이터 조회 카운터)")
        except Exception as e:
            logger.error(f"search_count 컬렉션 생성 실패: {e}")

    def _make_point_id(self, simulator: str, date_str: str) -> int:
        """simulator+날짜 → 고정 정수 ID (upsert 키로 사용)"""
        raw = f"{simulator}_{date_str}"
        return int(hashlib.md5(raw.encode()).hexdigest()[:15], 16)

    def increment(self, simulator: str, team: str = ""):
        """
        시뮬레이터 조회 버튼 클릭 시 호출.
        오늘 날짜 포인트가 있으면 count+1, 없으면 신규 생성.
        """
        try:
            today = datetime.now().strftime("%Y-%m-%d")
            point_id = self._make_point_id(simulator, today)

            # 기존 포인트 조회
            try:
                results = self.client.retrieve(
                    collection_name=SEARCH_COUNT_COLLECTION,
                    ids=[point_id],
                    with_payload=True
                )
            except Exception:
                results = []

            if results:
                current_count = results[0].payload.get("count", 0)
                self.client.set_payload(
                    collection_name=SEARCH_COUNT_COLLECTION,
                    payload={
                        "count": current_count + 1,
                        "last_updated": datetime.now().isoformat(),
                    },
                    points=[point_id]
                )
                logger.info(f"📊 [{simulator}] {today} count={current_count + 1}")
            else:
                self.client.upsert(
                    collection_name=SEARCH_COUNT_COLLECTION,
                    points=[PointStruct(
                        id=point_id,
                        vector=[0.0],  # 더미 벡터
                        payload={
                            "simulator": simulator,
                            "date": today,
                            "count": 1,
                            "last_updated": datetime.now().isoformat(),
                            "team": team,
                        }
                    )]
                )
                logger.info(f"📊 [{simulator}] {today} 신규 기록 count=1")
        except Exception as e:
            logger.error(f"search_count 기록 실패 [{simulator}]: {e}")

    def get_stats(self, simulator: str = None, days: int = 30) -> List[Dict]:
        """
        조회 통계 반환 (집계/대시보드용).
        simulator=None이면 전체 시뮬레이터.
        """
        try:
            from qdrant_client.models import Filter, FieldCondition, MatchValue, Range
            from datetime import timedelta

            since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
            conditions = [
                FieldCondition(key="date", range=Range(gte=since))
            ]
            if simulator:
                conditions.append(
                    FieldCondition(key="simulator", match=MatchValue(value=simulator))
                )

            results = self.client.scroll(
                collection_name=SEARCH_COUNT_COLLECTION,
                scroll_filter=Filter(must=conditions),
                limit=1000,
                with_payload=True
            )[0]

            return [
                {
                    "simulator": r.payload.get("simulator"),
                    "date": r.payload.get("date"),
                    "count": r.payload.get("count", 0),
                    "team": r.payload.get("team", ""),
                }
                for r in results
            ]
        except Exception as e:
            logger.error(f"search_count 통계 조회 실패: {e}")
            return []


def record_simulator_count(simulator: str, team: str = ""):
    """app.py에서 호출하는 퍼블릭 함수"""
    global SEARCH_COUNT_SYSTEM
    if SEARCH_COUNT_SYSTEM:
        SEARCH_COUNT_SYSTEM.increment(simulator, team)
    else:
        logger.warning("SEARCH_COUNT_SYSTEM 미초기화 — 카운트 건너뜀")


# ──────────────────────────────────────────────────────────────────────────────
# 시뮬레이터 Query + Result 상세 로그 (Qdrant simulator_logs 컬렉션)
# ──────────────────────────────────────────────────────────────────────────────
SIMULATOR_LOG_COLLECTION = os.getenv("SIMULATOR_LOG_COLLECTION", "simulator_logs")

class SimulatorLogSystem:
    """
    시뮬레이터 조회마다 Query(입력값) + Result(결과값)를 Qdrant에 기록.
    - 컬렉션: simulator_logs
    - 벡터: 더미 1차원 (집계 전용, 검색 불필요)
    - payload 공통: simulator / timestamp / date / hour / team
    - payload 가변: query_params(입력) / result_summary(결과 요약)
    """

    def __init__(self, qdrant_client: QdrantClient):
        self.client = qdrant_client
        self._ensure_collection()

    def _ensure_collection(self):
        try:
            existing = [c.name for c in self.client.get_collections().collections]
            if SIMULATOR_LOG_COLLECTION not in existing:
                self.client.create_collection(
                    collection_name=SIMULATOR_LOG_COLLECTION,
                    vectors_config=VectorParams(size=1, distance=Distance.COSINE)
                )
                logger.info(f"✅ {SIMULATOR_LOG_COLLECTION} 컬렉션 생성")
        except Exception as e:
            logger.error(f"simulator_logs 컬렉션 생성 실패: {e}")

    def log(self, simulator: str, query_params: dict, result_summary: dict,
            team: str = "", success: bool = True):
        """
        시뮬레이터 조회 1건을 로그에 기록.

        Args:
            simulator: 시뮬레이터 이름 (예: "국내운임비교")
            query_params: 입력값 dict
                예) {"destination": "창원", "weight_kg": 100}
                예) {"items": [{"code": "6008243", "length_m": 300, "rolls": 3}]}
            result_summary: 결과 요약 dict
                예) {"recommended": "화물", "직송": 150000, "화물": 120000}
                예) {"recommended_vehicle": "18톤", "total_weight_kg": 12915}
            team: 부서 모드
            success: 결과 정상 도출 여부
        """
        try:
            now = datetime.now()
            log_id = hashlib.md5(
                f"{simulator}_{now.isoformat()}_{json.dumps(query_params, ensure_ascii=False)}".encode()
            ).hexdigest()
            # MD5 → 정수 변환 (Qdrant ID는 정수 또는 UUID)
            log_id_int = int(log_id[:15], 16)

            payload = {
                "simulator"     : simulator,
                "timestamp"     : now.isoformat(),
                "date"          : now.strftime("%Y-%m-%d"),
                "hour"          : now.hour,
                "team"          : team,
                "success"       : success,
                "query_params"  : json.dumps(query_params,  ensure_ascii=False),
                "result_summary": json.dumps(result_summary, ensure_ascii=False),
            }

            self.client.upsert(
                collection_name=SIMULATOR_LOG_COLLECTION,
                points=[PointStruct(
                    id=log_id_int,
                    vector=[0.0],
                    payload=payload
                )]
            )
            logger.info(f"📋 [{simulator}] 로그 기록 완료 | 입력={list(query_params.keys())} | 결과={list(result_summary.keys())}")

        except Exception as e:
            logger.error(f"simulator_logs 기록 실패 [{simulator}]: {e}")


def record_simulator_log(simulator: str, query_params: dict, result_summary: dict,
                         team: str = "", success: bool = True):
    """
    app.py에서 시뮬레이터 결과 도출 후 호출하는 퍼블릭 함수.

    사용 예시:
        record_simulator_log(
            simulator="국내운임비교",
            query_params={"destination": "창원", "weight_kg": 100},
            result_summary={"recommended": "화물", "직송": 150000, "화물": 120000},
            team=st.session_state.get("selected_team", "")
        )
    """
    global SIMULATOR_LOG_SYSTEM
    if SIMULATOR_LOG_SYSTEM:
        SIMULATOR_LOG_SYSTEM.log(simulator, query_params, result_summary, team, success)
    else:
        logger.warning("SIMULATOR_LOG_SYSTEM 미초기화 — 로그 건너뜀")


def record_doc_click(answer_id: str):
    """
    📎 참고문서 버튼 클릭 시 호출.
    answer_logs 해당 포인트의 doc_click_count를 +1 갱신.
    """
    global LOGGING_SYSTEM
    if not LOGGING_SYSTEM or not answer_id:
        return
    try:
        results = LOGGING_SYSTEM.client.retrieve(
            collection_name=ANSWER_LOG_COLLECTION,
            ids=[answer_id],
            with_payload=True
        )
        if results:
            current = results[0].payload.get("doc_click_count", 0)
            LOGGING_SYSTEM.client.set_payload(
                collection_name=ANSWER_LOG_COLLECTION,
                payload={"doc_click_count": current + 1},
                points=[answer_id]
            )
            logger.info(f"📎 doc_click_count={current + 1} ({answer_id[:8]}...)")
    except Exception as e:
        logger.error(f"record_doc_click 실패: {e}")


def record_sim_inquiry(simulator: str, team: str = ""):
    """
    시뮬레이터 '문의하기' 버튼 클릭 시 호출.
    search_count 컬렉션의 해당 시뮬레이터 오늘 포인트에 inquiry_count +1.
    """
    global SEARCH_COUNT_SYSTEM
    if not SEARCH_COUNT_SYSTEM:
        return
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        point_id = SEARCH_COUNT_SYSTEM._make_point_id(simulator, today)
        results = SEARCH_COUNT_SYSTEM.client.retrieve(
            collection_name=SEARCH_COUNT_COLLECTION,
            ids=[point_id],
            with_payload=True
        )
        if results:
            current = results[0].payload.get("inquiry_count", 0)
            SEARCH_COUNT_SYSTEM.client.set_payload(
                collection_name=SEARCH_COUNT_COLLECTION,
                payload={"inquiry_count": current + 1},
                points=[point_id]
            )
            logger.info(f"📋 [{simulator}] inquiry_count={current + 1}")
        else:
            # 조회 기록 없이 문의만 한 경우 신규 생성
            SEARCH_COUNT_SYSTEM.client.upsert(
                collection_name=SEARCH_COUNT_COLLECTION,
                points=[PointStruct(
                    id=point_id,
                    vector=[0.0],
                    payload={
                        "simulator": simulator,
                        "date": today,
                        "count": 0,
                        "inquiry_count": 1,
                        "last_updated": datetime.now().isoformat(),
                        "team": team,
                    }
                )]
            )
    except Exception as e:
        logger.error(f"record_sim_inquiry 실패: {e}")



# query_processor.py에 추가할 함수들
# Qdrant와 무관하게 엑셀에서 직접 기사 노선 데이터 빌드

import os, re
from datetime import date, timedelta
from collections import defaultdict
from typing import Optional, Dict, List, Tuple

# 엑셀 경로 (환경변수 또는 기본 경로)
_EXCEL_PATHS = [
    os.getenv("LOGIBOT_EXCEL_PATH", ""),
    "data/source_docs/Logibot-Data(기본)_V5.xlsx",
    "Logibot-Data(기본)_V5.xlsx",
]

def _find_excel_path() -> Optional[str]:
    for p in _EXCEL_PATHS:
        if p and os.path.exists(p):
            return p
    return None

def _excel_time_to_str(v) -> str:
    if isinstance(v, float):
        total_min = round(v * 24 * 60)
        h, m = divmod(total_min, 60)
        return f"{h}시{m:02d}분" if m else f"{h}시"
    if v is None: return ""
    return str(v).strip()

def _clean_cell(v) -> str:
    if v is None: return ""
    s = str(v).strip().replace("\xa0", " ")
    s = re.sub(r"\n+", " | ", s)
    return re.sub(r"\s{2,}", " ", s)

def build_driver_context_from_excel() -> str:
    """
    엑셀에서 직접 기사 노선 context_text 빌드.
    Qdrant 조회 실패 시 fallback으로 사용.
    모든 기사(공통+A+B) 포함.
    """
    excel_path = _find_excel_path()
    if not excel_path:
        logger.warning("엑셀 파일을 찾을 수 없습니다 - LOGIBOT_EXCEL_PATH 환경변수 설정 필요")
        return ""

    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb['지입 차량(기사) 노선 데이터']
        rows = list(ws.iter_rows(values_only=True))

        groups: Dict[tuple, list] = defaultdict(list)
        for row in rows[1:]:
            if not row[0]: continue
            key = (_clean_cell(row[0]), _clean_cell(row[2]), _clean_cell(row[3]))
            groups[key].append(row)

        segments = []
        for (driver, weekday, route_group), group_rows in sorted(groups.items()):
            if not driver or not weekday: continue
            affil = _clean_cell(group_rows[0][1])
            stops = []
            for row in sorted(group_rows, key=lambda r: (r[4] or 0)):
                seq    = str(row[4] or "")
                region = _clean_cell(row[5])
                dest   = _clean_cell(row[6])
                t      = _excel_time_to_str(row[7])
                note   = _clean_cell(row[10])
                note   = re.sub(r'\s*\|\s*\(순서가 연속으로 중복된것\)', '', note)
                note   = "" if note == "고정운행 노선" else note
                stop_str = f"{seq}. ({region}) {dest} {t}"
                if note:
                    stop_str += f" ※{note}"
                stops.append(stop_str)

            route_note = ""
            if route_group in ("A", "B"):
                route_note = (f"\n※ 서울 기사는 격주 교대 운행 "
                              f"(이 데이터는 '{route_group}노선'입니다)")

            text = (f"[지입기사 납품 동선 | {driver} | {weekday} | 노선그룹:{route_group}]\n"
                    f"기사명: {driver} | 소속: {affil} | 요일: {weekday}\n"
                    f"납품 동선:\n" + "\n".join(stops) + route_note)
            segments.append(text)

        context = "\n\n".join(segments)
        logger.info(f"엑셀 직접 로드: 기사+요일 {len(segments)}개 세그먼트")
        return context

    except Exception as e:
        logger.error(f"엑셀 직접 로드 실패: {e}")
        return ""

def clean_cache():
    """캐시 정리"""
    global response_cache
    current_time = datetime.now()
    
    with cache_lock:
        expired_keys = [
            k for k, v in response_cache.items() 
            if (current_time - v['timestamp']).seconds > CACHE_TTL
        ]
        
        for key in expired_keys:
            del response_cache[key]


def setup_rag_chain():
    """RAG 체인 설정"""
    logger.info("RAG 체인 초기화...")
    
    llm = OnPremiseGemmaLLM()
    embeddings = OllamaEmbeddings(model=OLLAMA_EMBEDDING_MODEL, base_url=OLLAMA_HOST)
    qdrant_client = QdrantClient(url=f"http://{QDRANT_HOST}:{QDRANT_PORT}", api_key=QDRANT_API_KEY)
    vectorstore = Qdrant(client=qdrant_client, collection_name=QDRANT_COLLECTION, embeddings=embeddings)
    
    rag_chain = RAGChainWrapper(vectorstore, llm, embeddings, qdrant_client)
    logger.info("✅ RAG 체인 완료")
    return rag_chain


def search_ddg(query: str) -> str:
    """웹 검색"""
    try:
        results = DDGS().text(query, max_results=5, region='kr-kr')
        
        if not results:
            return ""
        
        context = []
        for i, r in enumerate(results, 1):
            context.append(f"[출처 {i}] {r['title']}\n{r['body']}")
        
        return "\n\n".join(context)
        
    except Exception as e:
        logger.error(f"웹 검색 오류: {e}")
        return ""


def format_answer(answer: str) -> str:
    """LaTeX 변환, 숫자 점수 제거, 마크다운 특수문자 정리"""
    if not answer:
        return ""

    # 유사도 점수 패턴 제거
    # RAG 내부 similarity score(0.XXX / 1.000)가 답변에 노출되는 현상 방지
    # 보호: 전화번호(010-...), 수식 결과(= 0.856), 표 내부
    # 제거 조건: 앞이 일반 공백(= 제외)이고 뒤가 공백 또는 줄끝인 단독 소수점
    answer = re.sub(r'(?<=[^=\d])\s(0\.\d{3}|1\.000)(?=\s|$)', ' ', answer, flags=re.MULTILINE)
    answer = re.sub(r'^(0\.\d{3}|1\.000)(?=\s|$)', '', answer, flags=re.MULTILINE)

    # LaTeX 변환
    answer = re.sub(r'\\frac\{(.+?)\}\{(.+?)\}', r'(\1 / \2)', answer)
    answer = answer.replace(r'\times', '×').replace(r'\cdot', '·')
    answer = answer.replace(r'\left\lfloor', '[').replace(r'\right\rfloor', ']')
    answer = answer.replace(r'\left(', '(').replace(r'\right)', ')')
    answer = answer.replace('[ ', '').replace(' ]', '')
    answer = answer.replace(r'\,', ' ')

    # ── 마크다운 특수문자 정리 ──────────────────────────────────────────
    # **(답변)** 같이 답변 구조 태그로 쓰인 ** 제거
    # 단, 표(|...|) 안에 있는 **bold**는 앱에서 HTML로 변환되므로 유지
    lines = answer.split('\n')
    cleaned = []
    for line in lines:
        # 표 행은 건드리지 않음
        if line.strip().startswith('|'):
            cleaned.append(line)
            continue
        # 헤더(##) 줄도 유지 (md_to_html_answer가 처리)
        if re.match(r'^#{1,3}\s', line.strip()):
            cleaned.append(line)
            continue
        # **(레이블)** 패턴: 줄 맨 앞에 오는 구조 태그 형태 제거
        # 예) **(답변)**, **(핵심)**, **(요약)** 등
        line = re.sub(r'^\s*\*\*\([^)]+\)\*\*\s*', '', line)
        # 줄 중간에 **텍스트:** 형태도 텍스트만 남김 (콜론 포함)
        line = re.sub(r'\*\*([^*]+):\*\*', r'\1:', line)
        # 나머지 모든 **텍스트** → 텍스트만 남김 (** 특수문자 화면 노출 방지)
        # 표/헤더 줄은 이미 위에서 continue 처리됐으므로 여기선 일반 텍스트만 대상
        line = re.sub(r'\*\*([^*\n]+)\*\*', r'\1', line)
        cleaned.append(line)
    answer = '\n'.join(cleaned)

    # 연속 빈 줄 2개 이상 → 1개로 압축
    answer = re.sub(r'\n{3,}', '\n\n', answer)

    return answer.strip()


def build_conversation_context(context_history: List[Dict]) -> str:
    """
    대화 맥락 구성.
    - 최근 5턴까지 참조 (기존 3턴)
    - 답변은 최대 300자 (기존 150자) → 맥락 정확도 향상
    - 자재코드·수치 등 핵심 키워드가 있으면 전체 보존
    """
    if not context_history:
        return ""

    # 최근 5턴 사용
    recent = context_history[-5:]
    context_str = "=== 이번 대화 히스토리 (최신순) ===\n"

    for idx, ctx in enumerate(reversed(recent), 1):
        q = ctx.get('query', '')
        a = ctx.get('answer', '')

        # 자재코드(7자리 숫자) 또는 수치 계산 포함 시 답변 전체 보존
        import re as _re
        has_code    = bool(_re.search(r'\b\d{7}\b', q + a))
        has_calc    = any(kw in q for kw in ['계산', '직경', '배차', '파렛트', '톤', '박스', '포장'])
        answer_limit = len(a) if (has_code or has_calc) else 300

        context_str += f"\n[{idx}번째 이전 대화]\n"
        context_str += f"사용자: {q}\n"
        context_str += f"AI 답변: {a[:answer_limit]}{'...(이하 생략)' if len(a) > answer_limit else ''}\n"

    context_str += "\n위 대화 히스토리를 반드시 참고하여 일관성 있게 답변하세요.\n"
    return context_str


def _format_driver_route_answer(query: str, context_text: str) -> str:
    """
    지입기사 납품 동선 답변 - Python 직접 포맷팅 (LLM 미사용).
    V5 텍스트 형식: "[지입기사 납품 동선 | 기사명 | 요일 | 노선그룹:xxx]"
    각 경유지: "순서. (권역) 거래처명 시간 ※특이사항"
    """
    import re as _re
    from datetime import date as _date, datetime as _dt, timedelta as _td

    # ── 0. 현재 위치 조회 감지 ───────────────────────────────────────────
    LOCATION_KW = ["현재 위치","지금 어디","현재 어디","어디 있","지금 위치","예상 위치","지금쯤","어디쯤"]
    is_location_query = any(k in query for k in LOCATION_KW)

    # ── 1. 범위 감지 ─────────────────────────────────────────────────────
    scope_busan  = any(k in query for k in ["부산","부산공장","부산 기사"])
    scope_seoul  = any(k in query for k in ["서울","중부","중부물류","수도권"])
    specific_name = next((n for n in ["김병일","김영철","이용구","심효섭"] if n in query), None)

    # ── 1.5 다음주 질문 감지 ─────────────────────────────────────────────
    NEXT_WEEK_KW = ["다음주","다음 주","내주","next week"]
    is_next_week_query = any(k in query for k in NEXT_WEEK_KW)

    # ── 2. 격주 판별 ─────────────────────────────────────────────────────
    BIWEEKLY_ANCHOR_DATE = _date(2026, 5, 25)  # 5/28 적용 시작주의 월요일
    _today  = _date.today()
    _mon    = _today - _td(days=_today.weekday())
    _weeks  = (_mon - BIWEEKLY_ANCHOR_DATE).days // 7
    _this_week_driver = "이용구" if _weeks % 2 == 0 else "심효섭"
    _next_week_driver = "심효섭" if _this_week_driver == "이용구" else "이용구"

    # 다음주 질문 + 서울 기사 이름 지정 시 → 다음주 담당 기사로 교체
    # 예) "다음주 이용구 기사 노선" (이번주 B주) → 심효섭 기사 노선 안내
    _next_week_redirect_msg = ""
    SEOUL_DRIVERS_LIST = ["이용구", "심효섭"]
    if is_next_week_query and specific_name in SEOUL_DRIVERS_LIST:
        _next_wg = "A" if _this_week_driver == "이용구" else "B"
        if specific_name == _this_week_driver:
            # 이번주 담당 기사의 다음주 노선 → 다음주엔 상대방이 담당
            _next_week_redirect_msg = (
                f"> 📅 다음 주({_next_wg}주)는 **{_next_week_driver} 기사**가 서울 납품을 담당합니다.\n"
                f"> {specific_name} 기사 대신 {_next_week_driver} 기사 노선을 안내드립니다.\n"
            )
            specific_name = _next_week_driver  # 다음주 담당으로 교체
        # 이미 다음주 담당 기사를 물어본 경우는 그대로

    # ── 3. stop_line 파싱 함수 ───────────────────────────────────────────
    def _parse_stop(line: str):
        """
        "6. (대저) 동양알앤비 14시05분 ※발주에 따른 변동운행 노선 | (순서가 연속으로 중복된것)"
        → (seq, region, dest, time_str, note)
        """
        # "(순서가 연속으로 중복된것)" 제거
        line = _re.sub(r'\s*\|\s*\(순서가 연속으로 중복된것\)', '', line)
        line = _re.sub(r'\(순서가 연속으로 중복된것\)', '', line).strip()

        seq_m = _re.match(r'^(\d+)\.\s*', line)
        seq = int(seq_m.group(1)) if seq_m else 0
        rest = line[seq_m.end():] if seq_m else line

        region_m = _re.match(r'\(([^)]+)\)\s*', rest)
        region = region_m.group(1) if region_m else ""
        rest = rest[region_m.end():] if region_m else rest

        note = ""
        note_m = _re.search(r'※(.+)$', rest)
        if note_m:
            note = note_m.group(1).strip()
            rest = rest[:note_m.start()].strip()

        time_m = _re.search(r'(\d+시(?:\d+분)?(?:\s*이후)?)\s*$', rest)
        time_str = time_m.group(1).strip() if time_m else ""
        dest = rest[:time_m.start()].strip() if time_m else rest.strip()

        return seq, region, dest, time_str, note

    # ── 4. context_text 파싱 (V5 형식) ──────────────────────────────────
    # {driver_name: {weekday: [(seq, region, dest, time, note), ...]}}
    from collections import defaultdict
    driver_day_stops: dict = defaultdict(lambda: defaultdict(list))

    HEADER_PAT = r'\[지입기사 납품 동선 \| (.+?) \| (.+?) \| 노선그룹:(.+?)\]'
    v5_headers = list(_re.finditer(HEADER_PAT, context_text))

    # 이번 주 격주 그룹 (A or B)
    _wg = "B" if _weeks % 2 == 0 else "A"
    # 기사별 route_group 저장 (출력 단계에서 이번주/다음주 구분에 사용)
    driver_route_group: dict = {}

    if v5_headers:
        for idx_h, m in enumerate(v5_headers):
            driver_name  = m.group(1).strip()
            weekday      = m.group(2).strip()
            route_grp    = m.group(3).strip()  # '공통', 'A', 'B'
            body_start   = m.end()
            body_end     = v5_headers[idx_h + 1].start() if idx_h + 1 < len(v5_headers) else len(context_text)
            body         = context_text[body_start:body_end]

            # route_group 저장 (기사명 → route_group 매핑)
            if driver_name not in driver_route_group:
                driver_route_group[driver_name] = route_grp

            in_stops = False
            for line in body.split('\n'):
                line = line.strip()
                if '납품 동선:' in line:
                    in_stops = True
                    continue
                if (in_stops and line
                        and not line.startswith('기사명:')
                        and not line.startswith('소속:')
                        and not line.startswith('요일:')
                        and not line.startswith('※ 서울')):
                    parsed = _parse_stop(line)
                    if parsed[2] and parsed[0] > 0:  # Fix: dest 있고 seq 유효(0 제외)
                        # 중복 방지: 동일 seq+dest가 이미 있으면 추가하지 않음
                        existing = driver_day_stops[driver_name][weekday]
                        if not any(e[0]==parsed[0] and e[2]==parsed[2] for e in existing):
                            existing.append(parsed)

    if not driver_day_stops:
        # 파싱 실패 시 specific_name 기준 필터링
        if specific_name:
            filtered = []
            capture = False
            for line in context_text.split('\n'):
                if specific_name in line and '지입기사 납품 동선' in line:
                    capture = True
                elif '[지입기사 납품 동선' in line and specific_name not in line:
                    capture = False
                if capture:
                    filtered.append(line)
            if filtered:
                return '\n'.join(filtered)
        return context_text

    # ── 5. 기본 정보 ─────────────────────────────────────────────────────
    BUSAN_DRIVERS = ["김병일", "김영철"]
    SEOUL_DRIVERS = ["이용구", "심효섭"]
    DAY_ORDER = {"월요일":0,"화요일":1,"수요일":2,"목요일":3,"금요일":4}
    DRIVER_INFO = {
        "김병일": {"tel":"010-3587-4581","car":"3.5톤 카고","area":"부산·경남권"},
        "김영철": {"tel":"010-7123-6231","car":"1톤 카고","area":"울산·마산·창원권"},
        "이용구": {"tel":"010-9263-4190","car":"2.5톤 카고","area":"서울·경기·인천권"},
        "심효섭": {"tel":"010-5291-6593","car":"2.5톤 카고","area":"서울 도심권"},
    }

    def _get_info(name_key: str) -> dict:
        for k, v in DRIVER_INFO.items():
            if k in name_key:
                return {"short": k, **v}
        return {"short": name_key, "tel":"-","car":"-","area":"-"}

    def _should_include(driver_key: str) -> bool:
        short = _get_info(driver_key)["short"]
        if specific_name:
            return specific_name in short
        if scope_busan and not scope_seoul:
            return short in BUSAN_DRIVERS
        if scope_seoul and not scope_busan:
            return short in SEOUL_DRIVERS
        return True

    # ── 6. 요일별 경유지 패턴 비교 → 묶기 ────────────────────────────────
    def _compress_days(day_stops: dict) -> list:
        """
        day_stops: {weekday: [(seq,region,dest,time,note), ...]}
        → 동일 패턴 요일 묶기
        반환: [(day_label, first_day, [(seq,region,dest,time,note), ...]), ...]
        """
        if not day_stops:
            return []

        def pat_key(stops):
            # 순서와 중복을 제거하고 내용(권역+거래처)만 비교
            return frozenset((r, d) for _, r, d, _, _ in stops)

        groups = defaultdict(list)
        for day, stops in day_stops.items():
            groups[pat_key(stops)].append(day)

        result = []
        for pat, days in groups.items():
            days_sorted = sorted(days, key=lambda d: DAY_ORDER.get(d, 99))
            idxs = [DAY_ORDER.get(d, 99) for d in days_sorted]
            consec = len(idxs) > 1 and all(idxs[i+1]-idxs[i]==1 for i in range(len(idxs)-1))

            if len(days_sorted) == 5:
                label = "월~금 (매일)"
            elif consec and len(days_sorted) >= 3:
                label = f"{days_sorted[0][:1]}~{days_sorted[-1][:1]}요일"
            elif consec and len(days_sorted) == 2:
                label = f"{'·'.join(d[:1] for d in days_sorted)}요일"
            else:
                # 단일 요일 or 비연속 → 각 요일명 그대로 사용
                label = "·".join(days_sorted)

            first_day = days_sorted[0]
            stops = day_stops[first_day]
            # (label, first_day, stops) 형식으로 저장 — 정렬 키로 first_day 사용
            result.append((label, first_day, stops))

        # first_day 기준으로 정렬 (월→화→수→목→금)
        result.sort(key=lambda x: DAY_ORDER.get(x[1], 99))
        return result

    # ── 7. 현재 위치 질문 처리 ───────────────────────────────────────────
    if is_location_query:
        target_name = specific_name
        if not target_name:
            if scope_seoul and not scope_busan:
                target_name = _this_week_driver
            else:
                return "현재 위치를 조회할 기사 이름을 포함해 질문해 주세요. (예: '김병일 기사 지금 어디?')"

        # 두 기사 모두 매주 운행 (노선만 격주 교대)
        # 이용구=B노선, 심효섭=A노선 → 항상 운행 중이므로 차단하지 않음
        # 다만 이번주 주담당 여부를 안내
        _is_main_this_week = (target_name in SEOUL_DRIVERS and target_name == _this_week_driver)
        _week_notice = (
            f"(이번 주 {_wg}주 주담당)"
            if _is_main_this_week
            else f"(이번 주 {_wg}주 보조노선 · 다음 주 주담당)"
        ) if target_name in SEOUL_DRIVERS else ""

        target_key = next((k for k in driver_day_stops if target_name in k), None)
        if not target_key:
            return f"{target_name} 기사의 노선 데이터를 찾을 수 없습니다."

        day_stops = driver_day_stops[target_key]
        now = _dt.now()
        weekday_names = ["월요일","화요일","수요일","목요일","금요일","토요일","일요일"]
        today_name = weekday_names[now.weekday()]
        if now.weekday() >= 5:
            return f"⛔ 오늘은 **{today_name}**입니다. {target_name} 기사는 주말에 운행하지 않습니다."
        today_stops = day_stops.get(today_name, [])
        if not today_stops:
            return f"오늘({today_name}) {target_name} 기사의 노선 데이터가 없습니다."
        now_min = now.hour * 60 + now.minute
        stops_timed = [(s, int(_re.search(r'(\d+)시(\d+)?', s[3]).group(1))*60 + int(_re.search(r'(\d+)시(\d+)?', s[3]).group(2) or 0))
                       for s in today_stops if _re.search(r'(\d+)시', s[3])]
        if not stops_timed:
            return f"{target_name} 기사의 오늘 납품 시간 정보를 파싱할 수 없습니다."
        first_min = stops_timed[0][1]
        last_min  = stops_timed[-1][1]
        if now_min < first_min:
            diff = first_min - now_min
            return (f"🕐 현재 **{now.hour:02d}:{now.minute:02d}** 기준\n\n"
                    f"아직 출발 전입니다. 첫 납품지 **{stops_timed[0][0][2]}** 까지 약 **{diff}분** 남았습니다.")
        if now_min > last_min + 30:
            return (f"🕐 현재 **{now.hour:02d}:{now.minute:02d}** 기준\n\n"
                    f"오늘 납품이 완료되었을 것으로 예상됩니다.")
        location_msg = ""
        for i, (s, smin) in enumerate(stops_timed):
            if now_min <= smin:
                location_msg = (f"🚗 **{s[2]}** {'납품 중' if now_min==smin else '이동 중'} "
                                f"(예정 **{s[3]}**, 약 {smin-now_min}분 후)")
                break
        if not location_msg:
            location_msg = f"📍 **{stops_timed[-1][0][2]}** 근처"
        _notice_line = f"\n> 📅 {_week_notice}" if _week_notice else ""
        return (f"🕐 현재 **{now.hour:02d}:{now.minute:02d}** ({today_name}) 기준\n\n"
                f"### {target_name} 기사 예상 위치\n{location_msg}\n\n"
                f"> ⚠️ 예상 위치는 납품 시각 기준 추정값입니다. 기사에게 직접 확인해 주세요."
                f"{_notice_line}")

    # ── 8. 일반 노선 조회 답변 생성 ──────────────────────────────────────
    lines = []

    if specific_name:
        if _next_week_redirect_msg:
            lines.append(_next_week_redirect_msg)
        lines.append(f"**{specific_name} 기사** 납품 동선 정보입니다.\n")
    elif scope_busan and not scope_seoul:
        lines.append("**부산공장 지입기사** 납품 동선 정보입니다.\n")
    elif scope_seoul and not scope_busan:
        lines.append("**서울(중부물류센터) 지입기사** 납품 동선 정보입니다.\n")
    else:
        lines.append("지입기사 전원의 납품 동선 정보입니다.\n")

    groups_to_show = []
    if specific_name:
        if specific_name in ["김병일", "김영철"]:
            groups_to_show.append(("🚚 부산공장 지입기사", ["김병일","김영철"]))
        else:
            groups_to_show.append(("🚌 서울(중부물류센터) 지입기사", ["이용구","심효섭"]))
    else:
        if not (scope_seoul and not scope_busan):
            groups_to_show.append(("🚚 부산공장 지입기사", ["김병일","김영철"]))
        if not (scope_busan and not scope_seoul):
            groups_to_show.append(("🚌 서울(중부물류센터) 지입기사", ["이용구","심효섭"]))

    for group_title, name_list in groups_to_show:
        is_seoul_group = ("이용구" in name_list)

        # driver_day_stops에서 해당 그룹 기사 추출 (_should_include 적용)
        group_drivers = [(dk, dv) for dk, dv in driver_day_stops.items()
                         if any(n in dk for n in name_list) and _should_include(dk)]

        if not group_drivers and is_seoul_group:
            # driver_day_stops 자체가 비어있거나 서울 기사 데이터 없음
            # → Qdrant 재적재 필요 안내
            lines.append(f"## {group_title}\n")
            lines.append(f"> ⚠️ Qdrant에서 노선 데이터를 가져오지 못했습니다.")
            lines.append(f"> `python data_loader.py ... --reset` 으로 재적재 후 다시 질문해 주세요.\n")
            continue
        elif not group_drivers:
            continue

        lines.append(f"## {group_title}\n")

        if is_seoul_group:
            lines.append(
                f"> 🗓️ **이번 주({_wg}주) 운행: {_this_week_driver} 기사** ｜ 다음 주: {_next_week_driver} 기사\n"
                f"> 두 기사는 격주로 교대 운행합니다.\n"
            )
            # 이번 주 기사 먼저 정렬
            group_drivers = sorted(group_drivers,
                                   key=lambda x: (0 if _this_week_driver in x[0] else 1))

        for driver_key, day_stops in group_drivers:
            info  = _get_info(driver_key)
            short = info["short"]
            drv_rg = driver_route_group.get(driver_key, "공통")

            # 서울 기사: 이번 주/다음 주 구분 태그
            if is_seoul_group:
                is_this_week = (drv_rg == _wg or drv_rg == "공통")
                if specific_name and not is_this_week:
                    # 명시적으로 다음 주 기사를 물어본 경우
                    week_tag = "  🔵 **다음 주 운행 노선** (미리보기)"
                elif specific_name and is_this_week:
                    week_tag = "  🟢 **이번 주 운행 노선**"
                elif is_this_week:
                    week_tag = "  🟢 **이번 주 운행**"
                else:
                    week_tag = "  🔵 다음 주 운행"
            else:
                week_tag = ""

            lines.append(f"### {short} 기사{week_tag}")
            lines.append(f"- 📱 **{info['tel']}**  |  🚛 {info['car']}  |  📍 {info['area']}")
            lines.append(f"- 운행: **월~금 (주 5일) 매일 운행**\n")

            # 요일 묶기
            compressed = _compress_days(day_stops)
            if not compressed:
                lines.append("_동선 데이터 없음_\n")
                continue

            for day_label, _first_day, stops in compressed:
                lines.append(f"**{day_label}** 납품 동선:\n")
                lines.append("| 순서 | 권역 | 거래처 | 납품시간 | 운행유형 |")
                lines.append("|:---:|:----:|------|:------:|:-----:|")
                for seq, region, dest, time_str, note in sorted([s for s in stops if s[0] > 0], key=lambda x:(x[0], x[3])):
                    run_type = "🔀 변동" if "변동" in note else "✅ 고정"
                    lines.append(f"| {seq} | {region} | {dest} | {time_str} | {run_type} |")
                lines.append("")

    if not specific_name:
        lines.append("---")
        lines.append("**💡 운행 특이사항**")
        if not (scope_seoul and not scope_busan):
            lines.append("- 부산공장 기사: 미성폴리머(김해)/신항 등 추가 운행은 오전에만 가능")
        if not (scope_busan and not scope_seoul):
            lines.append(f"- 서울 기사: 격주 교대 운행 (이번 주: **{_this_week_driver}** / 다음 주: {_next_week_driver})")
        lines.append("")

    lines.append("📞 납품 일정 변경·추가 문의는 **물류팀 담당자**에게 연락해 주세요.")
    return "\n".join(lines)



def process_query(query: str, rag_chain: RAGChainWrapper, learning_system: LearningSystem, 
                 logging_system: LoggingSystem, context: List[Dict] = None, team: str = "") -> Tuple[str, List[Dict], bool]:
    """쿼리 처리"""

    _start_ts = time.time()  # ① 응답시간 측정 시작
    _now      = datetime.now()
    _hour     = _now.hour
    _session_turn = len([c for c in (context or []) if c.get("role") == "user"])  # ⑥ 세션 턴 수

    # 로깅만 수행
    query_id = logging_system.log_query(
        query,
        metadata={"context_length": len(context or [])},
        team=team,
        hour=_hour,
        cache_hit=False  # 캐시 히트 여부는 아래에서 갱신
    )

    # 캐시 키: 질문 + 직전 대화 턴의 쿼리를 포함 → 맥락이 다르면 다른 답변
    prev_queries = "||".join(c.get("query", "") for c in (context or [])[-3:])
    cache_key = hashlib.md5(f"{query}|{prev_queries}".encode()).hexdigest()
    
    # 캐시 확인
    with cache_lock:
        if cache_key in response_cache:
            cached_data = response_cache[cache_key]
            if (datetime.now() - cached_data['timestamp']).seconds < CACHE_TTL:
                logger.info("✅ 캐시 응답")
                # ② 캐시 히트 → query_log 갱신
                try:
                    if query_id:
                        logging_system.client.set_payload(
                            collection_name=QUERY_LOG_COLLECTION,
                            payload={"cache_hit": True},
                            points=[query_id]
                        )
                except Exception:
                    pass
                return cached_data['answer'], cached_data['sources'], cached_data.get('has_table', False)
    
    # ── 시뮬레이터 유도 체크 ─────────────────────────────────────
    # ── 중량 기반 차량 톤수 직접 계산 ──────────────────────────────────
    # 질문에 kg/ton 중량이 명시된 경우 차량 데이터 기준으로 즉시 답변
    import re as _re
    _WEIGHT_KG_PAT  = _re.compile(r'(\d+(?:\.\d+)?)\s*(?:KG|kg|킬로그램)', _re.IGNORECASE)
    _WEIGHT_TON_PAT = _re.compile(r'(\d+(?:\.\d+)?)\s*(?:톤|TON|ton)(?!\s*차량|\s*트럭|\s*짜리\s*차|\s*급\s*차)', _re.IGNORECASE)
    _PLT_CAR_KW = ["PLT","plt","파렛트","팔레트","차량","배차","몇 톤","몇톤","운송","운반"]
    _is_car_query = any(k in query for k in _PLT_CAR_KW)

    _weight_kg = None
    _kg_hit  = _WEIGHT_KG_PAT.search(query)
    _ton_hit = _WEIGHT_TON_PAT.search(query)
    if _kg_hit:
        _weight_kg = float(_kg_hit.group(1))
    elif _ton_hit and _is_car_query:
        _weight_kg = float(_ton_hit.group(1)) * 1000

    if _weight_kg and _is_car_query:
        _wt = _weight_kg / 1000
        _CAR_TABLE = [
            (1.32,  "1톤(1.2톤)",  "~1.32톤"),
            (2.75,  "2.5톤",       "1.33~2.75톤"),
            (3.85,  "3.5톤",       "2.76~3.85톤"),
            (5.50,  "5톤",         "3.86~5.5톤"),
            (8.20,  "8톤",         "5.6~8.2톤"),
            (12.0,  "11톤",        "8.3~12톤"),
            (19.0,  "18톤",        "12.1~19톤"),
            (25.5,  "25톤",        "19.1~25.5톤"),
            (float("inf"), "트레일러", "25.5톤 초과"),
        ]
        _rec_car   = next((car for lim,car,_ in _CAR_TABLE if _wt<=lim), "트레일러")
        _rec_range = next((rng for lim,_,rng in _CAR_TABLE if _wt<=lim), "")
        _wt_answer = (
            f"**총 중량 {_weight_kg:,.0f}kg({_wt:.3f}톤)** 기준으로 안내드립니다.\n\n"
            f"| 항목 | 내용 |\n|---|---|\n"
            f"| 📦 입력 중량 | {_weight_kg:,.0f}kg ({_wt:.3f}톤) |\n"
            f"| 🚛 권장 차량 | **{_rec_car}** |\n"
            f"| 📋 해당 중량 범위 | {_rec_range} |\n\n"
            f"※ 중량 외 **부피(길이×폭×높이)** 도 차량 선택에 영향을 줍니다. "
            f"정확한 배차는 담당자에게 문의하시거나 좌측 **[시뮬레이터]** 를 이용해주세요."
        )
        logging_system.log_answer(
            query=query, answer=_wt_answer, sources=[],
            metadata={"weight_calc": True, "weight_kg": _weight_kg},
            team=team, response_ms=int((time.time()-_start_ts)*1000),
            domain="weight_calc", session_turn=_session_turn, hour=_hour,
        )
        return _wt_answer, [], False
    # ─────────────────────────────────────────────────────────────────

    guide_msg = check_simulator_intent(query)
    if guide_msg:
        logging_system.log_answer(
            query=query, answer=guide_msg, sources=[],
            metadata={"simulator_guide": True}, team=team,
            response_ms=int((time.time() - _start_ts) * 1000),
            domain="simulator_guide", session_turn=_session_turn, hour=_hour,
        )
        return guide_msg, [], False
    # ─────────────────────────────────────────────────────────────

    # learning_history에서 유사 상호작용 검색
    similar_interactions = learning_system.search_similar_interactions(query, limit=2, min_score=0.85)
    
    history_context = ""
    if similar_interactions:
        history_context = "참고: 과거 유사 질문 (learning_history)\n"
        for idx, interaction in enumerate(similar_interactions, 1):
            history_context += f"\n[{idx}] {interaction['query']}\n{interaction['answer'][:200]}...\n"
    
    # 대화 맥락 구성
    conversation_context = build_conversation_context(context or [])
    
    answer = ""
    sources = []
    use_web_search = False
    has_table = False
    
    table_keywords = ["표로", "표 형식", "테이블", "표를", "표 만들"]
    is_table_request = any(kw in query for kw in table_keywords)
    search_k = 7  # V4 동일: 기본 검색 문서 수
    list_keywords = ["몇 개", "전부", "목록", "리스트", "어디", "전체", "모두"]
    if any(kw in query for kw in list_keywords):
        search_k = 15  # V4 동일: 목록 질문은 더 넓게

    # 웹서치를 허용할 명시적 키워드 (사내 데이터에 없는 외부 정보가 필요한 경우)
    WEB_ALLOWED_KW = [
        "최신", "뉴스", "시세", "환율", "날씨", "법률", "규정", "관세",
        "HS코드", "incoterms", "인코텀즈", "선적 서류", "수출 서류",
        "B/L", "invoice", "packing list", "원산지증명", "세관"
    ]
    _allow_web = any(kw.lower() in query.lower() for kw in WEB_ALLOWED_KW)

    try:
        filtered_docs = rag_chain.hybrid_search(query, k=search_k)

        if not filtered_docs:
            # 로컬 문서 없을 때: 웹서치 허용 키워드 있으면 웹서치, 없으면 LLM 자체 지식으로 답변
            if _allow_web:
                use_web_search = True
            else:
                # LLM 자체 지식으로 답변 시도
                fallback_prompt = rag_chain.prompt_template.format(
                    context="관련 내부 데이터가 없습니다. 일반 물류 지식을 바탕으로 답변하세요.",
                    input=query,
                    history_context=history_context if history_context else "없음",
                    conversation_context=conversation_context if conversation_context else "없음"
                )
                try:
                    answer = rag_chain.llm._call(fallback_prompt)
                except Exception:
                    answer = "해당 정보를 내부 데이터에서 찾을 수 없습니다. 담당자에게 직접 문의해 주세요."
        else:
            context_text = "\n\n---\n\n".join([
                doc.page_content
                for doc, _ in filtered_docs
            ])

            sources = [{
                'name': (
                    doc.metadata.get('file_name')
                    or os.path.basename(doc.metadata.get('source', ''))
                    or doc.metadata.get('sheet_name', '검색 결과')
                ),
                'page_content': doc.page_content,
                'sheet_name': doc.metadata.get('sheet_name', ''),
                'internal_score': float(score),
                'page': doc.metadata.get('page', 'N/A')
            } for doc, score in filtered_docs]

            _domain_for_limit = rag_chain._detect_domain(query)
            if _domain_for_limit in ("driver_route", "operation_rule"):
                _ctx_limit = 14000  # summary(~7000자) + 개별 Q&A 청크 여유분
            elif _domain_for_limit in ("general", "personnel"):
                _ctx_limit = 8000  # Fix: 지게차 기사 등 운영규칙+현황 동시 참조 필요
            else:
                _ctx_limit = 4000
            if len(context_text) > _ctx_limit:
                context_text = context_text[:_ctx_limit] + "\n..."

            if is_table_request:
                has_table = True

            # ── driver_route: LLM으로 의도 먼저 분류 → 처리 방식 결정 ─────
            if _domain_for_limit == "driver_route":

                # 1단계: LLM에게 질문 의도 분류 요청 (경량 호출)
                intent_prompt = f"""다음 질문의 의도를 아래 3가지 중 하나로만 답하세요. 다른 말은 절대 하지 마세요.

질문: "{query}"

선택지:
- ROUTE_LOOKUP : 특정 기사의 납품 동선/노선/일정을 조회하는 질문
- COMPARE_OR_CONFIRM : 비교, 확인, 예/아니오, 의견, 데이터에 없는 정보를 묻는 질문
- GENERAL_INFO : 연락처, 소속, 차량 정보, 거래처 납품 문의처 등 기사 기본 정보 질문
  ※ "거래처에 납품하려는데 누구에게 문의", "거래처 담당자" 같은 질문은 GENERAL_INFO

답:"""
                try:
                    intent_raw = rag_chain.llm._call(intent_prompt).strip().upper()
                    if "ROUTE_LOOKUP" in intent_raw:
                        intent = "ROUTE_LOOKUP"
                    elif "GENERAL_INFO" in intent_raw:
                        intent = "GENERAL_INFO"
                    else:
                        intent = "COMPARE_OR_CONFIRM"
                except Exception:
                    # LLM 분류 실패 시 키워드 폴백
                    COMPARE_KW = [
                        "같은 노선", "동일한 노선", "항상 같", "차이가 있",
                        "다른가요", "다른지", "비교", "맞나요", "하나요",
                        "인가요", "할까요", "몇 명", "겹치는", "효율",
                    ]
                    intent = "COMPARE_OR_CONFIRM" if any(k in query for k in COMPARE_KW) else "ROUTE_LOOKUP"
                    logger.warning(f"의도 분류 LLM 실패 → 키워드 폴백: {intent}")

                logger.info(f"driver_route 의도 분류: {intent} | 질문: {query[:40]}")

                if intent == "ROUTE_LOOKUP":
                    # 노선 조회 → Python 포맷팅 (기존 방식)
                    answer = _format_driver_route_answer(query, context_text)
                elif intent == "GENERAL_INFO":
                    # 기사 기본정보(연락처/소속/차량) → DRIVER_INFO에서 직접 추출
                    DRIVER_INFO_MAP = {
                        "김병일": {"tel":"010-3587-4581","car":"3.5톤 카고","area":"부산·경남권","base":"부산공장"},
                        "김영철": {"tel":"010-7123-6231","car":"1톤 카고",  "area":"울산·마산·창원권","base":"부산공장"},
                        "이용구": {"tel":"010-9263-4190","car":"2.5톤 카고","area":"서울·경기·인천권","base":"중부물류센터"},
                        "심효섭": {"tel":"010-5291-6593","car":"2.5톤 카고","area":"서울 도심권","base":"중부물류센터"},
                    }
                    target = next((n for n in DRIVER_INFO_MAP if n in query), None)
                    if target:
                        d = DRIVER_INFO_MAP[target]
                        answer = (
                            f"**{target} 기사** 기본 정보입니다.\n\n"
                            f"| 항목 | 내용 |\n|---|---|\n"
                            f"| 📱 연락처 | {d['tel']} |\n"
                            f"| 🚛 차량 | {d['car']} |\n"
                            f"| 📍 담당 권역 | {d['area']} |\n"
                            f"| 🏭 소속 | {d['base']} |"
                        )
                    else:
                        # ── 거래처 문의 → 납품 기사 소속별 담당자 안내 ──
                        _CUSTOMER_DRIVER = {
                            # 부산공장 기사 (김영철/김병일) 납품 거래처
                            "GM대우 KD":"김영철","동양 ENG":"김영철","동일벨트":"김영철",
                            "동일종합산업":"김영철","동일종합상사":"김영철","동일팬유통":"김영철",
                            "동일상사":"김영철","모던테크":"김영철","승리상사":"김영철",
                            "한국벨트":"김영철","현대내자":"김영철","현대테크젠":"김영철",
                            "동양알앤비":"김병일","대신화물":"김병일","동일삼광산업":"김병일",
                            "동일상공사":"김병일","동일알앤씨":"김병일","동일종합물산":"김병일",
                            "반도상사":"김병일","성우상사":"김병일","유일상사":"김병일",
                            # 중부물류센터 기사 (이용구/심효섭) 납품 거래처
                            "동일벨트산업":"이용구","동일부품":"이용구","동일팬벨트":"이용구",
                            "명진":"이용구","서울팬벨트":"이용구","서울팬벨트(하남)":"이용구",
                            "인왕산업":"이용구","한길산업":"이용구","흥진사":"이용구",
                            "삼흥정밀":"심효섭","성보산업사":"심효섭","신우기업":"심효섭",
                            "신흥알앤테크":"심효섭","신흥폴리테크":"심효섭","유니온벨티노":"심효섭",
                            "제일산업사":"심효섭","조선통상":"심효섭","조일상공":"심효섭",
                            "태영산업사":"심효섭","흥국상사":"심효섭",
                        }
                        # 공백 제거 후 거래처 매칭 (긴 이름 우선 → "동일벨트산업"이 "동일벨트"보다 먼저)
                        _q_ns = query.replace(" ","")
                        _matched_customer = next(
                            (c for c in sorted(_CUSTOMER_DRIVER, key=len, reverse=True)
                             if c.replace(" ","") in _q_ns), None
                        )
                        if _matched_customer:
                            _driver = _CUSTOMER_DRIVER[_matched_customer]
                            _d_info = DRIVER_INFO_MAP.get(_driver, {})
                            if _driver in ["김영철","김병일"]:
                                _contact = (
                                    f"**{_matched_customer}** 납품은 **{_driver} 기사** ({_d_info.get('area','')})가 담당합니다.\n\n"
                                    f"제품 관련 문의는 아래 담당자에게 연락해 주세요.\n"
                                    f"- 김동우 팀원 (내선: 9133) — 전동내수 담당\n"
                                    f"- 이정희 주임 (내선: 9341) — 전동내수 공정 직책자"
                                )
                            else:
                                _contact = (
                                    f"**{_matched_customer}** 납품은 **{_driver} 기사** ({_d_info.get('area','')})가 담당합니다.\n\n"
                                    f"제품 관련 문의는 아래 담당자에게 연락해 주세요.\n"
                                    f"- 이정환 팀원 (내선: 9834) — 중부물류센터 담당\n"
                                    f"- 오승현 팀원 (내선: 9832) — 중부물류센터 담당"
                                )
                            answer = _contact
                            logging_system.log_answer(
                                query=query, answer=answer, sources=[],
                                metadata={"customer_lookup": True},
                                team=team, response_ms=int((time.time()-_start_ts)*1000),
                                domain="driver_route", session_turn=_session_turn, hour=_hour,
                            )
                            return answer, [], False

                        # 거래처 매칭 없음 → scope 기반 필터링
                        _scope_busan = any(k in query for k in ["부산","부산공장","부산 기사"])
                        _scope_seoul = any(k in query for k in ["서울","중부","중부물류","수도권"])

                        if _scope_busan and not _scope_seoul:
                            _filtered = {n:d for n,d in DRIVER_INFO_MAP.items() if d["base"] == "부산공장"}
                            _title = "**부산공장 지입기사 기본 정보**"
                        elif _scope_seoul and not _scope_busan:
                            _filtered = {n:d for n,d in DRIVER_INFO_MAP.items() if d["base"] == "중부물류센터"}
                            _title = "**서울(중부물류센터) 지입기사 기본 정보**"
                        else:
                            _filtered = DRIVER_INFO_MAP
                            _title = "**지입기사 전체 기본 정보**"

                        lines_info = [f"{_title}\n",
                                      "| 기사명 | 연락처 | 차량 | 권역 | 소속 |",
                                      "|---|---|---|---|---|"]
                        for name, d in _filtered.items():
                            lines_info.append(f"| {name} | {d['tel']} | {d['car']} | {d['area']} | {d['base']} |")
                        answer = "\n".join(lines_info)
                else:
                    # COMPARE_OR_CONFIRM → LLM 답변
                    _d_prompt = get_domain_prompt("driver_route")
                    formatted_prompt = _d_prompt.format(
                        context=context_text,
                        input=query,
                        history_context=history_context if history_context else "없음",
                        conversation_context=conversation_context if conversation_context else "없음"
                    )
                    answer = rag_chain.llm._call(formatted_prompt)

                has_table = "|" in answer and "---" in answer
            else:
                # 도메인별 프롬프트 적용
                _d_prompt = get_domain_prompt(_domain_for_limit)
                formatted_prompt = _d_prompt.format(
                    context=context_text,
                    input=query,
                    history_context=history_context if history_context else "없음",
                    conversation_context=conversation_context if conversation_context else "없음"
                )
                answer = rag_chain.llm._call(formatted_prompt)

            if "|" in answer and "---" in answer:
                has_table = True

            # 웹서치 전환 조건: 답변이 완전히 비어있고 + 웹허용 키워드 있을 때만
            if (not answer or len(answer.strip()) < 10) and _allow_web:
                use_web_search = True

    except Exception as e:
        import traceback
        logger.error(f"RAG 검색/답변 실패: {e}\n{traceback.format_exc()}")
        # 예외 발생 시에도 웹서치 허용 키워드 없으면 웹서치 건너뜀
        if _allow_web:
            use_web_search = True
        else:
            answer = "일시적인 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."
    
    # 웹 검색 (WEB_ALLOWED_KW에 해당하는 질문만 여기 도달)
    if use_web_search:
        logger.info(f"🌐 웹서치 발동: '{query[:40]}' (허용 키워드 매칭)")

        # driver_route는 Python 포맷팅으로 처리되므로 여기 도달하면 안 됨 (안전망)
        _domain_check = rag_chain._detect_domain(query)
        if _domain_check == "driver_route":
            _test_docs = rag_chain.fetch_whole_docs(["지입 차량(기사) 노선 데이터"], limit=1)
            if _test_docs:
                return (
                    "지입기사 납품 동선 데이터는 찾았으나 답변 생성 중 오류가 발생했습니다.\n\n"
                    "잠시 후 다시 시도해 주세요.",
                    [], False
                )
            else:
                return (
                    "지입기사 납품 동선 데이터를 Qdrant에서 찾을 수 없습니다.\n\n"
                    "`python data_loader.py`로 재인덱싱 후 다시 질문해 주세요.",
                    [], False
                )

        web_context = search_ddg(query)
        sources = [{'name': 'Web Search', 'score': 0.5, 'page': 'N/A'}]
        
        if web_context:
            enhanced_prompt = f"""당신은 DRB 물류 전문 AI 어시스턴트입니다.
아래 웹 검색 결과를 참고하여 질문에 답변하세요.
내부 데이터에 없는 내용이므로 웹 검색 결과를 바탕으로 친절하게 안내하세요.

[웹 검색 결과]
{web_context}

[질문]
{query}

{"⚠️ 표 형식으로 작성하세요." if is_table_request else ""}

답변:
"""
            try:
                orig_timeout = rag_chain.llm.timeout
                rag_chain.llm.timeout = 120
                answer = rag_chain.llm._call_with_max_tokens(enhanced_prompt, max_tokens=3000)
                rag_chain.llm.timeout = orig_timeout
                answer += "\n\n🌐 웹 검색 기반"
                if "|" in answer and "---" in answer:
                    has_table = True
            except Exception as e:
                logger.error(f"웹서치 LLM 호출 실패: {e}")
                answer = "답변 생성에 실패했습니다. 잠시 후 다시 시도해 주세요."
        else:
            answer = "관련 정보를 찾을 수 없습니다. 담당자에게 직접 문의해 주세요."
    
    # learning_history 저장은 👍 따봉 버튼 클릭 시에만 (submit_feedback → update_feedback 경로)
    # 여기서 저장하지 않음
    
    # 답변 포맷팅
    answer = format_answer(answer)

    # ① 응답시간 계산
    _response_ms = int((time.time() - _start_ts) * 1000)

    # ③ 최종 도메인 판별 (use_web_search면 "web", 아니면 rag_chain에서 판별)
    _final_domain = "web" if use_web_search else getattr(rag_chain, "_last_domain", "")
    if not _final_domain:
        try:
            _final_domain = rag_chain._detect_domain(query)
        except Exception:
            _final_domain = "unknown"

    # 로깅
    _answer_id = logging_system.log_answer(
        query=query,
        answer=answer,
        sources=sources,
        metadata={"has_table": has_table, "use_web_search": use_web_search},
        team=team,
        response_ms=_response_ms,
        domain=_final_domain,
        session_turn=_session_turn,
        hour=_hour,
    )
    
    # 캐시 저장
    with cache_lock:
        response_cache[cache_key] = {
            'answer': answer,
            'sources': sources,
            'has_table': has_table,
            'timestamp': datetime.now(),
            'answer_id': _answer_id,   # 📎 클릭 시 doc_click_count 갱신용
        }
    
    if len(response_cache) > 100:
        clean_cache()

    # answer_id를 rag_chain에 임시 보관 → get_rag_response에서 반환값으로 전달
    try:
        rag_chain._last_answer_id = _answer_id or ""
    except Exception:
        pass

    return answer, sources, has_table


# 초기화
try:
    RAG_CHAIN = setup_rag_chain()
    EMAIL_NOTIFIER = EmailNotifier()
    LEARNING_SYSTEM = LearningSystem(RAG_CHAIN.qdrant_client, RAG_CHAIN.embeddings, EMAIL_NOTIFIER)
    LOGGING_SYSTEM = LoggingSystem(RAG_CHAIN.qdrant_client, RAG_CHAIN.embeddings)
    SEARCH_COUNT_SYSTEM = SearchCountSystem(RAG_CHAIN.qdrant_client)
    SIMULATOR_LOG_SYSTEM = SimulatorLogSystem(RAG_CHAIN.qdrant_client)
    logger.info("✅ 시스템 완료")
    logger.info("=" * 60)
    logger.info("📊 컬렉션 용도:")
    logger.info("  - logistics_data: 메인 문서 (답변 생성 시 사용)")
    logger.info("  - learning_history: 👍 긍정 피드백 데이터 (따봉 클릭 시에만 저장)")
    logger.info("  - bad_feedback_history: 부정 피드백 데이터 (참고용 + 이메일 알림)")
    logger.info("  - query_logs: 질문 로그 (로깅 전용)")
    logger.info("  - answer_logs: 답변 로그 (로깅 전용)")
    logger.info("  - search_count: 시뮬레이터 일자별 조회 횟수")
    logger.info("  - simulator_logs: 시뮬레이터 Query+Result 상세 로그")
    logger.info("=" * 60)
    if EMAIL_NOTIFIER.enabled:
        logger.info(f"📧 이메일 알림: 활성화 → {', '.join(EMAIL_NOTIFIER.email_to)}")
    else:
        logger.info("📧 이메일 알림: 비활성화")
    logger.info("=" * 60)
except Exception as e:
    logger.exception("❌ 초기화 실패")
    RAG_CHAIN = None
    LEARNING_SYSTEM = None
    LOGGING_SYSTEM = None
    EMAIL_NOTIFIER = None
    SEARCH_COUNT_SYSTEM = None
    SIMULATOR_LOG_SYSTEM = None


def get_rag_response(query: str, context: List[Dict] = None, team: str = "") -> Dict[str, Any]:
    """RAG 응답 생성"""
    global RAG_CHAIN, LEARNING_SYSTEM, LOGGING_SYSTEM
    
    answer, sources, has_table = process_query(query, RAG_CHAIN, LEARNING_SYSTEM, LOGGING_SYSTEM, context, team=team)

    # 출처 정보 구성 (중복 제거)
    unique_sources_data = []
    seen = set()
    for s in (sources or []):
        if isinstance(s, dict):
            meta = s.get('metadata', {})
            text = s.get('page_content', s.get('content', ''))
        else:
            meta = getattr(s, 'metadata', {})
            text = getattr(s, 'page_content', '')

        f_name = meta.get('file_name', meta.get('name', '참고 문서'))
        f_name = re.sub(r'\s?\d\.\d{3}\b', '', str(f_name)).strip()

        if f_name and text and f_name not in seen:
            unique_sources_data.append({"name": f_name, "content": text})
            seen.add(f_name)

    return {
        "answer": answer,
        "sources": unique_sources_data[:3],
        "has_table": has_table,
        "cached": False,
        "answer_id": getattr(RAG_CHAIN, "_last_answer_id", "") if RAG_CHAIN else "",
    }


def submit_feedback(query: str, feedback_score: float, answer: str = "",
                    sources: List[Dict] = [], reason: str = "", team: str = ""):
    """
    피드백 제출 (부정 피드백 시 사유 포함 자동 이메일 발송)
    - feedback_score >= 0.5 : 👍 긍정 → learning_history 저장
    - feedback_score < 0.5  : 👎 부정 → bad_feedback_history 저장 + 이메일
    """
    global LEARNING_SYSTEM
    if LEARNING_SYSTEM:
        LEARNING_SYSTEM.update_feedback(query, feedback_score, answer, sources, reason=reason, team=team)
    else:
        logger.error("❌ LEARNING_SYSTEM 미초기화")
        
def analyze_logistics_data(df: pd.DataFrame) -> str:
    try:
        # 중량 관련 컬럼 탐색 (대소문자 무시)
        weight_col = [c for c in df.columns if '중량' in str(c) or 'weight' in str(c).lower()]
        if not weight_col:
            return "❌ 엑셀 파일 내에 '중량' 관련 컬럼을 찾을 수 없습니다."
        
        total_weight = df[weight_col[0]].sum()
        
        # 배차 로직 (단순화)
        if total_weight <= 1000: truck = "1톤 트럭"
        elif total_weight <= 5000: truck = "5톤 카고"
        elif total_weight <= 11000: truck = "11톤 윙바디"
        else: truck = "25톤 트레일러"
            
        return f"📊 **분석 결과**: 합계 중량 {total_weight:,.1f}kg, 추천 차량은 **{truck}** 입니다."
    except Exception as e:
        return f"❌ 엑셀 분석 중 오류: {str(e)}"

def analyze_pdf_logistics(file_bytes):
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        full_text = ""
        for page in doc:
            full_text += page.get_text()
        
        # 중량 숫자 추출 (정규표현식: 중량 뒤의 숫자와 단위 매칭)
        weights = re.findall(r'(?:중량|Weight)\s*[:：]?\s*([\d,.]+)', full_text)
        if weights:
            total_w = sum([float(w.replace(',', '')) for w in weights])
            return f"📄 **PDF 분석 완료**: 감지된 총 중량은 **{total_w:,.1f}kg**입니다. 배차 분석을 진행할까요?"
        
        return "📄 PDF 텍스트는 추출했으나, '중량' 키워드와 수치를 찾지 못했습니다."
    except Exception as e:
        return f"❌ PDF 처리 실패: {str(e)}"
    
def _calc_loadable_plt(plt_w: float, plt_l: float, car_w: float, car_l: float) -> int:
    """
    파렛트(plt_w × plt_l)를 차량 적재함(car_w × car_l)에
    가로·세로 두 방향 + 혼합 배열로 최대 몇 개 올릴 수 있는지 계산. 단위: m
    """
    best = 0
    orientations = [(plt_w, plt_l), (plt_l, plt_w)]

    # 단일 방향
    for pw, pl in orientations:
        if pw <= car_w:
            cols = int(car_w / pw)
            rows = int(car_l / pl)
            best = max(best, cols * rows)

    # 혼합 배열: 앞 구간 방향1, 나머지 구간 방향2
    for pw1, pl1 in orientations:
        for pw2, pl2 in orientations:
            if (pw1, pl1) == (pw2, pl2):
                continue
            cols1 = int(car_w / pw1) if pw1 <= car_w else 0
            cols2 = int(car_w / pw2) if pw2 <= car_w else 0
            if cols1 == 0 or cols2 == 0:
                continue
            for n1 in range(1, int(car_l / pl1) + 1):
                rem = car_l - n1 * pl1
                if rem < 0:
                    break
                n2 = int(rem / pl2)
                best = max(best, cols1 * n1 + cols2 * n2)

    return best


# ── 운임 비교 시뮬레이터 ──────────────────────────────────────────────────────

# 단거리 권역 (장거리 아닌 것)
_SHORT_DISTANCE_ZONES = {"경남권", "경북권", "부산권"}

# 장거리 권역 (20% 가산)
_LONG_DISTANCE_ZONES = {"강원권", "경기권", "서울권", "인천권", "전남권", "전북권", "충남권", "충북권"}

# 중량 → 차종 매핑 (오름차순)
_WEIGHT_TO_VEHICLE = [
    (1_000,  "1톤"),
    (2_500,  "2.5톤"),
    (3_500,  "3.5톤"),
    (5_000,  "5톤"),
    (8_000,  "8톤"),
    (11_000, "11톤"),
    (18_000, "18톤"),
    (25_000, "25톤"),
    (float("inf"), "트레일러"),
]

# 도착지 → 권역 매핑 (직송 시트 기준)
_ZONE_HEADERS = {
    "(강원권)": "강원권", "(경기권)": "경기권", "(경남권)": "경남권",
    "(경북권)": "경북권", "(부산권)": "부산권", "(서울권)": "서울권",
    "(인천권)": "인천권", "(전남권)": "전남권", "(전북권)": "전북권",
    "(충남권)": "충남권", "(충북권)": "충북권",
}


def _get_vehicle_for_weight(weight_kg: int) -> str:
    """중량(kg) → 적합 차종"""
    for limit, vehicle in _WEIGHT_TO_VEHICLE:
        if weight_kg <= limit:
            return vehicle
    return "트레일러"


def _get_zone_for_dest(dest: str, all_rows: list) -> Optional[str]:
    """
    직송 운임표 전체 rows에서 도착지가 속한 권역 반환.
    all_rows: [(출발지, 도착지, content), ...]
    """
    current_zone = None
    dest = dest.strip()
    for _, dest_val, _ in all_rows:
        dest_val = str(dest_val).strip()
        if dest_val in _ZONE_HEADERS:
            current_zone = _ZONE_HEADERS[dest_val]
        if dest_val == dest or dest in dest_val:
            return current_zone
    return None


def calculate_fare_comparison(dest: str, weight_kg: int) -> dict:
    """
    직송 / 화물 / 택배 운임을 계산해서 비교 결과 반환.
    - 직송: 중량→차종, 도착지→권역(하드코딩), 장거리 20% 가산
    - 화물/택배: 도착지 매칭 → 단가 × 중량
    """
    result = {"직송": None, "화물": None, "택배": None, "추천": None, "도착지": dest, "중량": weight_kg}

    # 권역별 도착지 하드코딩 (엑셀 데이터 기반)
    _DEST_TO_ZONE = {}
    _ZONE_DEST_MAP = {
        "강원권": ["강릉","도계","동해","묵호","북평","사북","삼척","석포","속초","양양","영양","영월","옥계","원주","정선","철원","춘천","태백","평창","평해","횡성"],
        "경기권": ["가평","강화군","과천","광명","광주","고양","구리","군포","기흥","김포","남양주","반월","부천","부평","성남","송탄","수색","수원","시흥","의정부","안산","안성","안양","양주","양평","여주","연천","오산","용인","의왕","이천","파주","평택","포천","하남","화성"],
        "경남권": ["거창","거제도","군위","기장","김해","남해","녹산","대저","덕계","마산","명지","밀양","사천","산청","삼천포","삼랑진","서창","신항","양산","언양","옥포","온산","용원","울산","웅상","웅촌","원동","의령","일광","장유","정관","진동","진례","진성","진양","진영","진주","진해","창녕","창원","철마","칠원","통영","하동","함안","함양","합천","현풍"],
        "경북권": ["감포","건천","경산","경주","고령","고성","구미","김천","달성","대구","문경","봉화","상주","성주","안동","영덕","영주","영천","예천","왜관","울진","월성","의성","점촌","청도","청송","칠곡","포항","풍기","후포"],
        "부산권": ["감만","감천","개금","구서","금사동","다대포","동래","만덕","망미동","반여동","사상","서면","신평","안락동","엄궁","연산동","용당","우암","자성대","장림","재송","중앙동","초읍","하단","학장","해운대"],
        "서울권": ["서울"],
        "인천권": ["영종도","옹진군","인천"],
        "전남권": ["강진","고흥","곡성","광양","광주","목포","벌교","순천","여수","여천","장성","장흥","함평","해남"],
        "전북권": ["고창","군산","김제","나주","남원","담양","무안","무주","부안","순창","영광","영암","완도","완주","이리","익산","전주","정읍","화순"],
        "충남권": ["공주","금산","논산","당진","대전","대천","보령","부여","서산","서천","신탄진","아산","연기","예산","조치원","천안","태안"],
        "충북권": ["괴산","단양","보은","세종","옥천","음성","제천","진천","청원","청주","충원","충주"],
    }
    for zone, cities in _ZONE_DEST_MAP.items():
        for city in cities:
            _DEST_TO_ZONE[city] = zone

    dest_strip = dest.strip()
    vehicle    = _get_vehicle_for_weight(weight_kg)

    # 권역 판별
    zone    = _DEST_TO_ZONE.get(dest_strip)
    is_long = zone in _LONG_DISTANCE_ZONES if zone else False

    logger.info(f"[운임 조회 시작] 도착지:{dest_strip} / 권역:{zone or '미매핑'} / 장거리:{is_long} / 차종:{vehicle} / 중량:{weight_kg}kg")

    try:
        client = QdrantClient(url=f"http://{QDRANT_HOST}:{QDRANT_PORT}", api_key=QDRANT_API_KEY)

        # ── 직송 운임 조회 ──────────────────────────────────────────────────
        # data_loader V5: fare_type="직송", source="운임_테이블.xlsx" 으로 저장
        direct_pts = client.scroll(
            collection_name=QDRANT_COLLECTION,
            scroll_filter=models.Filter(must=[models.FieldCondition(
                key="fare_type", match=models.MatchValue(value="직송")
            )]),
            limit=300
        )[0]

        for pt in direct_pts:
            payload  = pt.payload
            content  = payload.get("text") or payload.get("page_content", "")
            dest_val = str(payload.get("destination", "")).strip()
            origin   = str(payload.get("departure", "부산")).strip()

            # 도착지 매칭 (정확 일치 우선, 부분 포함 허용)
            if dest_strip != dest_val and dest_strip not in dest_val:
                continue

            # fares 딕셔너리에서 바로 차종 요금 추출
            fares_raw = payload.get("fares", {})
            if isinstance(fares_raw, dict):
                fare_val = fares_raw.get(vehicle, "")
                if fare_val:
                    try:
                        fare = int(str(fare_val).replace(",", "").strip())
                        if fare > 0:
                            final_fare = round(fare * 1.2) if is_long else fare
                            result["직송"] = {
                                "금액": final_fare,
                                "기본금액": fare,
                                "차종": vehicle,
                                "출발지": origin,
                                "권역": zone or "미분류",
                                "장거리": is_long,
                            }
                            logger.info(
                                f"[직송 산출] {origin}→{dest_val} / {vehicle} / "
                                f"기본:{fare:,}원"
                                + (f" / 장거리20%↑→{final_fare:,}원" if is_long else f" / {final_fare:,}원")
                            )
                            break
                    except Exception:
                        pass

            # fares 딕셔너리 없으면 text 파싱 fallback
            if not result["직송"] and content:
                for line in content.split("\n"):
                    if vehicle in line and ":" in line:
                        nums = re.findall(r'[\d,]+', line.split(":")[-1])
                        for n in nums:
                            try:
                                fare = int(n.replace(",", ""))
                                if fare > 10000:
                                    final_fare = round(fare * 1.2) if is_long else fare
                                    result["직송"] = {
                                        "금액": final_fare,
                                        "기본금액": fare,
                                        "차종": vehicle,
                                        "출발지": origin,
                                        "권역": zone or "미분류",
                                        "장거리": is_long,
                                    }
                                    break
                            except Exception:
                                continue
                        if result["직송"]:
                            break

            if result["직송"]:
                break

        # ── 화물/택배 운임 조회 ────────────────────────────────────────────
        parcel_pts = client.scroll(
            collection_name=QDRANT_COLLECTION,
            scroll_filter=models.Filter(must=[models.FieldCondition(
                key="fare_type", match=models.MatchValue(value="화물택배")
            )]),
            limit=300
        )[0]

        logger.info(f"[화물/택배] Qdrant 조회 건수: {len(parcel_pts)}")

        for pt in parcel_pts:
            payload  = pt.payload
            content  = payload.get("text") or payload.get("page_content", "")
            dest_val = str(payload.get("destination", "")).strip()

            # 도착지 먼저 text에서 파싱 시도
            if not dest_val:
                for part in content.split("|"):
                    part = part.strip()
                    if "도착지" in part:
                        dest_val = part.split(":")[-1].strip()
                        break

            if dest_strip != dest_val and dest_strip not in dest_val:
                continue

            logger.info(f"[화물/택배 매칭] 도착지:{dest_val} / content:{content[:80]}")

            # payload fares 직접 추출
            cargo_rate  = None
            parcel_rate = None

            hw = payload.get("fare_hwamul", "")
            tk = payload.get("fare_taekbae", "")
            if hw:
                try:
                    cargo_rate = int(str(hw).replace(",", "").strip())
                except Exception:
                    pass
            if tk:
                try:
                    parcel_rate = int(str(tk).replace(",", "").strip())
                except Exception:
                    pass

            # fallback: text 파싱
            if not cargo_rate or not parcel_rate:
                for part in content.split("|"):
                    part = part.strip()
                    if "화물" in part and ":" in part and not cargo_rate:
                        nums = re.findall(r'\d+', part.split(":")[-1].replace(",", ""))
                        if nums:
                            try:
                                cargo_rate = int(nums[0])
                            except Exception:
                                pass
                    if "택배" in part and ":" in part and not parcel_rate:
                        nums = re.findall(r'\d+', part.split(":")[-1].replace(",", ""))
                        if nums:
                            try:
                                parcel_rate = int(nums[0])
                            except Exception:
                                pass

            if cargo_rate and cargo_rate > 0 and result["화물"] is None:
                fare = cargo_rate * weight_kg
                result["화물"] = {"금액": fare, "단가": cargo_rate}
                logger.info(f"[화물 산출] {dest_val} / {cargo_rate:,}원/kg × {weight_kg}kg = {fare:,}원")

            if parcel_rate and parcel_rate > 0 and result["택배"] is None:
                fare = parcel_rate * weight_kg
                result["택배"] = {"금액": fare, "단가": parcel_rate}
                logger.info(f"[택배 산출] {dest_val} / {parcel_rate:,}원/kg × {weight_kg}kg = {fare:,}원")

            if result["화물"] and result["택배"]:
                break

    except Exception as e:
        logger.error(f"운임 비교 계산 오류: {e}")
        return result

    # ── 추천 결정 (직송 vs 택배만 비교) ──────────────────────────────────
    compare_targets = {k: v["금액"] for k, v in result.items()
                       if k in ("직송", "택배") and isinstance(v, dict) and v.get("금액")}
    if compare_targets:
        result["추천"] = min(compare_targets, key=compare_targets.get)
        logger.info(f"[운임 비교 결과] {compare_targets} → 추천: {result['추천']}")
    else:
        logger.warning(f"[운임 비교] 조회된 운임 없음 - 직송:{result['직송']} 화물:{result['화물']} 택배:{result['택배']}")

    return result


def get_freight_meta() -> dict:
    """Qdrant에서 직송/화물택배 메타 정보 조회 (selectbox 옵션용)"""
    try:
        client = QdrantClient(url=f"http://{QDRANT_HOST}:{QDRANT_PORT}", api_key=QDRANT_API_KEY)
        # data_loader V5: fare_type 키로 저장
        direct_pts = client.scroll(collection_name=QDRANT_COLLECTION,
            scroll_filter=models.Filter(must=[models.FieldCondition(
                key="fare_type", match=models.MatchValue(value="직송"))]),
            limit=5)[0]
        parcel_pts = client.scroll(collection_name=QDRANT_COLLECTION,
            scroll_filter=models.Filter(must=[models.FieldCondition(
                key="fare_type", match=models.MatchValue(value="화물택배"))]),
            limit=5)[0]
    except Exception as e:
        logger.error(f"운임 메타 조회 오류: {e}")
        return {}

    fare_ok   = len(direct_pts) > 0
    parcel_ok = len(parcel_pts) > 0
    logger.info(f"운임 메타 조회: 직송={fare_ok}, 화물택배={parcel_ok}")
    return {"loaded": fare_ok and parcel_ok}


def _load_vehicle_candidates(plt_w: float, plt_l: float) -> list:
    """
    Qdrant '차량 데이터' 문서에서 차량 후보 목록을 파싱해 반환.
    data_loader V5 텍스트 형식:
      | 1톤(1.2톤) | ~ 1.32 | 2.8 | 1.6 |
      | 트레일러(츄레라) | ~ 25 (폭 2.6m 이하) | 12 | 2.34 |
      | 로브이(Low-v, 로베드, Low-bed) | ~ 24 | 6 | 2.34 |
    """
    # ── 하드코딩 차량 목록 (Qdrant 조회 실패 시 fallback) ──────────────
    FALLBACK_VEHICLES_RAW = [
        ("1톤(1.2톤)",                   "~ 1.32",            2.8,  1.6),
        ("2.5톤",                        "1.33 ~ 2.75",       4.3,  1.8),
        ("3.5톤(신규)",                  "2.76 ~ 3.85",       4.8,  2.0),
        ("5톤",                          "3.86 ~ 5.5",        6.2,  2.34),
        ("8톤",                          "5.6 ~ 8.2",         7.4,  2.34),
        ("11톤",                         "8.3 ~ 12",          9.0,  2.34),
        ("18톤",                         "12.1 ~ 19",         10.1, 2.34),
        ("25톤",                         "19.1 ~ 25.5",       10.1, 2.34),
        ("트레일러(츄레라)",              "~ 25 (폭 2.6m 이하)", 12.0, 2.34),
        ("로브이(Low-v, 로베드, Low-bed)","~ 24",              6.0,  2.34),
    ]

    def _parse_max_weight(wt_str: str) -> Optional[float]:
        cleaned = re.sub(r'\(.*?\)', '', str(wt_str)).strip()
        nums = re.findall(r'[\d.]+', cleaned)
        return float(nums[-1]) if nums else None

    def _build_vehicles(raw_list) -> list:
        vehicles = []
        for name, wt_str, length, width in raw_list:
            is_lowbed = any(kw in name for kw in
                            ('로브이', 'Low-v', 'Low-bed', '로베드', 'low-bed', 'low-v'))
            skip_kw = ('높이기준', '특이사항', '톤수', '차량톤수')
            if any(k in name for k in skip_kw):
                continue
            if not name.strip():
                continue

            max_weight_ton = _parse_max_weight(wt_str)

            if is_lowbed:
                vehicles.append({
                    "name":           name,
                    "spec":           "높이 2.6m 이상 제품 전용 특수차량",
                    "max_plt":        999,
                    "max_weight_ton": max_weight_ton,
                    "is_lowbed":      True,
                    "length":         length,
                    "width":          width,
                })
            else:
                try:
                    l = float(length)
                    w = float(width)
                except Exception:
                    continue
                max_plt = _calc_loadable_plt(plt_w, plt_l, w, l)
                vehicles.append({
                    "name":           name,
                    "spec":           f"길이 {l}m / 폭 {w}m",
                    "max_plt":        max_plt,
                    "max_weight_ton": max_weight_ton,
                    "is_lowbed":      False,
                    "length":         l,
                    "width":          w,
                })
        return vehicles

    # ── Qdrant 조회 시도 ───────────────────────────────────────────────
    content = ""
    try:
        client = QdrantClient(url=f"http://{QDRANT_HOST}:{QDRANT_PORT}", api_key=QDRANT_API_KEY)

        # 1차: domain=vehicle + type=summary (V5 요약 청크)
        pts = client.scroll(
            collection_name="logistics_data",
            scroll_filter=models.Filter(must=[
                models.FieldCondition(key="domain", match=models.MatchValue(value="vehicle")),
                models.FieldCondition(key="type",   match=models.MatchValue(value="summary")),
            ]),
            limit=3
        )[0]

        # 2차: domain=vehicle 전체
        if not pts:
            pts = client.scroll(
                collection_name="logistics_data",
                scroll_filter=models.Filter(must=[
                    models.FieldCondition(key="domain", match=models.MatchValue(value="vehicle"))
                ]),
                limit=15
            )[0]

        # 3차: source=차량 데이터
        if not pts:
            pts = client.scroll(
                collection_name="logistics_data",
                scroll_filter=models.Filter(must=[
                    models.FieldCondition(key="source", match=models.MatchValue(value="차량 데이터"))
                ]),
                limit=15
            )[0]

        if pts:
            best = max(pts, key=lambda p: len(p.payload.get("text", p.payload.get("page_content", ""))))
            content = best.payload.get("text") or best.payload.get("page_content", "")

    except Exception as e:
        logger.warning(f"차량 데이터 Qdrant 조회 실패 → fallback 사용: {e}")

    # ── content 파싱 ──────────────────────────────────────────────────
    if content:
        raw_list = []
        for line in content.split('\n'):
            # 마크다운 표 행만 처리 (|---|--- 구분선 제외)
            if '|' not in line:
                continue
            parts = [p.strip() for p in line.strip().strip('|').split('|')]
            # 빈 첫 컬럼 제거
            if parts and parts[0] == '':
                parts = parts[1:]
            if len(parts) < 4:
                continue
            name, wt_str = parts[0], parts[1]
            # 헤더/구분선/특이사항 행 스킵
            # Fix: '' 제거 — 모든 문자열은 ''로 시작하므로 전체 차량명이 SKIP되는 버그 수정
            skip = ('차량톤수', '---', '높이기준', '특이사항')
            if any(name.startswith(s) for s in skip) or not name:
                continue
            # 길이/폭 파싱
            try:
                length_m = re.search(r'([\d.]+)', parts[2])
                width_m  = re.search(r'([\d.]+)', parts[3])
                if not length_m or not width_m:
                    continue
                raw_list.append((name, wt_str, float(length_m.group(1)), float(width_m.group(1))))
            except Exception:
                continue

        if raw_list:
            vehicles = _build_vehicles(raw_list)
            if vehicles:
                logger.info(f"차량 후보 {len(vehicles)}개 (Qdrant 파싱)")
                return vehicles

    # ── fallback: 하드코딩 데이터 사용 ────────────────────────────────
    logger.warning("차량 데이터 파싱 실패 → 내장 하드코딩 데이터 사용")
    vehicles = _build_vehicles(FALLBACK_VEHICLES_RAW)
    logger.info(f"차량 후보 {len(vehicles)}개 (하드코딩 fallback)")
    return vehicles


def get_db_transport_advice(total_pallets: float, total_weight_kg: float = 0.0,
                             plt_w: float = 1.1, plt_l: float = 1.1):
    """
    단일 자재 배차 추천 (하위호환 유지).
    내부적으로 get_split_dispatch_advice 호출 후 첫 번째 차량 반환.
    """
    result = get_split_dispatch_advice(
        items=[{"pallets": total_pallets, "weight_kg": total_weight_kg,
                "plt_w": plt_w, "plt_l": plt_l}]
    )
    if not result or not result.get("trucks"):
        return None
    first = result["trucks"][0]
    return {
        "name"          : first["name"],
        "spec"          : first["spec"],
        "max_plt"       : first["max_plt"],
        "max_weight_ton": first["max_weight_ton"],
        "weight_ok"     : first["weight_ok"],
        "is_lowbed"     : first.get("is_lowbed", False),
    }


def get_split_dispatch_advice(items: list) -> dict:
    """
    복수 자재 혼적 또는 단일 자재 분할 배차 추천.

    items: [{"pallets": int, "weight_kg": float, "plt_w": float, "plt_l": float}, ...]

    핵심 설계 원칙:
    ─ 혼적 시 차량 적재함에서 실제 배열 가능한 PLT 수는
      '가장 큰 파렛트' 기준으로 계산 (각 자재별 합산 방식 ❌)
      → 하나의 적재함 공간을 자재들이 공유하기 때문
    ─ 부피(PLT) + 중량 동시 만족하는 가장 작은 차량 선택
    ─ 한 차량 불가 → 중량 기준으로 최대한 적재 후 나머지 재귀 분할
    """
    try:
        total_plt       = sum(it["pallets"] for it in items)
        total_weight_kg = sum(it["weight_kg"] for it in items)

        # 혼적 부피 계산 기준: 가장 큰 파렛트 (면적 = w*l 기준)
        rep_item    = max(items, key=lambda it: it["plt_w"] * it["plt_l"])
        rep_plt_w   = rep_item["plt_w"]
        rep_plt_l   = rep_item["plt_l"]

        # 차량 목록 로드 (대표 파렛트 사이즈 기준 max_plt 계산)
        all_vehicles = _load_vehicle_candidates(rep_plt_w, rep_plt_l)
        if not all_vehicles:
            return {"trucks": [], "total_plt": total_plt, "total_weight_kg": total_weight_kg,
                    "split": False, "error": "차량 DB를 불러올 수 없습니다."}

        # 일반 차량만, 최대중량 오름차순 정렬 (동일 중량이면 max_plt 오름차순)
        normal_vehicles = sorted(
            [v for v in all_vehicles if not v["is_lowbed"]],
            key=lambda x: (x["max_weight_ton"] or 0, x["max_plt"])
        )

        # 일반 차량이 없으면 전체 사용
        if not normal_vehicles:
            normal_vehicles = sorted(
                all_vehicles,
                key=lambda x: (x["max_weight_ton"] or 0, x["max_plt"])
            )
        if not normal_vehicles:
            return {"trucks": [], "total_plt": total_plt, "total_weight_kg": total_weight_kg,
                    "split": False, "error": "사용 가능한 차량이 없습니다."}

        def _best_single_vehicle(need_plt: int, need_weight_kg: float) -> Optional[dict]:
            """
            need_plt, need_weight_kg 를 단일 차량으로 처리 가능한
            후보 중 가장 작은 차량 반환.
            ─ 1순위: 부피 OK + 중량 OK  → 최대중량 기준 가장 작은 차량
            ─ 2순위: 부피 OK + 중량 초과 → 최대중량 기준 가장 작은 차량 + 경고
            """
            need_weight_ton = need_weight_kg / 1000.0
            ok_both, ok_vol = [], []

            for v in normal_vehicles:
                # 부피: 대표 파렛트 기준 이 차량에 몇 PLT 올릴 수 있는가
                cap_plt = v["max_plt"]   # _load_vehicle_candidates에서 이미 계산됨
                if cap_plt < need_plt:
                    continue             # 부피 부족 → 제외

                wt_ok = True
                if need_weight_ton > 0 and v["max_weight_ton"] is not None:
                    wt_ok = need_weight_ton <= v["max_weight_ton"]

                entry = {
                    **v,
                    "weight_ok"         : wt_ok,
                    "assigned_plt"      : need_plt,
                    "assigned_weight_kg": need_weight_kg,
                    "load_ratio_vol"    : (need_plt / cap_plt) * 100,
                    "load_ratio_wt"     : (
                        (need_weight_ton / v["max_weight_ton"]) * 100
                        if v["max_weight_ton"] else 0.0
                    ),
                }
                if wt_ok:
                    ok_both.append(entry)
                else:
                    ok_vol.append(entry)

            if ok_both:
                # 중량 OK 중 가장 작은 차량 (최대중량 오름차순 → 첫 번째)
                return ok_both[0]
            if ok_vol:
                # 중량 초과지만 부피는 맞는 가장 작은 차량
                return ok_vol[0]
            return None

        def _split_dispatch(remain_plt: int, remain_weight_kg: float,
                            depth: int = 0) -> list:
            """
            remain_plt / remain_weight_kg 를 배차.

            전략:
            1) 단일 차량으로 부피+중량 모두 OK → 즉시 반환
            2) 불가 시 → 각 차량이 이번에 실을 수 있는 최대 PLT를
               (부피 cap, 중량 cap 동시 만족) 기준으로 계산하고,
               '가장 많이 싣되 가장 작은 차량' 선택 → 나머지 재귀
               (차량 대수를 최소화하면서 각 차량을 최대 활용)
            """
            if depth > 30 or remain_plt <= 0:
                return []

            wt_per_plt = remain_weight_kg / remain_plt if remain_plt > 0 else 0

            # ── 1) 단일 차량 가능한지 먼저 시도 ──────────────────────────
            best = _best_single_vehicle(remain_plt, remain_weight_kg)
            if best and best["weight_ok"]:
                return [best]

            # ── 2) 분할: 각 차량별로 이번에 실을 수 있는 최대 PLT 계산 ──
            # (부피 한도 & 중량 한도 동시 만족)
            # 결과: (실을_PLT, 차량) 쌍 중 실을_PLT 최대 → 동률이면 차량 작은 것
            best_load  = 0          # 이번 차량에 실을 최대 PLT
            chosen     = None       # 선택된 차량
            chosen_wkg = 0.0

            for v in normal_vehicles:   # 작은 차량부터 순회
                cap_vol = v["max_plt"]
                cap_wt_kg = (v["max_weight_ton"] * 1000.0
                             if v["max_weight_ton"] else float("inf"))

                # 중량 제약으로 이 차량에 실을 수 있는 최대 PLT
                if wt_per_plt > 0 and cap_wt_kg < float("inf"):
                    max_by_wt = int(cap_wt_kg / wt_per_plt)
                else:
                    max_by_wt = remain_plt

                load = min(cap_vol, max_by_wt, remain_plt)
                if load <= 0:
                    continue

                load_wkg = wt_per_plt * load
                # 이 차량 중량 OK 재확인
                if v["max_weight_ton"] and load_wkg / 1000.0 > v["max_weight_ton"]:
                    continue

                # 더 많이 실을 수 있는 차량 발견 시 갱신
                # (동률이면 이미 작은 차량이 chosen이므로 갱신 안 함)
                if load > best_load:
                    best_load  = load
                    chosen     = v
                    chosen_wkg = load_wkg

            # 아무 차량도 선택 못 했으면 가장 큰 차량에 강제 1PLT
            if chosen is None:
                chosen    = max(normal_vehicles,
                                key=lambda x: (x["max_plt"], x["max_weight_ton"] or 0))
                best_load = 1
                chosen_wkg = wt_per_plt * best_load

            cap       = chosen["max_plt"]
            left_plt  = remain_plt - best_load
            left_wkg  = remain_weight_kg - chosen_wkg
            wt_ok_f   = (chosen_wkg / 1000.0 <= chosen["max_weight_ton"]
                         if chosen["max_weight_ton"] else True)

            truck = {
                **chosen,
                "weight_ok"         : wt_ok_f,
                "assigned_plt"      : best_load,
                "assigned_weight_kg": chosen_wkg,
                "load_ratio_vol"    : (best_load / cap) * 100,
                "load_ratio_wt"     : (
                    (chosen_wkg / 1000.0 / chosen["max_weight_ton"]) * 100
                    if chosen["max_weight_ton"] else 0.0
                ),
            }
            return [truck] + _split_dispatch(left_plt, left_wkg, depth + 1)

        trucks   = _split_dispatch(total_plt, total_weight_kg)
        is_split = len(trucks) > 1

        return {
            "trucks"         : trucks,
            "total_plt"      : total_plt,
            "total_weight_kg": total_weight_kg,
            "split"          : is_split,
            "error"          : None,
        }

    except Exception as e:
        logger.error(f"⚠️ 분할 배차 계산 오류: {e}")
        return {"trucks": [], "total_plt": 0, "total_weight_kg": 0,
                "split": False, "error": str(e)}
# 메인 함수
if __name__ == "__main__":
    test_queries = [
        "6004010 자재 정보",
        "컨베어벨트 목록을 표로 보여줘"
    ]
    
    conversation_context = []
    
    for test_q in test_queries:
        try:
            print(f"\n{'='*60}\n질문: {test_q}\n{'='*60}")
            
            start_time = time.time()
            result = get_rag_response(test_q, context=conversation_context)
            elapsed = time.time() - start_time
            
            print(f"\n{result['answer']}")
            print(f"\n표 포함: {result['has_table']}")
            print(f"\n⏱️ 응답 시간: {elapsed:.2f}초\n{'='*60}")
            
            conversation_context.append({
                "query": test_q,
                "answer": result['answer'],
                "timestamp": datetime.now().isoformat()
            })
            
        except Exception as e:
            logger.exception(f"테스트 실패: {e}")