"""
data_loader.py  ─  Logibot V5 Qdrant 적재 모듈
=====================================================
[시트 → 도메인 매핑]
  물류팀 운영 규칙          → operation_rule
  포장량 산출 데이터         → packaging
  수출 포장량 산출 수식      → export_rule
  차량 데이터               → vehicle
  용차 차량 노선 데이터      → route
  지입 차량(기사) 노선 데이터→ driver_route
  컨베어벨트 규격 데이터     → conveyor
  컨베어벨트 직경 산출 수식  → conveyor_formula
  주름혹벨트 우든박스 데이터 → sidewall
  물류팀 현황 데이터         → personnel
  크롤러 러버트랙 규격 데이터→ crawler
  파렛트, 박스 데이터        → pallet_box
"""

import os, uuid, logging, re
from datetime import date, datetime, timedelta
from typing import List, Dict, Any

import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from langchain_ollama import OllamaEmbeddings
from qdrant_client import QdrantClient
from qdrant_client.models import (
    Distance, VectorParams, PointStruct,
    PayloadSchemaType,
)

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# ── 환경 변수 ──────────────────────────────────────────────────
QDRANT_COLLECTION      = os.getenv("QDRANT_COLLECTION",      "logistics_data")
QDRANT_HOST            = os.getenv("QDRANT_HOST",            "localhost")
QDRANT_PORT            = int(os.getenv("QDRANT_PORT",        "6333"))
QDRANT_API_KEY         = os.getenv("QDRANT_API_KEY",         None)
OLLAMA_EMBEDDING_MODEL = os.getenv("OLLAMA_EMBEDDING_MODEL", "granite-embedding:278m")
OLLAMA_HOST            = os.getenv("OLLAMA_HOST",            "http://localhost:11434")
VECTOR_DIM             = 768
BATCH_SIZE             = 64

# ── 격주 기준 앵커 (5/28이 속한 주 월요일) ──────────────────────
BIWEEKLY_ANCHOR = date(2026, 5, 25)  # 5/28 적용 시작주의 월요일 (B주=이용구 기준)   # 현재 A주 시작일


# ══════════════════════════════════════════════════════════════
#  유틸리티
# ══════════════════════════════════════════════════════════════

def _clean(v) -> str:
    """None / 특수문자 / 개행 정리 → 단순 문자열"""
    if v is None:
        return ""
    s = str(v).strip()
    s = s.replace("\xa0", " ")        # 비공백 스페이스
    s = re.sub(r"\n+", " | ", s)      # 개행 → 구분자
    s = re.sub(r"\s{2,}", " ", s)     # 연속 공백
    return s


def _excel_time_to_str(v) -> str:
    """엑셀 소수점 시간(0.534...) → '12시50분' 문자열"""
    if isinstance(v, float):
        total_min = round(v * 24 * 60)
        h, m = divmod(total_min, 60)
        return f"{h}시{m:02d}분" if m else f"{h}시"
    return _clean(v)


def get_week_group(target: date = None) -> str:
    """오늘(또는 target)이 A주인지 B주인지 반환.
    앵커주(BIWEEKLY_ANCHOR) = B주 = 이용구 운행주.
    query_processor.py get_week_group()과 동일 기준 유지."""
    if target is None:
        target = date.today()
    monday = target - timedelta(days=target.weekday())
    weeks_elapsed = (monday - BIWEEKLY_ANCHOR).days // 7
    return "B" if weeks_elapsed % 2 == 0 else "A"  # Fix: 앵커주=B주(이용구) 기준 통일


def get_active_driver_schedule(df: pd.DataFrame, target: date = None) -> pd.DataFrame:
    """
    오늘 날짜 기준 활성 노선만 필터링.
    - 노선그룹 == '공통' : 항상 포함
    - 노선그룹 == 'A' or 'B' : 이번 주 그룹만 포함
    """
    if target is None:
        target = date.today()
    wg = get_week_group(target)
    mask = (df["노선그룹"] == "공통") | (df["노선그룹"] == wg)
    return df[mask].copy()


# ══════════════════════════════════════════════════════════════
#  시트별 Document 생성 함수
#  각 함수는 {"text": str, "metadata": dict} 리스트를 반환
# ══════════════════════════════════════════════════════════════

def load_operation_rules(ws) -> List[Dict]:
    """물류팀 운영 규칙 → ① 압축 summary 1개 + ② Q&A 개별 청크 51개"""
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 행1: 제목, 행2: 헤더(카테고리/질문/답변)

    qa_list = []
    for row in rows[2:]:
        if not row[1]:
            continue
        category = _clean(row[0]) or "일반_운영"
        question = _clean(row[1])
        answer   = _clean(row[2]) if len(row) > 2 else ""
        if not question:
            continue
        qa_list.append((category, question, answer))

    # ① 압축 summary 청크 (fetch_whole_docs 시 6000자 context 내 전체 Q&A 커버용)
    # 각 Q&A를 한 줄로 압축하여 하나의 문서로 저장
    summary_lines = ["[물류팀 운영 규칙 - 전체 Q&A 요약]"]
    for cat, q, a in qa_list:
        a_short = a[:100] + "..." if len(a) > 100 else a
        summary_lines.append(f"[{cat}] Q:{q} → A:{a_short}")
    summary_text = "\n".join(summary_lines)
    docs.append({
        "text": summary_text,
        "metadata": {
            "domain":  "operation_rule",
            "type":    "summary",
            "source":  "물류팀 운영 규칙",
        }
    })

    # ② 개별 Q&A 청크 (벡터 검색 정밀도용)
    for category, question, answer in qa_list:
        text = f"[물류팀 운영 규칙 | {category}]\n질문: {question}\n답변: {answer}"
        docs.append({
            "text": text,
            "metadata": {
                "domain":   "operation_rule",
                "type":     "qa",
                "category": category,
                "question": question,
                "answer":   answer,
                "source":   "물류팀 운영 규칙",
            }
        })
    logger.info(f"  물류팀 운영 규칙: summary 1개 + Q&A {len(qa_list)}개 = 총 {len(docs)}건")
    return docs


def load_packaging_data(ws) -> List[Dict]:
    """포장량 산출 데이터 → 포장재 종류별 청크"""
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 행1: 헤더 (포장재_키워드 / 구분 / B01 / B02 / N18 / N19)
    header = rows[0]
    # 자재그룹명 추출: "자재그룹 B01 박스당 무게(KG)\nB01 : 랩트벨트" → "B01"
    groups = []
    for h in header[2:6]:
        m = re.search(r'(B0[1-9]|N1[0-9])', str(h) if h else "")
        groups.append(m.group(1) if m else str(h))

    for row in rows[1:]:
        if not row[1]:
            continue
        keywords = _clean(row[0])
        label    = _clean(row[1])
        weights  = {groups[i]: _clean(row[2+i]) for i in range(len(groups)) if 2+i < len(row)}
        is_sleeve = "슬리브" in label

        weight_lines = "\n".join(
            f"  - {g}({get_group_name(g)}): {w} KG/박스" for g, w in weights.items() if w
        )
        note = "\n※ 슬리브 항목은 현재 파렛트 포장으로 변경 중인 항목입니다." if is_sleeve else ""

        text = (
            f"[포장량 산출 데이터]\n"
            f"포장재 종류: {label}\n"
            f"검색 키워드: {keywords}\n"
            f"자재그룹별 1박스당 중량:\n{weight_lines}{note}"
        )
        docs.append({
            "text": text,
            "metadata": {
                "domain":    "packaging",
                "label":     label,
                "keywords":  keywords,
                "is_sleeve": is_sleeve,
                "weights":   weights,
                "source":    "포장량 산출 데이터",
            }
        })
    logger.info(f"  포장량 산출 데이터: {len(docs)}건")
    return docs


def get_group_name(code: str) -> str:
    mapping = {"B01": "랩트벨트", "B02": "로엣지벨트", "N18": "리브드벨트", "N19": "타이밍벨트"}
    return mapping.get(code, code)


def load_export_formula(ws) -> List[Dict]:
    """수출 포장량 산출 수식 → 규칙 전문 1개 청크"""
    rows = list(ws.iter_rows(values_only=True))
    lines = [_clean(row[0]) for row in rows if row[0] and _clean(row[0])]
    text = "[수출 포장량 산출 규칙]\n" + "\n".join(lines)
    docs = [{"text": text, "metadata": {"domain": "export_rule", "source": "수출 포장량 산출 수식"}}]
    logger.info(f"  수출 포장량 산출 수식: {len(docs)}건")
    return docs


def load_vehicle_data(ws) -> List[Dict]:
    """차량 데이터 → 차량 전체를 하나의 표 청크 + 톤수별 개별 청크"""
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 헤더: 차량톤수 / 최대중량(ton) / 적재함길이(m) / 적재함폭(m)
    table_lines = ["[차량 데이터 - 전체 목록]", "| 차량톤수 | 최대중량(ton) | 적재함길이(m) | 적재함폭(m) |",
                   "|---------|------------|------------|----------|"]
    for row in rows[1:]:
        if not row[0]: continue
        ton  = _clean(row[0])
        wmax = _clean(row[1])
        leng = _clean(row[2])
        wid  = _clean(row[3])
        if not ton: continue
        table_lines.append(f"| {ton} | {wmax} | {leng} | {wid} |")
        # 개별 청크
        text = (f"[차량 데이터]\n"
                f"차량톤수: {ton}\n"
                f"최대적재중량: {wmax} ton\n"
                f"적재함 길이: {leng} m\n"
                f"적재함 폭: {wid} m")
        docs.append({"text": text,
                     "metadata": {"domain": "vehicle", "ton": ton,
                                  "max_weight": wmax, "length": leng, "width": wid,
                                  "source": "차량 데이터"}})
    # 전체 표 청크
    docs.insert(0, {"text": "\n".join(table_lines),
                    "metadata": {"domain": "vehicle", "type": "summary", "source": "차량 데이터"}})
    logger.info(f"  차량 데이터: {len(docs)}건")
    return docs


def load_route_data(ws) -> List[Dict]:
    """용차 차량 노선 데이터 → 도착지별 청크 + 권역별 묶음 청크"""
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 헤더: 출발지 / 도착지 / 권역 / 거리기준 / 예상소요시간 / 비고
    region_map: Dict[str, List[str]] = {}

    for row in rows[1:]:
        if not row[1]: continue
        depart = _clean(row[0]) or "부산"
        dest   = _clean(row[1])
        region = _clean(row[2])
        dist   = _clean(row[3])
        time_  = _clean(row[4])
        note   = _clean(row[5])

        if not dest: continue

        text = (f"[용차 차량 노선]\n"
                f"출발지: {depart} | 도착지: {dest} | 권역: {region}\n"
                f"거리 기준: {dist} | 예상 소요시간: {time_}")
        if note:
            text += f"\n비고: {note}"

        docs.append({"text": text,
                     "metadata": {"domain": "route", "departure": depart,
                                  "destination": dest, "region": region,
                                  "distance_type": dist, "travel_time": time_,
                                  "source": "용차 차량 노선 데이터"}})
        # 권역 묶음용
        region_map.setdefault(region, []).append(f"{dest}({time_})")

    # 권역별 묶음 청크 (한 번에 조회할 때 유리)
    for region, destinations in region_map.items():
        text = (f"[용차 노선 권역 요약 | {region}]\n"
                f"권역: {region}\n"
                f"해당 도착지 목록: {', '.join(destinations)}")
        docs.append({"text": text,
                     "metadata": {"domain": "route", "type": "region_summary",
                                  "region": region, "source": "용차 차량 노선 데이터"}})

    logger.info(f"  용차 차량 노선 데이터: {len(docs)}건")
    return docs


def load_driver_route(ws) -> List[Dict]:
    """
    지입기사 노선 데이터 → 기사+요일 단위 청크
    - 공통 노선: 항상 적재
    - A/B 노선: 메타데이터에 route_group 저장, 검색 시 필터 가능
    - 예상납품시간 소수점 → 한국어 시간 변환
    """
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 헤더: 기사명/소속/요일/노선그룹/순서/권역/거래처명/예상납품시간/적용시작일/적용종료일/조건_비고

    # 기사+요일+노선그룹 단위로 그룹핑
    from collections import defaultdict
    groups: Dict[tuple, List] = defaultdict(list)
    for row in rows[1:]:
        if not row[0]: continue
        key = (_clean(row[0]), _clean(row[2]), _clean(row[3]))  # 기사명, 요일, 노선그룹
        groups[key].append(row)

    for (driver, weekday, route_group), group_rows in groups.items():
        if not driver or not weekday: continue

        affil        = _clean(group_rows[0][1])
        apply_start  = group_rows[0][8]   # datetime or None
        apply_end    = group_rows[0][9]

        # 경유지 목록 조합
        stops = []
        for row in sorted(group_rows, key=lambda r: (r[4] or 0)):
            seq    = str(int(row[4])) if row[4] else ""  # Fix: 0 or "" = "" 방지
            region = _clean(row[5])
            dest   = _clean(row[6])
            t      = _excel_time_to_str(row[7])
            note   = _clean(row[10])
            stop_str = f"{seq}. ({region}) {dest} {t}"
            if note and note not in ("고정운행 노선", ""):
                stop_str += f" ※{note}"
            stops.append(stop_str)

        route_note = ""
        if route_group in ("A", "B"):
            route_note = (f"\n※ 서울(중부물류센터) 기사는 격주 교대 운행합니다. "
                          f"현재 운행 중인 노선그룹은 시스템 날짜 기준으로 자동 판별됩니다. "
                          f"(이 데이터는 '{route_group}노선'입니다)")

        text = (f"[지입기사 납품 동선 | {driver} | {weekday} | 노선그룹:{route_group}]\n"
                f"기사명: {driver} | 소속: {affil} | 요일: {weekday}\n"
                f"납품 동선:\n" + "\n".join(stops) + route_note)

        docs.append({
            "text": text,
            "metadata": {
                "domain":       "driver_route",
                "driver":       driver,
                "affiliation":  affil,
                "weekday":      weekday,
                "route_group":  route_group,       # '공통' / 'A' / 'B'
                "apply_start":  str(apply_start)[:10] if apply_start else "",
                "apply_end":    str(apply_end)[:10]   if apply_end   else "",
                "stop_count":   len(stops),
                "source":       "지입 차량(기사) 노선 데이터",
            }
        })

    logger.info(f"  지입기사 노선 데이터: {len(docs)}건")
    return docs


def load_conveyor_data(ws) -> List[Dict]:
    """컨베어벨트 규격 데이터 → 자재코드별 청크"""
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 헤더: 자재코드 / 자재내역 / 자재그룹 / M당순중량 / 포규격 / 포무게 /
    #        입힘무게 / 코팅후포두께 / 포두께 / 제품폭 / 심체수 / 상고무두께 / 하고무두께
    # (우측에 포규격 참조표가 있지만 col 14~17은 무시)

    for row in rows[1:]:
        code = _clean(row[0])
        if not code or not code.isdigit(): continue

        desc     = _clean(row[1])
        group    = _clean(row[2])
        weight_m = _clean(row[3])
        fab_spec = _clean(row[4])
        fab_w    = _clean(row[5])
        coat_w   = _clean(row[6])
        coat_t   = _clean(row[7])
        fab_t    = _clean(row[8])
        width    = _clean(row[9])
        ply      = _clean(row[10])
        top_t    = _clean(row[11])
        bot_t    = _clean(row[12])

        text = (
            f"[컨베어벨트 규격 | 자재코드:{code}]\n"
            f"자재코드: {code}\n"
            f"자재내역: {desc}\n"
            f"자재그룹: {group}\n"
            f"M당 순중량: {weight_m} KG\n"
            f"포규격: {fab_spec} | 포두께: {fab_t} mm | 제품폭: {width} mm\n"
            f"심체수(PLY): {ply} | 상고무두께: {top_t} mm | 하고무두께: {bot_t} mm\n"
            f"포무게: {fab_w} g | 입힘무게: {coat_w} g | 코팅후포두께: {coat_t} mm"
        )
        docs.append({
            "text": text,
            "metadata": {
                "domain":       "conveyor",
                "material_code": code,
                "description":  desc,
                "material_group": group,
                "weight_per_m": weight_m,
                "fabric_spec":  fab_spec,
                "width_mm":     width,
                "ply":          ply,
                "top_rubber":   top_t,
                "bot_rubber":   bot_t,
                "coat_thick":   coat_t,
                "source":       "컨베어벨트 규격 데이터",
            }
        })

    logger.info(f"  컨베어벨트 규격 데이터: {len(docs)}건")
    return docs


def load_conveyor_formula(ws) -> List[Dict]:
    """컨베어벨트 직경 산출 수식 → 규칙 전문 1개 청크"""
    rows = list(ws.iter_rows(values_only=True))
    lines = []
    for row in rows:
        v0 = _clean(row[0])
        v1 = _clean(row[1]) if len(row) > 1 else ""
        if v0:
            lines.append(f"{v0}" + (f": {v1}" if v1 else ""))
    text = "[컨베어벨트 직경 산출 공식 및 규칙]\n" + "\n".join(lines)
    docs = [{"text": text, "metadata": {"domain": "conveyor_formula", "source": "컨베어벨트 직경 산출 수식"}}]
    logger.info(f"  컨베어벨트 직경 산출 수식: {len(docs)}건")
    return docs


def load_sidewall_data(ws) -> List[Dict]:
    """주름혹벨트 우든박스 → 자재코드별 청크"""
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 헤더: 자재코드 / 자재내역 / 수량(M) / 순중량(KG) / 박스규격 / 박스중량기준 / 총중량계산식

    # 중량 기준 규칙: 행2(첫 번째 데이터행)에만 있음 → 전체 공통 규칙으로 추출
    WEIGHT_RULE = ""
    CALC_RULE = ""
    for row in rows[1:]:
        if row[5]:
            WEIGHT_RULE = _clean(row[5])
        if row[6]:
            CALC_RULE = _clean(row[6])
        if WEIGHT_RULE and CALC_RULE:
            break

    # 박스 규격 → 우든박스 중량 추정 함수
    def _estimate_box_weight(box_size_str: str) -> str:
        """
        박스 규격(W×L×H mm)으로 우든박스 중량 추정.
        기준: 큰 사이즈(부피 80m³ 이상) → 약 4000KG
             중간 사이즈(20~80m³) → 약 2000~3000KG
             작은 사이즈(20m³ 미만) → 약 1000KG
        """
        import re as _re
        nums = [int(n) for n in _re.findall(r'\d+', box_size_str)]
        if len(nums) < 3:
            return "규격 미확인"
        vol_m3 = (nums[0] * nums[1] * nums[2]) / 1e9
        if vol_m3 >= 80:
            return f"약 4000KG (큰 사이즈, 부피 {vol_m3:.1f}m³ 기준)"
        elif vol_m3 >= 20:
            return f"약 2000~3000KG (중간 사이즈, 부피 {vol_m3:.1f}m³ 기준)"
        else:
            return f"약 1000KG (작은 사이즈, 부피 {vol_m3:.1f}m³ 기준)"

    for row in rows[1:]:
        code = _clean(row[0])
        if not code or not str(code).isdigit(): continue

        desc      = _clean(row[1])
        qty_m     = _clean(row[2])
        net_w     = _clean(row[3])
        box_size  = _clean(row[4])
        box_w_note= _clean(row[5]) or WEIGHT_RULE  # 빈 경우 공통 규칙 사용
        calc_note = _clean(row[6]) or CALC_RULE

        # 우든박스 중량 추정값 계산
        est_box_weight = _estimate_box_weight(box_size) if box_size else "규격 미확인"

        # 총중량 계산 예시 (수치가 있을 때)
        total_example = ""
        try:
            q = float(qty_m)
            n = float(net_w)
            belt_w = q * n
            total_example = f"\n벨트 중량: {q} M × {n} KG/M = {belt_w:,.0f} KG\n우든박스 중량: {est_box_weight}\n총 중량 예상: {belt_w:,.0f} KG + 우든박스 중량"
        except Exception:
            pass

        text = (
            f"[주름혹벨트 우든박스 | 자재코드:{code}]\n"
            f"자재코드: {code}\n"
            f"자재내역: {desc}\n"
            f"수량: {qty_m} M | 순중량: {net_w} KG/M\n"
            f"우든박스 규격(W×L×H): {box_size} mm\n"
            f"우든박스 중량 추정: {est_box_weight}\n"
            f"우든박스 중량 기준 규칙: {box_w_note}\n"
            f"총중량 계산식: (수량_M × 순중량_KG) + 우든박스_중량_KG"
            f"{total_example}"
        )
        docs.append({
            "text": text,
            "metadata": {
                "domain":            "sidewall",
                "material_code":     str(code),
                "description":       desc,
                "qty_m":             qty_m,
                "net_weight":        net_w,
                "box_size":          box_size,
                "est_box_weight_kg": est_box_weight,
                "source":            "주름혹벨트 우든박스 사이즈 데이터",
            }
        })

    logger.info(f"  주름혹벨트 우든박스: {len(docs)}건")
    return docs


def load_personnel_data(ws) -> List[Dict]:
    """물류팀 현황 데이터 → 팀원별 청크 + 전체 요약 청크
    컬럼: 구분 / 성명 / 직책 / 전화번호 / 담당 공정 및 업무 범위 / 부재시 대응자
    """
    docs = []
    rows = list(ws.iter_rows(values_only=True))

    # 헤더에서 '부재시 대응자' 컬럼 인덱스 동적 탐지
    header = [str(c).strip() if c else "" for c in rows[0]]
    absence_col = None
    for i, h in enumerate(header):
        if "부재" in h or "대응자" in h:
            absence_col = i
            break

    all_lines = ["[물류팀 현황 - 전체 목록]",
                 "| 구분 | 성명 | 직책 | 내선번호 | 담당공정 | 부재시 대응자 |",
                 "|-----|-----|-----|--------|--------|------------|"]

    for row in rows[1:]:
        if not row[1]: continue
        div     = _clean(row[0])
        name    = _clean(row[1])
        pos     = _clean(row[2])
        tel     = _clean(row[3])
        duty    = _clean(row[4])
        absence = _clean(row[absence_col]) if absence_col and absence_col < len(row) and row[absence_col] else ""

        if not name: continue

        # 부재 대응자 정보 포함 텍스트 생성
        absence_line = f"\n부재시 대응자: {absence}" if absence else ""
        text = (f"[물류팀 현황 | {name} {pos}]\n"
                f"구분: {div} | 성명: {name} | 직책: {pos}\n"
                f"내선번호: {tel}\n"
                f"담당공정/업무: {duty}"
                f"{absence_line}")

        docs.append({
            "text": text,
            "metadata": {
                "domain":   "personnel",
                "name":     name,
                "position": pos,
                "tel":      tel,
                "duty":     duty,
                "division": div,
                "absence":  absence,
                "source":   "물류팀 현황 데이터",
            }
        })
        all_lines.append(f"| {div} | {name} | {pos} | {tel} | {duty} | {absence} |")

    # 전체 요약 청크
    docs.insert(0, {"text": "\n".join(all_lines),
                    "metadata": {"domain": "personnel", "type": "summary", "source": "물류팀 현황 데이터"}})
    logger.info(f"  물류팀 현황 데이터: {len(docs)}건")
    return docs


def load_crawler_data(ws) -> List[Dict]:
    """크롤러 러버트랙 → 자재코드별 청크"""
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 헤더: 자재그룹 / 자재코드 / 1PC당중량 / 자재내역 / 1파렛트당최대적재수 / 파렛트사이즈 / 설명및로직

    # 공통 로직 (첫 번째 행에서 한 번만 읽음)
    common_logic = ""
    for row in rows[1:]:
        if row[6] and _clean(row[6]):
            common_logic = _clean(row[6])
            break

    for row in rows[1:]:
        code = _clean(row[1])
        if not code or not str(code).isdigit(): continue

        group   = _clean(row[0])
        weight  = _clean(row[2])
        desc    = _clean(row[3])
        max_pc  = _clean(row[4])
        plt_size= _clean(row[5])

        text = (
            f"[크롤러 러버트랙 | 자재코드:{code}]\n"
            f"자재그룹: {group} | 자재코드: {code}\n"
            f"자재내역: {desc}\n"
            f"1PC당 중량: {weight} KG\n"
            f"1파렛트 최대적재수: {max_pc} PC\n"
            f"파렛트 사이즈(L×W): {plt_size} mm\n"
            f"배차 추천 로직: {common_logic}"
        )
        docs.append({
            "text": text,
            "metadata": {
                "domain":          "crawler",
                "material_code":   str(code),
                "material_group":  group,
                "description":     desc,
                "weight_per_pc":   weight,
                "max_pc_per_pallet": max_pc,
                "pallet_size":     plt_size,
                "source":          "크롤러 러버트랙 규격 데이터",
            }
        })

    logger.info(f"  크롤러 러버트랙: {len(docs)}건")
    return docs


def load_pallet_box_data(ws) -> List[Dict]:
    """파렛트·박스 데이터 → 공정별 묶음 청크 + 전체 요약 청크"""
    docs = []
    rows = list(ws.iter_rows(values_only=True))
    # 헤더: 공정 / 포장재종류 / 포장재구분상세

    from collections import defaultdict
    process_map: Dict[str, List[str]] = defaultdict(list)

    for row in rows[1:]:
        if not row[0]: continue
        process = _clean(row[0])
        p_type  = _clean(row[1]).strip()
        size    = _clean(row[2])
        if not process or not size: continue
        process_map[process].append(f"  - {p_type}: {size}")

    # 전체 요약 청크
    summary_lines = ["[파렛트·박스 데이터 - 전체 목록]"]
    for proc, items in process_map.items():
        summary_lines.append(f"\n## {proc}")
        summary_lines.extend(items)
    docs.append({"text": "\n".join(summary_lines),
                 "metadata": {"domain": "pallet_box", "type": "summary", "source": "파렛트, 박스 데이터"}})

    # 공정 → 동의어 매핑 (벡터 검색 유사도 향상)
    PROCESS_ALIAS = {
        "전동내수": "내수용 · 국내용 · 내수 PLT",
        "전동수출": "수출용 · 해외용 · 수출 PLT",
        "크롤러":   "크롤러 러버트랙",
        "컨베어벨트": "컨베어 벨트 · 컨베이어",
    }
    # 공정별 청크
    for process, items in process_map.items():
        alias = PROCESS_ALIAS.get(process, "")
        alias_line = f"\n검색 키워드: {alias}" if alias else ""
        text = (f"[파렛트·박스 데이터 | {process}]\n"
                f"공정: {process}{alias_line}\n" + "\n".join(items))
        docs.append({
            "text": text,
            "metadata": {
                "domain":  "pallet_box",
                "process": process,
                "source":  "파렛트, 박스 데이터",
            }
        })

    logger.info(f"  파렛트·박스 데이터: {len(docs)}건")
    return docs


def load_fare_jiksung(ws) -> List[Dict]:
    """
    직송 운임 시트 → 출발지+도착지+차종별 요금 청크
    컬럼: 출발지 / 도착지 / 1톤 / 2.5톤 / 3.5톤 / 5톤 / 8톤 / 11톤 / 18톤 / 25톤 / 트레일러 / Low
    """
    docs = []
    rows = list(ws.iter_rows(values_only=True))

    # 헤더 행 찾기 (출발지, 도착지 가 있는 행)
    header_row_idx = None
    for i, row in enumerate(rows):
        if row[0] and "출발지" in str(row[0]):
            header_row_idx = i
            break
    if header_row_idx is None:
        logger.warning("  직송 시트: 헤더 행 없음")
        return docs

    # 차종 컬럼 목록 (3번째 컬럼부터)
    header = rows[header_row_idx]
    ton_cols = []
    for col_idx, h in enumerate(header[2:], start=2):
        label = _clean(h)
        if label:
            ton_cols.append((col_idx, label))

    # 전체 요약 청크 (표 형식)
    summary_lines = [
        "[직송 운임표 - 전체]",
        "| 출발지 | 도착지 | " + " | ".join(t for _, t in ton_cols) + " |",
        "|--------|--------|" + "|".join(["------"] * len(ton_cols)) + "|",
    ]

    current_depart = ""
    for row in rows[header_row_idx + 1:]:
        if not row[1]:
            continue
        depart = _clean(row[0]) or current_depart
        if _clean(row[0]):
            current_depart = depart
        dest = _clean(row[1])
        if not dest:
            continue

        # 차종별 요금 dict
        fares = {}
        for col_idx, label in ton_cols:
            v = row[col_idx] if col_idx < len(row) else None
            fares[label] = _clean(v) if v is not None else ""

        # 요금 라인 (빈 값 제외)
        fare_lines = "\n".join(
            f"  - {label}: {val}원" for label, val in fares.items() if val
        )
        text = (
            f"[직송 운임 | {depart} → {dest}]\n"
            f"출발지: {depart} | 도착지: {dest}\n"
            f"운송 방식: 직송\n"
            f"차종별 요금:\n{fare_lines}"
        )
        docs.append({
            "text": text,
            "metadata": {
                "domain":      "fare",
                "fare_type":   "직송",
                "departure":   depart,
                "destination": dest,
                "fares":       fares,
                "source":      "운임_테이블.xlsx",
            }
        })

        # 요약 표 행 추가
        summary_lines.append(
            f"| {depart} | {dest} | " +
            " | ".join(fares.get(t, "") for _, t in ton_cols) + " |"
        )

    # 전체 요약 청크 삽입
    if docs:
        docs.insert(0, {
            "text": "\n".join(summary_lines),
            "metadata": {
                "domain":    "fare",
                "fare_type": "직송",
                "type":      "summary",
                "source":    "운임_테이블.xlsx",
            }
        })

    logger.info(f"  직송 운임: {len(docs)}건")
    return docs


def load_fare_hwamul(ws) -> List[Dict]:
    """
    화물·택배 운임 시트 → 출발지+도착지+화물+택배 청크
    컬럼: 출발지 / 도착지 / 화물(원) / 택배(원)
    """
    docs = []
    rows = list(ws.iter_rows(values_only=True))

    # 헤더 행 찾기
    header_row_idx = None
    for i, row in enumerate(rows):
        if row[0] and "출발지" in str(row[0]):
            header_row_idx = i
            break
    if header_row_idx is None:
        logger.warning("  화물·택배 시트: 헤더 행 없음")
        return docs

    # 전체 요약 청크
    summary_lines = [
        "[화물·택배 운임표 - 전체]",
        "| 출발지 | 도착지 | 화물(원) | 택배(원) |",
        "|--------|--------|---------|---------|",
    ]

    current_depart = ""
    for row in rows[header_row_idx + 1:]:
        if not row[1]:
            continue
        depart = _clean(row[0]) or current_depart
        if _clean(row[0]):
            current_depart = depart
        dest    = _clean(row[1])
        hwamul  = _clean(row[2]) if len(row) > 2 else ""
        taekbae = _clean(row[3]) if len(row) > 3 else ""

        if not dest:
            continue

        text = (
            f"[화물·택배 운임 | {depart} → {dest}]\n"
            f"출발지: {depart} | 도착지: {dest}\n"
            f"운송 방식: 화물·택배\n"
            f"  - 화물: {hwamul}원\n"
            f"  - 택배: {taekbae}원"
        )
        docs.append({
            "text": text,
            "metadata": {
                "domain":      "fare",
                "fare_type":   "화물택배",
                "departure":   depart,
                "destination": dest,
                "fare_hwamul": hwamul,
                "fare_taekbae": taekbae,
                "source":      "운임_테이블.xlsx",
            }
        })
        summary_lines.append(f"| {depart} | {dest} | {hwamul} | {taekbae} |")

    if docs:
        docs.insert(0, {
            "text": "\n".join(summary_lines),
            "metadata": {
                "domain":    "fare",
                "fare_type": "화물택배",
                "type":      "summary",
                "source":    "운임_테이블.xlsx",
            }
        })

    logger.info(f"  화물·택배 운임: {len(docs)}건")
    return docs


def load_fare_excel(excel_path: str) -> List[Dict]:
    """
    운임_테이블.xlsx 전체 로드 → 두 시트 합산 반환
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    docs = []

    sheet_loaders = {
        "직송":     load_fare_jiksung,
        "화물,택배": load_fare_hwamul,
    }
    for sheet_name, loader_fn in sheet_loaders.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f"  ⚠️  운임 시트 없음: {sheet_name} (실제 시트명 확인 필요)")
            logger.warning(f"     실제 시트 목록: {wb.sheetnames}")
            continue
        ws = wb[sheet_name]
        docs.extend(loader_fn(ws))

    return docs


# ══════════════════════════════════════════════════════════════
#  Qdrant 적재
# ══════════════════════════════════════════════════════════════

def build_qdrant_client() -> QdrantClient:
    return QdrantClient(
        url=f"http://{QDRANT_HOST}:{QDRANT_PORT}",
        api_key=QDRANT_API_KEY,
        timeout=60,
    )


def ensure_collection(client: QdrantClient, collection: str):
    """컬렉션 없으면 생성, 있으면 재사용"""
    existing = [c.name for c in client.get_collections().collections]
    if collection not in existing:
        client.create_collection(
            collection_name=collection,
            vectors_config=VectorParams(size=VECTOR_DIM, distance=Distance.COSINE),
        )
        # 메타데이터 인덱스 (도메인 필터 속도 향상)
        client.create_payload_index(collection, "domain",      PayloadSchemaType.KEYWORD)
        client.create_payload_index(collection, "route_group", PayloadSchemaType.KEYWORD)
        client.create_payload_index(collection, "driver",      PayloadSchemaType.KEYWORD)
        logger.info(f"✅ 컬렉션 생성: {collection}")
    else:
        logger.info(f"✅ 기존 컬렉션 사용: {collection}")


def embed_and_upsert(
    client: QdrantClient,
    embedder: OllamaEmbeddings,
    collection: str,
    docs: List[Dict],
):
    """배치 임베딩 후 Qdrant upsert"""
    total = len(docs)
    for start in range(0, total, BATCH_SIZE):
        batch = docs[start : start + BATCH_SIZE]
        texts = [d["text"] for d in batch]
        vectors = embedder.embed_documents(texts)

        points = [
            PointStruct(
                id=str(uuid.uuid4()),
                vector=vec,
                payload={**meta, "text": text},
            )
            for (text, meta, vec) in zip(texts, [d["metadata"] for d in batch], vectors)
        ]
        client.upsert(collection_name=collection, points=points)
        logger.info(f"  upsert {start + len(batch)}/{total}")


# ══════════════════════════════════════════════════════════════
#  메인 진입점
# ══════════════════════════════════════════════════════════════

SHEET_LOADERS = {
    "물류팀 운영 규칙":            load_operation_rules,
    "포장량 산출 데이터":           load_packaging_data,
    "수출 포장량 산출 수식":        load_export_formula,
    "차량 데이터":                  load_vehicle_data,
    "용차 차량 노선 데이터":        load_route_data,
    "지입 차량(기사) 노선 데이터":  load_driver_route,
    "컨베어벨트 규격 데이터":       load_conveyor_data,
    "컨베어벨트 직경 산출 수식":    load_conveyor_formula,
    "주름혹벨트 우든박스 사이즈 데이터": load_sidewall_data,
    "물류팀 현황 데이터":           load_personnel_data,
    "크롤러 러버트랙 규격 데이터":  load_crawler_data,
    "파렛트, 박스 데이터":          load_pallet_box_data,
}


def load_all(excel_path: str, reset_collection: bool = False,
             fare_excel_path: str = None):
    """
    전체 시트 로드 → 임베딩 → Qdrant 적재

    Args:
        excel_path: Logibot-Data_기본__V5.xlsx 경로
        reset_collection: True면 기존 컬렉션 삭제 후 재생성
        fare_excel_path: 운임_테이블.xlsx 경로 (없으면 건너뜀)
    """
    logger.info("=" * 60)
    logger.info(f"📂 파일 로드: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    # 임베딩 모델 초기화
    embedder = OllamaEmbeddings(model=OLLAMA_EMBEDDING_MODEL, base_url=OLLAMA_HOST)

    # Qdrant 연결
    client = build_qdrant_client()

    if reset_collection:
        existing = [c.name for c in client.get_collections().collections]
        if QDRANT_COLLECTION in existing:
            client.delete_collection(QDRANT_COLLECTION)
            logger.info(f"🗑️  기존 컬렉션 삭제: {QDRANT_COLLECTION}")

    ensure_collection(client, QDRANT_COLLECTION)

    # 시트별 처리
    total_docs = 0
    for sheet_name, loader_fn in SHEET_LOADERS.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f"⚠️  시트 없음: {sheet_name}")
            continue
        logger.info(f"\n📋 처리 중: [{sheet_name}]")
        ws   = wb[sheet_name]
        docs = loader_fn(ws)
        if not docs:
            logger.warning(f"  ⚠️  생성된 문서 없음")
            continue
        embed_and_upsert(client, embedder, QDRANT_COLLECTION, docs)
        total_docs += len(docs)

    # ── 운임 테이블 별도 파일 처리 ────────────────────────────
    if fare_excel_path:
        logger.info(f"\n📋 처리 중: [운임_테이블.xlsx]")
        try:
            fare_docs = load_fare_excel(fare_excel_path)
            if fare_docs:
                embed_and_upsert(client, embedder, QDRANT_COLLECTION, fare_docs)
                total_docs += len(fare_docs)
                logger.info(f"  운임 테이블: {len(fare_docs)}건 적재 완료")
            else:
                logger.warning("  ⚠️  운임 테이블 생성된 문서 없음")
        except Exception as e:
            logger.error(f"  ❌ 운임 테이블 적재 실패: {e}")
    else:
        logger.info("\n⏭️  운임 테이블 경로 미지정 → 건너뜀 (--fare 옵션으로 지정 가능)")

    logger.info("=" * 60)
    logger.info(f"✅ 적재 완료 | 총 {total_docs}개 문서 → [{QDRANT_COLLECTION}]")
    return total_docs


# ══════════════════════════════════════════════════════════════
#  query_processor.py 에서 import해서 쓰는 헬퍼
# ══════════════════════════════════════════════════════════════

def get_driver_schedule_filter(target: date = None):
    """
    query_processor.py의 driver_route 검색 시
    오늘 날짜 기준 route_group 필터를 반환.
    
    사용 예시 (query_processor.py 내):
        from data_loader import get_driver_schedule_filter
        from qdrant_client.models import Filter, FieldCondition, MatchAny
        
        route_filter = get_driver_schedule_filter()
        results = qdrant_client.search(
            collection_name=QDRANT_COLLECTION,
            query_vector=query_vec,
            query_filter=route_filter,
            limit=10,
        )
    """
    from qdrant_client.models import Filter, FieldCondition, MatchAny
    wg = get_week_group(target)
    return Filter(
        must=[
            FieldCondition(
                key="route_group",
                match=MatchAny(any=["공통", wg])
            )
        ]
    )


if __name__ == "__main__":
    import sys

    args = sys.argv[1:]
    reset = "--reset" in args

    # 운임 파일 경로 우선순위:
    #   1) --fare <경로> 커맨드라인 인자
    #   2) .env의 LOGIBOT_FARE_PATH
    #   3) 기본값 data/source_docs/운임_테이블.xlsx
    if "--fare" in args:
        fare_idx = args.index("--fare")
        fare_path = args[fare_idx + 1] if fare_idx + 1 < len(args) else None
    else:
        fare_path = os.getenv(
            "LOGIBOT_FARE_PATH",
            os.path.join("data", "source_docs", "운임_테이블.xlsx")
        )

    if fare_path and not os.path.exists(fare_path):
        logger.warning(f"⚠️  운임 테이블 파일 없음: {fare_path}")
        logger.warning("   운임 데이터는 건너뜁니다. 파일이 있다면 .env에 LOGIBOT_FARE_PATH 설정")
        fare_path = None

    # 엑셀 경로 우선순위:
    #   1) 커맨드라인 인자 (python data_loader.py path/to/file.xlsx --reset)
    #   2) .env의 LOGIBOT_EXCEL_PATH
    #   3) 기본값 data/source_docs/Logibot-Data(기본)_V5.xlsx
    non_option_args = [a for a in args if not a.startswith("--")]
    if non_option_args:
        path = non_option_args[0]
    else:
        path = os.getenv(
            "LOGIBOT_EXCEL_PATH",
            os.path.join("data", "source_docs", "Logibot-Data(기본)_V5.xlsx")
        )

    if not os.path.exists(path):
        logger.error(f"❌ 엑셀 파일을 찾을 수 없습니다: {path}")
        logger.error("   해결 방법:")
        logger.error("   1) .env 파일에 LOGIBOT_EXCEL_PATH=<경로> 설정")
        logger.error("   2) 또는 직접 경로 지정: python data_loader.py <경로> --reset")
        sys.exit(1)

    load_all(path, reset_collection=reset, fare_excel_path=fare_path)